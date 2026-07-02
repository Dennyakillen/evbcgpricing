#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dry_run_e2e.py -- staged end-to-end dry run: FORE / MOTOR / EFTER (conductor, not a new probe)
==============================================================================================
Purpose
    ONE command that answers "can I safely launch a large-scale run across all
    three families?" -- split exactly as the pipeline is split:

      GATE   subscription + token + Blob data-plane role probe (E.3 / LB.46)
      FORE   data prep contracts: composes dry_run_pipeline (19 checks),
             dry_run_full_pipeline (local path contracts), window_coherence
             (window consistency), parquet freshness vs window end (LB.50),
             leftover-artifact scan (*.SURVIVALTEST, stray .bak)
      MOTOR  LIVE Azure/VM layer (the genuinely NEW part): start VM, SSH
             reachability, pipeline venv packages (ray/pandas/statsmodels),
             per-family code dirs, CONFIG DRIFT repo-vs-VM (md5 -- the exact
             drift class that blocked cluster-maj), disk headroom.
             Always deallocates in `finally` unless --keep-vm, then VERIFIES
             power state via get-instance-view (LB.60) -- a left-running VM
             is reported as a LEAK, not silently ignored.
      EFTER  local after-chain preconditions: blob.download_outputs/upload_
             outputs present, frozen placements on disk (run_step6 PLACEMENTS),
             xlwings importable (step 5/7 Excel COM is Windows-local, LB.44).

Design decision (VAKTEN trigger 1 -- why this is NOT probe generation five):
    This file contains NO new validation logic for FORE/EFTER -- it runs the
    EXISTING, proven tools as subprocesses and aggregates their exit codes.
    Role relationship: all_chain_validator = static chain derivation;
    dry_run_pipeline = pipe checks; THIS = staged conductor + live MOTOR layer
    + leak checks. Feeds BB.1 (probe consolidation) instead of fighting it.

Output
    Console: structural [STAGE]/[PASS]/[FAIL]/[WARN]/[SKIP]/[LEAK] lines.
    Excel receipt (openpyxl): workspace/validation_receipts/
    dry_run_e2e_<UTC timestamp>.xlsx -- columns Stage | Check | Status | Details.

Usage
    py -3.11 verify_tool\\dry_run_e2e.py                          # all stages, VM cycled
    py -3.11 verify_tool\\dry_run_e2e.py --stage fore             # offline-ish (token for 19-check part)
    py -3.11 verify_tool\\dry_run_e2e.py --skip-vm                # everything except live MOTOR
    py -3.11 verify_tool\\dry_run_e2e.py --keep-vm                # leave VM running (pre-launch mode)
    py -3.11 verify_tool\\dry_run_e2e.py --window 2022-07-01_2026-05-31
    py -3.11 verify_tool\\dry_run_e2e.py --plan                   # print check matrix, execute nothing

    Default window end = resolve_window_end() from orchestration/shared/window.py
    (single owner of the growing-window default -- reuse, never re-declare).

Lessons honored: E.3 (4h token), LB.46 (subscription), LB.50 (parquet-first),
    LB.60 (verify power state, don't trust the log line), LB.44 (xlwings local),
    SSH quoting discipline (single quotes, NO inner quotes -- pip show, ls, md5sum
    are chosen precisely because they need no inner quoting), R7 (exit codes lie:
    every subprocess's tail is captured into the receipt).

Developer : Jens Palmo (Senior Business Analyst, Evidensia). Author: Claude advisor.
Created   : 2026-07-02
"""
from __future__ import annotations

import argparse
import hashlib
import os
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

# ----------------------------------------------------------------------------- constants
REPO       = Path(os.environ.get("BCG_REPO", r"C:\Projekt\BCG"))
BA_ROOT    = Path(os.environ.get("BA_ROOT", r"C:\Projekt\Business_Analytics"))
RECEIPT_DIR = Path(os.environ.get("DRY_RUN_E2E_RECEIPT_DIR",
                                  str(REPO / "workspace" / "validation_receipts")))

SUBSCRIPTION = "ev-lz3-ai (SE)"
VM_RG, VM_NAME, VM_IP, VM_USER = "ev-openai-swce-rg-test", "bcg-poc-vm", "172.18.148.4", "azureuser"
STORAGE_ACCOUNT, STATUS_CONTAINER = "evbcgpricinginput", "runstatus"

FAMILIES = {
    "cluster": "~/bcg/cluster/code/src/config.yml",
    "site":    "~/bcg/site/code/src/config.yml",
    "bundle":  "~/bcg/bundle/code/src/config.yml",
}
LOCAL_CONFIGS = {
    "cluster": REPO / "Pipeline" / "02. Elasticity" / "2. Product Cluster Level Models" / "code" / "src" / "config.yml",
    "site":    REPO / "Pipeline" / "02. Elasticity" / "3. Product Site Level Models" / "code" / "src" / "config.yml",
    "bundle":  REPO / "Pipeline" / "02. Elasticity" / "5. Bundle Clinic Models" / "code" / "src" / "config.yml",
}
VENV_PIP = "~/bcg/cluster/.venv/bin/pip"   # site/bundle share cluster's venv

FORE_TOOLS = [
    ("dry_run_pipeline (19 pipe checks)",  REPO / "orchestration" / "dry_run_pipeline.py",      []),
    ("dry_run_full_pipeline (local paths)", REPO / "verify_tool" / "dry_run_full_pipeline.py",  []),
    ("window_coherence",                    REPO / "verify_tool" / "window_coherence.py",       ["--end", "{end}"]),
]
PARQUET = BA_ROOT / "parquet" / "transaction_data.parquet"

RESULTS: "list[tuple[str,str,str,str]]" = []   # (stage, check, status, details)


# ----------------------------------------------------------------------------- helpers
def log(stage: str, status: str, check: str, details: str = "") -> None:
    line = f"[{stage:<5}] [{status:<4}] {check}"
    if details:
        line += f"  -- {details}"
    print(line, flush=True)
    RESULTS.append((stage, check, status, details[:900]))


def sh(cmd: str, timeout: int = 90) -> "tuple[int,str]":
    """Run a shell command string (az/ssh live in PATH as launchers on Windows).
    Single-quote discipline for ssh remote parts: no inner quotes ever."""
    try:
        p = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=timeout)
        out = (p.stdout or "") + (p.stderr or "")
        return p.returncode, out.strip()
    except subprocess.TimeoutExpired:
        return 124, f"TIMEOUT {timeout}s"


def run_tool(label: str, script: Path, args: "list[str]", stage: str, timeout: int = 420) -> None:
    if not script.exists():
        log(stage, "FAIL", label, f"script missing: {script}")
        return
    cmd = [sys.executable, str(script)] + args
    try:
        p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, cwd=str(REPO))
        tail = " | ".join([l for l in (p.stdout + p.stderr).splitlines() if l.strip()][-4:])
        log(stage, "PASS" if p.returncode == 0 else "FAIL", label, f"rc={p.returncode}; {tail}")
    except subprocess.TimeoutExpired:
        log(stage, "FAIL", label, f"TIMEOUT {timeout}s")


def md5_local(p: Path) -> str:
    return hashlib.md5(p.read_bytes()).hexdigest() if p.exists() else "MISSING"


# ----------------------------------------------------------------------------- stages
def stage_gate() -> bool:
    """Subscription + token + data-plane role. Returns False only on hard stop."""
    rc, out = sh("az account show --query name -o tsv")
    ok = (rc == 0 and SUBSCRIPTION in out)
    log("GATE", "PASS" if ok else "FAIL", f"subscription == {SUBSCRIPTION}", out.splitlines()[0] if out else "")
    if not ok:
        log("GATE", "FAIL", "hard stop", "wrong/absent subscription (LB.46) -- az account set / az login first")
        return False

    rc, out = sh("az account get-access-token --query expiresOn -o tsv")
    log("GATE", "PASS" if rc == 0 else "FAIL", "token issued (E.3: renew every 4h; re-login AFTER PIM)", out)

    rc, out = sh(f"az storage blob list --account-name {STORAGE_ACCOUNT} "
                 f"--container-name {STATUS_CONTAINER} --auth-mode login "
                 f"--num-results 3 --query [].name -o tsv", timeout=60)
    if rc == 0:
        log("GATE", "PASS", "Blob DATA-PLANE role (AAD login-mode list)", out.replace("\n", ", ") or "(container empty)")
    elif "AuthorizationPermissionMismatch" in out:
        log("GATE", "WARN", "Blob DATA-PLANE role", "AuthorizationPermissionMismatch -- role not granted/propagated "
                                                    "(FAS T / Kent). Blob-dependent checks degrade; key-mode still works.")
    else:
        log("GATE", "WARN", "Blob DATA-PLANE role", out.splitlines()[-1] if out else f"rc={rc}")
    return True


def stage_fore(window_end: str) -> None:
    for label, script, args in FORE_TOOLS:
        run_tool(label, script, [a.format(end=window_end) for a in args], "FORE")

    # parquet freshness vs window end (LB.50: regenerate parquet FIRST)
    if PARQUET.exists():
        try:
            import pandas as pd  # type: ignore
            mx = pd.read_parquet(PARQUET, columns=["week_starting_monday"])["week_starting_monday"].max()
            fresh = str(mx)[:10] >= window_end[:8] + "01"   # at least into the end month
            log("FORE", "PASS" if fresh else "FAIL", "parquet freshness vs window end",
                f"max(week)={str(mx)[:10]}, window end={window_end}"
                + ("" if fresh else " -- regenerate parquet FIRST (LB.50)"))
        except Exception as e:  # pragma: no cover
            log("FORE", "WARN", "parquet freshness", f"could not read ({type(e).__name__}: {e}) -- verify manually")
    else:
        log("FORE", "FAIL", "parquet exists", str(PARQUET))

    # leftover artifacts (leak class on the FILE side)
    left = list(REPO.rglob("*.SURVIVALTEST")) + list(BA_ROOT.rglob("*.SURVIVALTEST"))
    log("FORE", "PASS" if not left else "FAIL", "no *.SURVIVALTEST leftovers",
        "; ".join(str(p) for p in left[:5]) or "clean")
    baks = [p for p in REPO.rglob("*.bak-*") if "archives" not in p.parts]
    log("FORE", "PASS" if not baks else "WARN", "stray .bak files outside archives",
        f"{len(baks)} found" + (f" e.g. {baks[0].name}" if baks else ""))


def stage_motor(keep_vm: bool) -> None:
    vm_started_here = False
    try:
        rc, out = sh(f"az vm start --resource-group {VM_RG} --name {VM_NAME}", timeout=300)
        log("MOTOR", "PASS" if rc == 0 else "FAIL", "az vm start", out.splitlines()[-1] if out else "started")
        vm_started_here = (rc == 0)
        if rc != 0:
            return

        # SSH reachability with patience (boot + sshd)
        ok = False
        for i in range(12):
            rc, out = sh(f"ssh -o ConnectTimeout=8 -o StrictHostKeyChecking=accept-new "
                         f"{VM_USER}@{VM_IP} 'echo SSH_OK'", timeout=30)
            if rc == 0 and "SSH_OK" in out:
                ok = True
                break
            time.sleep(10)
        log("MOTOR", "PASS" if ok else "FAIL", "SSH reachable", f"attempts={i+1}")
        if not ok:
            return

        # pipeline venv packages (pip show needs NO inner quotes -- by design)
        for pkg in ("ray", "pandas", "statsmodels"):
            rc, out = sh(f"ssh {VM_USER}@{VM_IP} '{VENV_PIP} show {pkg}'", timeout=45)
            ver = next((l.split(":", 1)[1].strip() for l in out.splitlines() if l.startswith("Version")), "?")
            log("MOTOR", "PASS" if rc == 0 else "FAIL", f"venv package {pkg}", f"version={ver}")

        # per-family code dirs + CONFIG DRIFT repo vs VM (the cluster-maj killer class)
        for fam, remote_cfg in FAMILIES.items():
            rc, out = sh(f"ssh {VM_USER}@{VM_IP} 'md5sum {remote_cfg}'", timeout=30)
            if rc != 0:
                log("MOTOR", "FAIL", f"{fam}: config on VM", out.splitlines()[-1] if out else "missing")
                continue
            remote_md5 = out.split()[0]
            local_md5 = md5_local(LOCAL_CONFIGS[fam])
            same = (remote_md5 == local_md5)
            log("MOTOR", "PASS" if same else "FAIL", f"{fam}: config drift repo vs VM",
                f"local={local_md5[:10]} vm={remote_md5[:10]}"
                + ("" if same else " -- scp the repo config before launching (drift = maj-crash class)"))

        rc, out = sh(f"ssh {VM_USER}@{VM_IP} 'df -h /home'", timeout=30)
        use = next((l.split()[4] for l in out.splitlines() if "/home" in l or l.endswith("/")), "?")
        pct = int(use.rstrip("%")) if use.rstrip("%").isdigit() else 0
        log("MOTOR", "PASS" if pct < 85 else "WARN", "disk headroom /home", f"use={use}")

    finally:
        if keep_vm:
            log("MOTOR", "WARN", "VM left RUNNING by --keep-vm",
                "~9 kr/h -- deallocate after launch: az vm deallocate ...")
        elif vm_started_here:
            sh(f"az vm deallocate --resource-group {VM_RG} --name {VM_NAME}", timeout=300)
            rc, out = sh(f"az vm get-instance-view --resource-group {VM_RG} --name {VM_NAME} "
                         f"--query instanceView.statuses[1].displayStatus -o tsv", timeout=60)
            deallocated = ("deallocated" in out.lower())
            log("MOTOR", "PASS" if deallocated else "LEAK", "VM deallocated + VERIFIED (LB.60)", out)


def stage_efter() -> None:
    # blob helpers present (PULL/PUSH for run_after)
    sys.path.insert(0, str(REPO / "orchestration" / "infrastructure"))
    try:
        import blob  # type: ignore
        for fn in ("download_outputs", "upload_outputs"):
            log("EFTER", "PASS" if hasattr(blob, fn) else "FAIL", f"blob.{fn} exists", "")
    except Exception as e:
        log("EFTER", "WARN", "import blob", f"{type(e).__name__}: {e} (azure sdk env?)")

    # frozen placements from run_step6.PLACEMENTS (derive, don't re-declare paths here)
    sys.path.insert(0, str(REPO / "verify_tool" / "run"))
    try:
        import run_step6  # type: ignore
        placements = getattr(run_step6, "PLACEMENTS", {})
        frozen = [(k, v) for k, v in placements.items()] if isinstance(placements, dict) else []
        if not frozen:
            log("EFTER", "WARN", "run_step6.PLACEMENTS", "not found/empty -- verify manually")
        for key, src in frozen:
            p = Path(str(src))
            log("EFTER", "PASS" if p.exists() else "FAIL", f"placement source: {key}", str(p))
    except Exception as e:
        log("EFTER", "WARN", "import run_step6", f"{type(e).__name__}: {e}")

    # Excel-COM leg (step 5/7 is Windows-local, LB.44)
    try:
        import xlwings  # type: ignore  # noqa: F401
        log("EFTER", "PASS", "xlwings importable (step 5/7 Excel COM)", "")
    except Exception as e:
        log("EFTER", "FAIL", "xlwings importable", f"{type(e).__name__}: {e}")


# ----------------------------------------------------------------------------- receipt
def write_receipt(window: str, stages: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    path = RECEIPT_DIR / f"dry_run_e2e_{ts}.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "dry_run_e2e"
    meta = [("Receipt", "dry_run_e2e staged FORE/MOTOR/EFTER"),
            ("Generated (UTC)", ts), ("Repo", str(REPO)),
            ("Window", window), ("Stages", stages),
            ("Developer", "Jens Palmo (Senior Business Analyst, Evidensia)")]
    for k, v in meta:
        ws.append([k, v]); ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append([])
    ws.append(["Stage", "Check", "Status", "Details"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for row in RESULTS:
        ws.append(list(row))
    for col, w in zip("ABCD", (8, 52, 8, 100)):
        ws.column_dimensions[col].width = w
    wb.save(path)
    return path


# ----------------------------------------------------------------------------- main
def main() -> int:
    ap = argparse.ArgumentParser(description="Staged end-to-end dry run: FORE/MOTOR/EFTER.")
    ap.add_argument("--stage", choices=["all", "fore", "motor", "efter"], default="all")
    ap.add_argument("--window", default=None, help="run window id, e.g. 2022-07-01_2026-05-31")
    ap.add_argument("--skip-vm", action="store_true", help="skip live MOTOR stage")
    ap.add_argument("--keep-vm", action="store_true", help="leave VM running after MOTOR (pre-launch)")
    ap.add_argument("--plan", action="store_true", help="print the check matrix, execute nothing")
    args = ap.parse_args()

    if args.window:
        window = args.window
        end = window.split("_")[-1]
    else:
        sys.path.insert(0, str(REPO / "orchestration" / "shared"))
        try:
            from window import resolve_window_end, WINDOW_ANCHOR  # type: ignore
            end = resolve_window_end(); window = f"{WINDOW_ANCHOR}_{end}"
        except Exception:
            end = "9999-12-31"; window = f"2022-07-01_{end}"
            log("GATE", "WARN", "window.py not importable", "pass --window explicitly")

    print(f"=== dry_run_e2e | window {window} | stage {args.stage}"
          f"{' | PLAN ONLY' if args.plan else ''} ===", flush=True)

    if args.plan:
        matrix = {
            "GATE":  ["subscription", "token", "Blob data-plane role (login-mode)"],
            "FORE":  [t[0] for t in FORE_TOOLS] + ["parquet freshness (LB.50)",
                                                    "*.SURVIVALTEST leftovers", "stray .bak"],
            "MOTOR": ["vm start", "ssh reachable", "venv: ray/pandas/statsmodels",
                      "config drift repo-vs-VM x3 families", "disk /home",
                      "deallocate + LB.60 verify (finally)"],
            "EFTER": ["blob.download/upload_outputs", "run_step6.PLACEMENTS on disk",
                      "xlwings (LB.44)"],
        }
        for st, checks in matrix.items():
            for c in checks:
                log(st, "PLAN", c)
        p = write_receipt(window, f"{args.stage} (plan)")
        print(f"[RECEIPT] {p}")
        return 0

    if not stage_gate():
        write_receipt(window, "gate (hard stop)")
        return 2
    if args.stage in ("all", "fore"):
        stage_fore(end)
    if args.stage in ("all", "motor") and not args.skip_vm:
        stage_motor(args.keep_vm)
    elif args.stage in ("all", "motor"):
        log("MOTOR", "SKIP", "live VM stage skipped by --skip-vm")
    if args.stage in ("all", "efter"):
        stage_efter()

    fails = [r for r in RESULTS if r[2] == "FAIL"]
    leaks = [r for r in RESULTS if r[2] == "LEAK"]
    p = write_receipt(window, args.stage)
    print(f"\n=== SUMMARY: PASS={sum(1 for r in RESULTS if r[2]=='PASS')} "
          f"FAIL={len(fails)} WARN={sum(1 for r in RESULTS if r[2]=='WARN')} "
          f"LEAK={len(leaks)} ===")
    print(f"[RECEIPT] {p}")
    return 1 if (fails or leaks) else 0


if __name__ == "__main__":
    sys.exit(main())
