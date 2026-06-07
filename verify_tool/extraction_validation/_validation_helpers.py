"""
_validation_helpers.py
=======================
Shared helpers for the validation suite: stdout-capturing Excel receipt
writer, formatting, hashing, and path resolution.

Developer: Jens Palmö (Senior Business Analyst, Evidensia Djursjukvård AB)
Created:   2026-06-07
Updated:   2026-06-07 (single-sheet "Logg" receipt format, matching verify_tool style)

RECEIPT FORMAT (matches existing verify_tool receipts from 28 May):
  - Single sheet named "Logg"
  - Row 1: title with timestamp
  - Row 2: interpreter / context
  - Row 3: developer
  - Row 4: header explaining the raw log preservation
  - Row 5: separator line
  - Row 6+: raw stdout verbatim, monospace, column alignment preserved

Used by all validate_*.py scripts in C:\Projekt\BCG\verify_tool\extraction_validation\.
"""
import hashlib
import io
import sys
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# ============================ STANDARD PATHS ============================
BCG_ROOT = Path(r"C:\Projekt\BCG")
BUSINESS_ROOT = Path(r"C:\Projekt\Business_Analytics")

# Cluster model paths
CLUSTER_MODEL_ROOT = BCG_ROOT / "Pipeline" / "02. Elasticity" / "2. Product Cluster Level Models"
OUR_CSV = CLUSTER_MODEL_ROOT / "data" / "0828_Sweden_weekly_model_data_P_C.csv"

# BCG facit paths
BCG_FACIT_CSV = BUSINESS_ROOT / "bcg_inputs" / "0828_Sweden_weekly_model_data_P_C.csv"
BCG_CLUSTER_SEED_XLSX = BUSINESS_ROOT / "bcg_inputs" / "0808_Sweden_Clinic_Cluster_Mapping.xlsx"
BCG_FTE_XLSX = BUSINESS_ROOT / "bcg_inputs" / "Sweden__Interpolated_Productivity_time_date_june25.xlsx"

# Receipt directory under verify_tool/receipts/
RECEIPT_ROOT = BCG_ROOT / "verify_tool" / "receipts"

# Standard windows
BCG_START = "2022-07-01"
BCG_END = "2025-06-28"


# ============================ FORMATTING ============================
def fmt_msek(amount):
    """Format amount as million SEK with 1 decimal."""
    return f"{amount/1e6:>10.1f} MSEK"


def fmt_pct(pct, decimals=2):
    """Format percentage with sign and decimals."""
    return f"{pct:>+7.{decimals}f}%"


def fmt_int(n):
    """Format integer with thousand separators."""
    return f"{n:>10,}"


def file_hash_short(path):
    """Compute short MD5 hash of a file for audit trail."""
    try:
        h = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()[:12]
    except Exception:
        return "n/a"


def now_iso():
    """Current timestamp in ISO format."""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def now_file_stamp():
    """Current timestamp suitable for file names."""
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")


def get_receipt_dir():
    """Get the dated receipt subfolder, creating it if needed."""
    today = datetime.now().strftime("%Y-%m-%d")
    dir_path = RECEIPT_ROOT / today
    dir_path.mkdir(parents=True, exist_ok=True)
    return dir_path


# ============================ CONSOLE LOGGING ============================
def section(title, char="="):
    """Print a section header."""
    line = char * 78
    print(line)
    print(title)
    print(line)


def subsection(title):
    """Print a subsection header."""
    print()
    print("-" * 78)
    print(title)
    print("-" * 78)


# ============================ STDOUT CAPTURE ============================
class _Tee:
    """File-like that writes to both real stdout AND a StringIO buffer."""

    def __init__(self, real_stdout, buffer):
        self._real = real_stdout
        self._buffer = buffer

    def write(self, data):
        self._real.write(data)
        self._buffer.write(data)

    def flush(self):
        self._real.flush()


@contextmanager
def capture_stdout():
    """Context manager that captures stdout while still printing to console.

    Usage:
        with capture_stdout() as buf:
            print("hello")
        log_text = buf.getvalue()
    """
    buf = io.StringIO()
    original = sys.stdout
    sys.stdout = _Tee(original, buf)
    try:
        yield buf
    finally:
        sys.stdout = original


# ============================ RECEIPT WRITER ============================
# XML 1.0 illegal chars - everything below 0x20 except \t, \n; plus 0x7F-0x9F
# (openpyxl uses XML 1.0 which is stricter than Python strings)
import re as _re
_ILLEGAL_XML_CHARS_RE = _re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x84\x86-\x9F\uFDD0-\uFDEF"
    r"\uFFFE\uFFFF]"
)


def _sanitize_for_excel(text):
    """Strip XML-illegal characters that openpyxl rejects.

    Replaces NULL bytes, bell, vertical tab, form feed, and other control
    chars with a visible placeholder. Preserves \\t, \\n, and printable text.
    """
    if text is None:
        return ""
    # Remove \r (Excel doesn't like raw CR inside cells); keep \n via row split
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return _ILLEGAL_XML_CHARS_RE.sub("", text)


def write_log_receipt(receipt_path, script_name, log_text, interpreter="py -3.11"):
    """
    Write a single-sheet "Logg" Excel receipt matching the verify_tool format.

    Args:
        receipt_path: Path where Excel file is saved.
        script_name: Identifier for the title row (e.g. "validate_extraction_coverage.py").
        log_text: Raw stdout text to preserve verbatim.
        interpreter: Interpreter identifier for row 2 (default "py -3.11").
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Logg"

    timestamp = now_iso()

    # Sanitize text (strip XML-illegal chars, normalize CRLF)
    log_text = _sanitize_for_excel(log_text)

    # Header rows (matches 28 May receipt format)
    ws.append([f"{script_name} - receipt {timestamp}"])
    ws.append([f"interpreter: {interpreter}"])
    ws.append([f"Developer: Jens Palmö (Senior Business Analyst, Evidensia Djursjukvård AB)"])
    ws.append([f"Raw stdout below, verbatim. Monospace preserves the column alignment."])
    ws.append(["=" * 78])
    ws.append([""])

    # Title rows styled (use generally-available fonts)
    ws["A1"].font = Font(bold=True, size=12, color="2F5496")
    for r in [2, 3, 4]:
        ws[f"A{r}"].font = Font(italic=True, color="666666")

    # Raw log lines, monospace ("Courier New" is universally installed; fallback "Consolas")
    mono_font = Font(name="Courier New", size=10)
    for line in log_text.splitlines():
        # Ensure value is a plain string and not interpreted as formula/etc.
        # Prepending with empty string is safe; openpyxl will treat as text.
        cell_value = str(line) if line else ""
        ws.append([cell_value])
        cell = ws[f"A{ws.max_row}"]
        cell.font = mono_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        # Force text type (avoid Excel inferring as formula/number/date)
        cell.number_format = "@"

    # Wide column for readability
    ws.column_dimensions["A"].width = 150

    wb.save(receipt_path)
    return receipt_path


# ============================ BACKWARDS COMPAT ============================
# Old multi-sheet writer kept as no-op for any script that still calls it.
# New scripts should use capture_stdout() + write_log_receipt() instead.
def write_receipt(receipt_path, title, sheets):
    """Legacy multi-sheet writer - replaced by single-sheet "Logg" format.

    Kept for backwards compatibility but emits a deprecation note in the log.
    """
    log_lines = [f"[DEPRECATED] write_receipt() was called for: {title}"]
    log_lines.append("Use capture_stdout() + write_log_receipt() instead.")
    log_lines.append("")
    log_lines.append("Sheets that would have been created:")
    for sheet in sheets:
        log_lines.append(f"  - {sheet.get('name', 'unknown')}: "
                         f"{len(sheet.get('rows', []))} rows")
    log_text = "\n".join(log_lines)
    return write_log_receipt(receipt_path, "legacy_call", log_text)
