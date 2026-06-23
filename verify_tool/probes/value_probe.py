#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
value_probe.py  --  VARDESOND: spara ett specifikt vardes ALLA forekomster i koden
===================================================================================
Foljer upp Lager 1:s INKONSEKVENS-fynd. Nar strukturkvittot sager "variabel X har
OLIKA default pa olika stallen" svarar denna sond: EXAKT vilken fil/rad har vilket
varde, sa du kan ratta till ETT. Ren statisk analys -- ingen Blob/VM/DW/PIM.

Forsta drivande fall (ur Lager 1, 2026-06-22): BCG_END_DATE har default '2025-06-28'
PA vissa stallen och '2025-06-29' pa andra (en dags glapp -> familjer kan inkludera
olika manga dagar -> tyst inkonsekvens i vaven). Denna sond visar exakt var, sa glappet
kan stangas.

Generell: ge --term (en variabel ELLER ett godtyckligt varde/strang) sa listar sonden
varje fil + radnummer + den fulla raden dar termen forekommer, grupperat per VARDE den
tilldelas (for env-variabler) eller per forekomst (for fri strang).

KOR (PowerShell, global py -3.11, ingen extern access):
    py -3.11 value_probe.py --root "C:\\Projekt\\BCG" --also-ba "C:\\Projekt\\Business_Analytics" --term BCG_END_DATE
    py -3.11 value_probe.py --root "C:\\Projekt\\BCG" --term evipricingmodelstprod   # spara ett kontonamn
    py -3.11 value_probe.py --root "C:\\Projekt\\BCG" --term 2026-04-30              # spara ett datum
    # valfritt --out <mapp>; valfritt --no-excel (bara konsol)

Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia), assisterad av Claude.
Beroende: openpyxl (global 3.11) for kvitto; annars std-lib (re, pathlib).
"""
from __future__ import annotations
import argparse
import datetime
import re
import sys
from pathlib import Path


def log(m): print(m, flush=True)


def safe_read_lines(p: Path):
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return p.read_text(encoding=enc).splitlines()
        except Exception:
            continue
    return []


def iter_files(root: Path, also_ba):
    skip = {".venv", "site-packages", "__pycache__", ".git"}
    roots = [root]
    if also_ba:
        ba = Path(also_ba)
        if ba.is_dir(): roots.append(ba)
    for r in roots:
        for p in r.rglob("*.py"):
            if not (set(p.parts) & skip): yield p
        for p in r.rglob("*.sql"):
            if not (set(p.parts) & skip): yield p


def main() -> int:
    ap = argparse.ArgumentParser(description="Vardesond: spara ett vardes alla forekomster (statisk).")
    ap.add_argument("--root", default=r"C:\Projekt\BCG")
    ap.add_argument("--also-ba", default=None)
    ap.add_argument("--term", required=True, help="Variabelnamn (t.ex. BCG_END_DATE) ELLER ett varde/strang.")
    ap.add_argument("--out", default=None)
    ap.add_argument("--no-excel", action="store_true")
    args = ap.parse_args()

    root = Path(args.root)
    if not root.is_dir():
        log(f"[ERROR] root finns ej: {root}"); return 2

    term = args.term
    files = list(iter_files(root, args.also_ba))
    log(f"[RUN] spar '{term}' i {len(files)} filer under {root}"
        + (f" + {args.also_ba}" if args.also_ba else ""))

    # Om termen ser ut som ett ENV-NAMN (versaler/understreck), forsok extrahera TILLDELAT varde.
    looks_env = bool(re.fullmatch(r"[A-Z][A-Z0-9_]+", term))
    # Monster for att fanga default-varde nar env-namnet forekommer:
    #   os.environ.get("NAME", "VALUE")  /  os.environ.get('NAME', 'VALUE')
    #   NAME = os.environ.get(..., "VALUE")  -- vi tar VALUE ur samma rad
    env_default = re.compile(
        re.escape(term) + r"""["']?\s*,\s*["']([^"']*)["']""")
    # Generell forekomst
    hits = []          # (fil_rel, radnr, rad, extraherat_varde_or_None)
    for p in files:
        lines = safe_read_lines(p)
        for i, line in enumerate(lines, 1):
            if term in line:
                val = None
                if looks_env:
                    m = env_default.search(line)
                    if m: val = m.group(1)
                hits.append((str(p.relative_to(root)) if p.is_relative_to(root) else str(p),
                             i, line.strip(), val))

    if not hits:
        log(f"[KPI] inga forekomster av '{term}'."); 
        log("[DONE]"); return 0

    # Gruppera per extraherat varde (for env) -> visar glappet direkt
    by_val = {}
    for f, n, line, val in hits:
        key = val if val is not None else "(ingen default pa raden)"
        by_val.setdefault(key, []).append((f, n, line))

    log(f"\n[KPI] {len(hits)} forekomster"
        + (f", grupperade per default-varde:" if looks_env else ":"))
    if looks_env and len([k for k in by_val if k != "(ingen default pa raden)"]) > 1:
        log(f"  ⚠ OLIKA varden funna: {sorted(k for k in by_val if k != '(ingen default pa raden)')}")
    for val, occ in sorted(by_val.items()):
        log(f"\n  VARDE: {val}   ({len(occ)} forekomst(er))")
        for f, n, line in occ:
            log(f"    {f}:{n}")
            log(f"        {line[:100]}")

    if not args.no_excel:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
            safe_term = re.sub(r"[^A-Za-z0-9_.-]", "_", term)[:40]
            out_dir = Path(args.out) if args.out else (
                root / "verify_tool" / "receipts" / datetime.date.today().isoformat() / "value")
            out_dir.mkdir(parents=True, exist_ok=True)
            fp = out_dir / f"value_{safe_term}_{stamp}.xlsx"
            wb = Workbook(); ws = wb.active; ws.title = "Vardesond"
            mono = Font(name="Consolas", size=10); bold = Font(name="Consolas", size=10, bold=True)
            warn = PatternFill("solid", fgColor="F8D7DA")
            ws.append([f"VARDESOND: '{term}'   {stamp}"]); ws["A1"].font = bold
            ws.append([f"Root: {root}"]); 
            multi = looks_env and len([k for k in by_val if k != "(ingen default pa raden)"]) > 1
            if multi:
                ws.append([f"VARNING: OLIKA varden funna -> {sorted(k for k in by_val if k != '(ingen default pa raden)')}"])
                ws.cell(ws.max_row, 1).fill = warn; ws.cell(ws.max_row,1).font = bold
            ws.append([]); ws.append(["VARDE", "FIL", "RAD", "KOD"])
            for c in range(1,5): ws.cell(ws.max_row, c).font = bold
            for val, occ in sorted(by_val.items()):
                for f, n, line in occ:
                    ws.append([val, f, n, line])
            for col, w in (("A",22),("B",50),("C",7),("D",80)):
                ws.column_dimensions[col].width = w
            for row in ws.iter_rows():
                for c in row:
                    if not c.font or c.font.name != "Consolas": c.font = mono
                    c.alignment = Alignment(vertical="top", wrap_text=True)
            wb.save(fp); log(f"\n[Saved] {fp}")
        except Exception as e:
            log(f"[WARN] kunde ej skriva Excel ({e}) -- konsol-utskrift racker.")
    log("[DONE]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
