#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
structure_probe.py  --  SOND LAGER 1: STRUKTURKVITTO (statisk arkitektur-karta)
================================================================================
Laser kodbasen statiskt (kor INGEN kod, ingen VM, ingen Blob, ingen DW) och
kartlagger HUR systemet ar byggt. Klassar varje fynd:
  FAKTA        -- sa har ar det byggt (neutralt, karta)
  AVVIKELSE    -- skiljer sig fran rimlig forvantan, granska (gult)
  INKONSEKVENS -- tva stallen som borde saga samma sak sager olika (rott)

Sex sektioner:
  A  Filberoenden (vem importerar vem)
  B  Env-variabel-register (vad styr vad; INKONSEKVENS om olika default)
  C  Datumlas-karta (G7-fallan synlig)
  D  Antaganden & frysningar (hardkodade datum, frusna lager)
  E  Lokal-vs-VM-karta (xlwings/COM-beroenden = ej Linux)
  F  Storage-konto-karta (FD.35: prod-default vs test-anvandning)

SJALVUPPTACKANDE: skannar monster (ast for Python, regex for SQL), antar inte
specifika funktionsnamn. Robust mot saknade/flyttade filer (rapporterar, kraschar ej).

KOR (PowerShell, global py -3.11, fran var som helst):
    py -3.11 structure_probe.py --root "C:\\Projekt\\BCG"
    # valfritt: --out <mapp>   (default: <root>\\verify_tool\\receipts\\<datum>\\structure)
    # valfritt: --also-ba "C:\\Projekt\\Business_Analytics"   (inkludera extraktionslagret)

Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia), assisterad av Claude.
Beroende: openpyxl (finns i global 3.11). Std-lib annars (ast, re, pathlib).
"""
from __future__ import annotations
import argparse
import ast
import datetime
import os
import re
import sys
from pathlib import Path

# ---- Klassificering -------------------------------------------------
FACT = "FAKTA"
DEV  = "AVVIKELSE"
INC  = "INKONSEKVENS"

# Kanda forvantningar (for AVVIKELSE-detektion). Justera vid behov.
EXPECTED_STORAGE = "evbcgpricinginput"        # TEST = det ratta (FD.35)
KNOWN_BAD_STORAGE = "evipricingmodelstprod"   # PROD = finns ej / fel default
XLWINGS_HINT = ("xlwings", "win32com", "pywintypes")   # lokal-bara (ej Linux)
DATE_LITERAL = re.compile(r"\b(20\d{2})-(\d{2})-(\d{2})\b")  # YYYY-MM-DD
SQL_DATE = re.compile(r"DATE\s+'(\d{4}-\d{2}-\d{2})'")
YEARFLAG = re.compile(r"YearFlag\s+IN\s*\(([^)]*)\)", re.IGNORECASE)

# Legitimt FASTA datum -- medvetna konstanter, INTE avvikelser:
#   2017-02-22 = parquetens/DW:ns startdatum (datahistorikens borjan)
#   2022-07-01 = vaxande fonstrets fasta ankare (LF.2)
# Dessa ska klassas FAKTA aven nar de ar hardkodade.
LEGIT_FIXED_DATES = {"2017-02-22", "2022-07-01"}
# Filer dar ett hardkodat SLUTDATUM ar extra misstankt (ska ga via env/resolve):
RUNNER_HINT = ("run_site", "run_cluster", "run_bundle", "run_data", "run_after")


def log(msg): print(msg, flush=True)


class Finding:
    __slots__ = ("section", "klass", "what", "where", "detail")
    def __init__(self, section, klass, what, where="", detail=""):
        self.section = section; self.klass = klass
        self.what = what; self.where = where; self.detail = detail


def rel(p: Path, root: Path) -> str:
    try: return str(p.relative_to(root))
    except Exception: return str(p)


def iter_py(root: Path, subdirs):
    """Python-filer under givna subdirs (om de finns)."""
    for sd in subdirs:
        d = root / sd
        if d.is_dir():
            for p in d.rglob("*.py"):
                if ".venv" in p.parts or "site-packages" in p.parts:
                    continue
                yield p


def iter_sql(root: Path):
    for p in root.rglob("*.sql"):
        if ".venv" in p.parts:
            continue
        yield p


def safe_read(p: Path) -> str:
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return p.read_text(encoding=enc)
        except Exception:
            continue
    return ""


# ---- Sektion A: filberoenden ----------------------------------------
def section_A(root, py_files, F):
    imports = {}   # fil -> set(importerade modulnamn, topp-nivå)
    modname = {}   # filnamn-stam -> rel-path (for att hitta interna beroenden)
    for p in py_files:
        modname[p.stem] = rel(p, root)
    for p in py_files:
        src = safe_read(p)
        names = set()
        try:
            tree = ast.parse(src)
            for n in ast.walk(tree):
                if isinstance(n, ast.Import):
                    for a in n.names: names.add(a.name.split(".")[0])
                elif isinstance(n, ast.ImportFrom):
                    if n.module: names.add(n.module.split(".")[0])
        except SyntaxError:
            F.append(Finding("A", DEV, f"Kunde ej parsa (syntaxfel): {p.name}", rel(p, root)))
            continue
        imports[rel(p, root)] = names
    # FAKTA: interna beroenden (importer som matchar en annan projektfil)
    internal_consumers = {}  # modul -> [filer som importerar den]
    for f, names in imports.items():
        internal = sorted(n for n in names if n in modname)
        if internal:
            F.append(Finding("A", FACT, f"{Path(f).name} importerar: {', '.join(internal)}", f))
        for n in internal:
            internal_consumers.setdefault(n, []).append(Path(f).name)
    # FAKTA: omvant -- vem beror PA en central modul
    for mod, consumers in sorted(internal_consumers.items()):
        if len(consumers) >= 2:
            F.append(Finding("A", FACT, f"'{mod}' anvands av {len(consumers)}: {', '.join(sorted(consumers))}",
                             modname.get(mod, ""), "ror du denna paverkas dessa"))


# ---- Sektion B: env-register ----------------------------------------
def section_B(root, py_files, F):
    # var: namn -> {default-värde -> [filer]}
    env = {}
    pat_get = re.compile(r"""os\.environ\.get\(\s*["']([^"']+)["']\s*(?:,\s*(.+?))?\)""")
    pat_idx = re.compile(r"""os\.environ\[\s*["']([^"']+)["']\s*\]""")
    pat_set = re.compile(r"""os\.environ\.setdefault\(\s*["']([^"']+)["']\s*,\s*["']([^"']*)["']""")
    for p in py_files:
        src = safe_read(p)
        for m in pat_get.finditer(src):
            name = m.group(1); default = (m.group(2) or "").strip()
            env.setdefault(name, {}).setdefault(default, []).append(p.name)
        for m in pat_idx.finditer(src):
            env.setdefault(m.group(1), {}).setdefault("(required, no default)", []).append(p.name)
        for m in pat_set.finditer(src):
            env.setdefault(m.group(1), {}).setdefault(f"setdefault={m.group(2)}", []).append(p.name)
    for name, defaults in sorted(env.items()):
        clean = {d: fs for d, fs in defaults.items()}
        # FAKTA-rad per variabel
        ds = "; ".join(f"[{d or 'ingen'}] i {', '.join(sorted(set(fs)))}" for d, fs in clean.items())
        F.append(Finding("B", FACT, f"{name}", "", ds))
        # INKONSEKVENS: olika icke-tomma default for samma variabel
        real_defaults = {d for d in clean if d and not d.startswith("(") and not d.startswith("setdefault")}
        # normalisera bort citattecken for jamforelse
        norm = {re.sub(r'^["\']|["\']$', '', d) for d in real_defaults}
        if len(norm) > 1:
            F.append(Finding("B", INC, f"{name} har OLIKA default pa olika stallen: {sorted(norm)}",
                             "", "tyst-bugg-risk: ett stalle kan anvanda fel varde"))


# ---- Sektion C: datumlas-karta (SKARPT: skiljer legitimt/dynamiskt/misstankt) ----
def section_C(root, py_files, sql_files, F):
    # SQL: DATE-literaler + YearFlag (alltid FAKTA -- det ar dar G7-injektionen biter)
    for p in sql_files:
        src = safe_read(p)
        dates = SQL_DATE.findall(src)
        if dates:
            F.append(Finding("C", FACT, f"SQL DATE-literaler i {p.name}: {sorted(set(dates))}",
                             rel(p, root), "datumfonster-las; skrivs om av _inject_dates (G7)"))
        yf = YEARFLAG.findall(src)
        if yf:
            F.append(Finding("C", FACT, f"YearFlag-vitlista i {p.name}",
                             rel(p, root), "andra G7-laset -- maste flyttas MED veckofonstret"))
    # Python: klassificera varje hardkodat datum-default
    pat_arg = re.compile(r"""default\s*=\s*["'](\d{4}-\d{2}-\d{2})["']""")
    for p in py_files:
        src = safe_read(p)
        is_runner = any(h in p.name.lower() for h in RUNNER_HINT)
        has_dynamic = ("resolve_window_end" in src)   # filen har dynamisk harledning
        for m in pat_arg.finditer(src):
            d = m.group(1)
            if d in LEGIT_FIXED_DATES:
                # Medvetet fast (ankare/parquet-start) -> FAKTA, inte avvikelse
                why = "parquet/DW-start" if d == "2017-02-22" else "vaxande-fonster-ankare (LF.2)"
                F.append(Finding("C", FACT, f"Fast datum '{d}' i {p.name}",
                                 rel(p, root), f"medvetet konstant: {why}"))
            elif is_runner and has_dynamic:
                # Slutdatum-default i en runner SOM HAR resolve_window_end -> FAKTA (fallback)
                F.append(Finding("C", FACT, f"Slutdatum-default '{d}' i {p.name}",
                                 rel(p, root), "runner har resolve_window_end -> detta ar fallback (ok)"))
            elif is_runner:
                # Slutdatum-default i runner UTAN dynamik -> AVVIKELSE (G7-risk)
                F.append(Finding("C", DEV, f"Slutdatum-default '{d}' i {p.name} UTAN dynamisk harledning",
                                 rel(p, root), "G7-risk: lagg resolve_window_end eller env-styrning"))
            else:
                # Annat hardkodat datum i icke-runner -> AVVIKELSE (granska, ofta ofarligt)
                F.append(Finding("C", DEV, f"Hardkodat datum '{d}' i {p.name}",
                                 rel(p, root), "granska: medveten konstant eller kvarvarande hardkodning?"))
        # Positiva markorer (FAKTA -- visar att datum HANTERAS dynamiskt)
        if "BCG_END_DATE" in src or "BCG_START_DATE" in src:
            F.append(Finding("C", FACT, f"{p.name} laser/injicerar BCG_START/END_DATE",
                             rel(p, root), "datum env-styrt (bra)"))
        if has_dynamic:
            F.append(Finding("C", FACT, f"{p.name} anvander resolve_window_end",
                             rel(p, root), "harleder slut ur parquetens data (bra)"))


# ---- Sektion D: antaganden & frysningar -----------------------------
def section_D(root, py_files, F):
    frozen_hints = ("frozen", "facit", "frusen", "_frozen", "weave_weight", "step5", "step-5", "bundle_facit")
    anchor = "2022-07-01"
    for p in py_files:
        src = safe_read(p)
        low = src.lower()
        if anchor in src:
            F.append(Finding("D", FACT, f"Ankare {anchor} refereras i {p.name}",
                             rel(p, root), "vaxande fonstrets fasta start (LF.2)"))
        for h in ("weave", "frozen", "facit", "step5"):
            if h in low and ("froz" in low or "facit" in low or "frus" in low):
                F.append(Finding("D", FACT, f"Mojlig frysnings-referens ('{h}') i {p.name}",
                                 rel(p, root), "kolla LF.9 -- frusna 2025-lager"))
                break
        if "top" in low and ("80" in src or "0.8" in src):
            F.append(Finding("D", FACT, f"Mojligt Top-80-filter i {p.name}", rel(p, root)))


# ---- Sektion E: lokal-vs-VM -----------------------------------------
def section_E(root, py_files, F):
    for p in py_files:
        src = safe_read(p)
        hit = [h for h in XLWINGS_HINT if h in src]
        if hit:
            # AVVIKELSE om filen ocksa ser ut att vara en VM/launcher-fil
            is_launcher = p.name.lower() in ("launcher.py",) or "launcher" in p.name.lower()
            klass = DEV if is_launcher else FACT
            note = ("LOKAL-BARA (Excel-COM, ej Linux). " +
                    ("VM-launcher kor detta -> kraschar pa VM (designfraga)" if is_launcher
                     else "kors lokalt pa Windows"))
            F.append(Finding("E", klass, f"{p.name} anvander {', '.join(hit)}", rel(p, root), note))
        if "setsid" in src or "scp_from_vm" in src or "ssh_run" in src:
            F.append(Finding("E", FACT, f"{p.name} har VM-orkestrering (ssh/scp/setsid)",
                             rel(p, root), "kor steg PA VM (Linux)"))


# ---- Sektion F: storage-konto-karta ---------------------------------
def section_F(root, py_files, F):
    acct_pat = re.compile(r"""["'](ev[a-z0-9]*(?:pricing|model|input|prod|test)[a-z0-9]*)["']""")
    seen = {}
    for p in py_files:
        src = safe_read(p)
        for m in acct_pat.finditer(src):
            a = m.group(1)
            seen.setdefault(a, []).append(p.name)
    for a, files in sorted(seen.items()):
        F.append(Finding("F", FACT, f"Storage-konto-referens '{a}'", "", f"i {', '.join(sorted(set(files)))}"))
        if a == KNOWN_BAD_STORAGE:
            F.append(Finding("F", DEV, f"'{a}' (PROD) refereras -- forvantat konto ar '{EXPECTED_STORAGE}' (TEST)",
                             "", "FD.35: default kan peka pa konto som ej finns/ej anvands"))
    if KNOWN_BAD_STORAGE in seen and EXPECTED_STORAGE in seen:
        F.append(Finding("F", INC, f"BADE '{KNOWN_BAD_STORAGE}' (prod) och '{EXPECTED_STORAGE}' (test) refereras",
                         "", "FD.35: tva konton i koden -- reda ut till ETT hem"))



# ---- NASTA STEG: oversatt fynd -> konkret atgard / nasta sond -------
def derive_next_steps(findings):
    """Varje INKONSEKVENS/AVVIKELSE -> en actionable rad: vad gora, ev. vilken sond."""
    steps = []  # (prioritet 1-3, atgard, motivering, ev_sond)
    inc = [f for f in findings if f.klass == INC]
    dev = [f for f in findings if f.klass == DEV]

    for f in inc:
        if "OLIKA default" in f.what and "BCG_END_DATE" in f.what:
            steps.append((1,
                "Ena BCG_END_DATE-default till ETT varde (28 vs 29 jun ar en dags glapp)",
                "Olika familjer kan inkludera olika manga dagar -> subtil inkonsekvens i vaven.",
                "Vardesond: lista exakt vilken fil har vilket default, ratta till ett."))
        elif "OLIKA default" in f.what:
            steps.append((1, f"Ena default: {f.what}",
                "Tva stallen sager olika -> tyst-bugg-risk.",
                "Vardesond per fil."))
        elif "prod" in f.what and "test" in f.what:
            steps.append((1,
                "Reda ut storage-konto till ETT hem (FD.35)",
                "Bade prod- och test-konto refereras; appen default:ade till prod som ej finns.",
                "Konto-sond: verifiera vilket konto som faktiskt har data (kor mot Blob)."))

    # Gruppera C-avvikelser (G7-risker) till ETT steg
    g7 = [f for f in dev if f.section == "C" and "G7-risk" in f.detail]
    if g7:
        files = sorted({Path(f.where).name for f in g7})
        steps.append((2,
            f"Gor slutdatum dynamiskt i: {', '.join(files)}",
            "Hardkodat slutdatum utan resolve_window_end -> G7-falla (ny manad tappas tyst).",
            "Ingen ny sond -- patcha med resolve_window_end (monstret finns i run_status.py)."))

    # launcher/xlwings pa VM
    xl = [f for f in dev if f.section == "E"]
    if xl:
        files = sorted({Path(f.where).name for f in xl})
        steps.append((2,
            f"Fa VM-korning att stoppa efter Steg 4 (galler: {', '.join(files)})",
            "launcher kor Steg 5 (xlwings) pa Linux-VM -> kraschar alltid; skapar forvirring.",
            "Additiv fix (ror ej BCG-karna): launcher-flagga eller kor steg 1-4 explicit."))

    # F-avvikelser (prod-konto referens) om ej redan tackt av INC
    fdev = [f for f in dev if f.section == "F" and "PROD" in f.what]
    if fdev and not any("FD.35" in st[1] for st in steps):
        steps.append((1, "Byt prod-konto-default till test (evbcgpricinginput)",
            "Default pekar pa konto som ej finns/ej anvands.",
            "Konto-sond + patcha default."))

    # Generella C-avvikelser (icke-G7) -> en samlad granskningsrad
    other_c = [f for f in dev if f.section == "C" and "G7-risk" not in f.detail]
    if other_c:
        files = sorted({Path(f.where).name for f in other_c})
        steps.append((3,
            f"Granska ovriga hardkodade datum (ofta ofarliga): {', '.join(files)}",
            "Kan vara medvetna konstanter (testdatum, valideringsfonster) -- bekrafta.",
            "Manuell blick; lagg legitima i LEGIT_FIXED_DATES sa de slutar flaggas."))

    steps.sort(key=lambda t: t[0])
    return steps

# ---- Excel-kvitto ---------------------------------------------------
def write_receipt(findings, next_steps, out_dir: Path, root: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception as e:
        log(f"[ERROR] openpyxl saknas ({e}). Kor med global py -3.11."); return None
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    fp = out_dir / f"00_structure_master_{stamp}.xlsx"
    wb = Workbook()
    mono = Font(name="Consolas", size=10)
    bold = Font(name="Consolas", size=10, bold=True)
    fills = {FACT: PatternFill("solid", fgColor="E8F0E8"),     # ljusgron-neutral
             DEV:  PatternFill("solid", fgColor="FFF3CD"),     # gul
             INC:  PatternFill("solid", fgColor="F8D7DA")}     # rod
    sect_names = {"A": "A Filberoenden", "B": "B Env-register", "C": "C Datumlas",
                  "D": "D Antaganden", "E": "E Lokal-vs-VM", "F": "F Storage-konto"}

    # Oversiktsflik
    ws = wb.active; ws.title = "Oversikt"
    ws.append([f"STRUKTURKVITTO (Lager 1)  {stamp}"]); ws["A1"].font = bold
    ws.append([f"Root: {root}"]); ws.append([])
    counts = {}
    for f in findings: counts[f.klass] = counts.get(f.klass, 0) + 1
    ws.append(["Klass", "Antal"]); 
    for r in (4,): 
        for c in (1,2): ws.cell(r, c).font = bold
    for k in (INC, DEV, FACT):
        ws.append([k, counts.get(k, 0)])
        ws.cell(ws.max_row, 1).fill = fills[k]
    ws.append([]); ws.append(["LAS FORST: INKONSEKVENS (rott) -> AVVIKELSE (gult) -> FAKTA (neutralt)"])
    ws.cell(ws.max_row, 1).font = bold
    # de viktigaste raderna overst
    ws.append([])
    ws.append(["TVARSNITT av ROTT+GULT (granska dessa):"]); ws.cell(ws.max_row,1).font = bold
    for f in findings:
        if f.klass in (INC, DEV):
            ws.append([f.klass, sect_names.get(f.section, f.section), f.what, f.where])
            ws.cell(ws.max_row, 1).fill = fills[f.klass]
    for col, w in (("A",16),("B",18),("C",70),("D",45)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        for c in row:
            if not c.font or c.font.name != "Consolas": c.font = mono

    # NASTA STEG-flik (det som gor sonden "ledande")
    ws = wb.create_sheet("NASTA STEG")
    ws.append(["NASTA STEG -- prioriterade atgarder ur fynden"]); ws["A1"].font = bold
    ws.append([]); ws.append(["PRIO", "ATGARD", "VARFOR", "SOND / METOD"])
    for c in range(1, 5): ws.cell(3, c).font = bold
    pf = {1: PatternFill("solid", fgColor="F8D7DA"), 2: PatternFill("solid", fgColor="FFF3CD"),
          3: PatternFill("solid", fgColor="E8F0E8")}
    for prio, atgard, varfor, sond in next_steps:
        ws.append([f"P{prio}", atgard, varfor, sond])
        ws.cell(ws.max_row, 1).fill = pf.get(prio, pf[3])
    for col, w in (("A",6),("B",55),("C",55),("D",55)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        for c in row:
            if not c.font or c.font.name != "Consolas": c.font = mono
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # En flik per sektion
    for sec in ("A", "B", "C", "D", "E", "F"):
        ws = wb.create_sheet(sect_names[sec][:31])
        ws.append(["KLASS", "VAD", "VAR", "DETALJ"])
        for c in range(1, 5): ws.cell(1, c).font = bold
        for f in findings:
            if f.section == sec:
                ws.append([f.klass, f.what, f.where, f.detail])
                ws.cell(ws.max_row, 1).fill = fills.get(f.klass, fills[FACT])
        for col, w in (("A",14),("B",75),("C",45),("D",55)):
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows():
            for c in row:
                if not c.font or c.font.name != "Consolas": c.font = mono
                c.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(fp)
    return fp


def main() -> int:
    ap = argparse.ArgumentParser(description="Strukturkvitto (Lager 1) -- statisk arkitektur-karta.")
    ap.add_argument("--root", default=r"C:\Projekt\BCG", help="Projektrot (BCG-repo).")
    ap.add_argument("--also-ba", default=None, help="Inkludera Business_Analytics-rot (extraktionslagret).")
    ap.add_argument("--out", default=None, help="Utdatamapp (default: <root>\\verify_tool\\receipts\\<datum>\\structure).")
    args = ap.parse_args()

    root = Path(args.root)
    if not root.is_dir():
        log(f"[ERROR] root finns ej: {root}"); return 2

    # Subdirs att skanna (de finns kanske inte alla -- iter_py hoppar saknade)
    subdirs = ["orchestration", "verify_tool", "tools",
               "Pipeline/02. Elasticity", "Pipeline"]
    py_files = list(dict.fromkeys(iter_py(root, subdirs)))  # unika, ordningsbevarande
    if args.also_ba:
        ba = Path(args.also_ba)
        if ba.is_dir():
            py_files += [p for p in ba.rglob("*.py") if ".venv" not in p.parts]
    sql_files = list(iter_sql(root))

    log(f"[RUN] root={root}")
    log(f"[RUN] {len(py_files)} Python-filer, {len(sql_files)} SQL-filer att skanna")

    F = []
    section_A(root, py_files, F)
    section_B(root, py_files, F)
    section_C(root, py_files, sql_files, F)
    section_D(root, py_files, F)
    section_E(root, py_files, F)
    section_F(root, py_files, F)

    # Konsolutskrift: bara ROTT+GULT (R7-anda -- struktur-larm syns direkt)
    inc = [f for f in F if f.klass == INC]
    dev = [f for f in F if f.klass == DEV]
    log(f"\n[KPI] {len(F)} fynd: {len(inc)} INKONSEKVENS, {len(dev)} AVVIKELSE, "
        f"{len(F)-len(inc)-len(dev)} FAKTA")
    for f in inc:
        log(f"  [INKONSEKVENS] {f.section}: {f.what}")
    for f in dev[:15]:
        log(f"  [AVVIKELSE]    {f.section}: {f.what}")
    if len(dev) > 15:
        log(f"  ... +{len(dev)-15} fler AVVIKELSE (se kvitto)")

    next_steps = derive_next_steps(F)
    log("\n[NASTA STEG] prioriterade atgarder:")
    for prio, atgard, varfor, sond in next_steps:
        log(f"  P{prio}: {atgard}")
        log(f"       -> {sond}")

    out_dir = Path(args.out) if args.out else (
        root / "verify_tool" / "receipts" / datetime.date.today().isoformat() / "structure")
    fp = write_receipt(F, next_steps, out_dir, root)
    if fp:
        log(f"\n[Saved] {fp}")
    log("[DONE]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
