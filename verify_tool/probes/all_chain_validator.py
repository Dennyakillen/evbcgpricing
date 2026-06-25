#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
all_chain_validator.py  --  end-to-end statisk validering av HELA modellkedjan
==============================================================================
Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia Djursjukvard AB).
Forfattare: Claude advisor, Leverans 2 (Phase Z).

SYFTE (tre i ett)
-----------------
EN sond som i ett enda svep:
  1. VALIDERAR att hela kedjan haller ihop (mat, gissa inte -- R7).
  2. ar ett REPLIKERINGSKONTRAKT: bevisar att en GitHub-klon har varje lank
     den behover, i ratt struktur (kor aven pa en frisk klon utan VM/azure).
  3. RENDERAR en fristaende FLODESKARTA (markdown) som en manniska -- eller en
     framtida AI-session -- kan lasa for att forsta hela kedjan UTAN att lasa
     189 kallfiler. Bifoga kartan som kontext istallet for kod.

KORNINGEN SKER I TRE DELAR (Jens arkitektur)
--------------------------------------------
    FORE  (lokal dator)  ->  MOTOR (Azure-VM)  ->  EFTER (lokal dator)
Koden ar alltsa INTE pa samma plats. FORE+EFTER kor lokalt (DW nas bara via VPN,
xlwings/COM kor bara pa Windows); MOTOR kor pa VM (Ray behover VM:ens RAM).
Sonden gor den tredelningen synlig -- och HARLEDER den ur statuskontraktet
(run_status.default_pipeline + PhaseLocation), den hardkodar den inte.

DEN BARANDE PRINCIPEN (kandidat-karnprincip)
--------------------------------------------
"Harled, deklarera inte tva ganger." Sanningen om vilka faser som finns bor pa
ETT stalle: run_status.default_pipeline(). Den har sonden HARLEDER faserna och
deras miljo (vm/local) darifran och DEKORERAR varje fas med flodesdetaljer
(skript, in/ut, env, kanda fallor). Foljd: lagger nagon till en fas i
default_pipeline utan att uppdatera FLOW_DETAIL, GAR SONDEN SONDER (FAIL) --
det ar en feature, inte en bug. En kopia ar en framtida divergens.

KOR (global Python 3.11, fran repo-roten C:\\Projekt\\BCG)
----------------------------------------------------------
    cd "C:\\Projekt\\BCG"
    py -3.11 verify_tool\\probes\\all_chain_validator.py            # allt statiskt (kor var som helst)
    py -3.11 verify_tool\\probes\\all_chain_validator.py --vm       # + levande MOTOR-kontroller (VM uppe)
    py -3.11 verify_tool\\probes\\all_chain_validator.py --no-receipt --no-flow-md

Det statiska laget kraver INGET utover kallkod (ingen azure, ingen VM, ingen DW).
Optional libs (openpyxl/pandas/duckdb) ger djupare kontroller om de finns, annars
degraderas de till INFO -- sonden faller aldrig pa en frisk klon.

Options:
    --vm           Kor aven VM-sidans levande kontroller (kraver VM uppe + ssh).
    --no-receipt   Skriv inte Excel-receipt.
    --no-flow-md   Rendera inte den fristaende flodeskartan (markdown).
"""
from __future__ import annotations

import argparse
import datetime
import os
import re
import subprocess
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Sokvagar (Z.0-fakta -- enda stallet de bor). Spegel av bundle_chain_validator.
# ---------------------------------------------------------------------------
REPO = Path(r"C:\Projekt\BCG")
ORCH = REPO / "orchestration"
RUNNERS = ORCH / "runners"
SHARED = ORCH / "shared"
INFRA = ORCH / "infrastructure"
ELAST = REPO / "Pipeline" / "02. Elasticity"
VERIFY = REPO / "verify_tool"

RUN_STATUS_PY = SHARED / "run_status.py"
BLOB_PY = INFRA / "blob.py"
AZURE_VM_PY = INFRA / "azure_vm.py"

# De tre modell-runnarna (kloner -- defekt i en = defekt i alla)
FAMILY_RUNNERS = {
    "cluster": RUNNERS / "run_cluster_model.py",
    "site":    RUNNERS / "run_site_model.py",
    "bundle":  RUNNERS / "run_bundle_model.py",
}
# Lokala output_summary-rotter per familj (newest sokes rekursivt darunder)
FAMILY_OUTPUT_ROOTS = {
    "cluster": ELAST / "2. Product Cluster Level Models" / "output",
    "site":    ELAST / "3. Product Site Level Models" / "output",
    "bundle":  ELAST / "5. Bundle Clinic Models" / "output",
}

# FORE/EFTER-orkestrerare
RUN_DATA_PY = RUNNERS / "run_data.py"
RUN_AFTER_PY = RUNNERS / "run_after.py"

# VM (for --vm)
VM_USER = "azureuser"
VM_HOST = os.environ.get("BCG_VM_HOST", "172.18.148.4")
VM_VENV_PY = "/home/azureuser/bcg/cluster/.venv/bin/python"

STAMP = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# ---------------------------------------------------------------------------
# FLOW_DETAIL -- den deklarativa flodesmodellen.
# ---------------------------------------------------------------------------
# Keyad pa fas-nyckel ur default_pipeline. Sonden KONTROLLERAR att nycklarna har
# == default_pipelines faser (synk-koll). Delen (FORE/MOTOR/EFTER) HARLEDS ur
# PhaseLocation + ordning, den star INTE har. Varje post dekorerar fasen med:
#   script    : huvudskript / orkestrerare (lista ok)
#   env_python: vilken python-miljo
#   inputs    : vad fasen laser
#   outputs   : vad fasen skriver
#   env_vars  : styrande miljovariabler
#   traps     : (LB-ref, kort text) -- den hardvunna kunskapen
#   run_cmd   : kommandot som kor fasen
#   todo      : (valfritt) saker som annu ej ar belagda mot kallan
FLOW_DETAIL: dict[str, dict] = {
    # ---------------- FORE (lokal) ----------------
    "extraction": {
        "script": [
            r"orchestration\runners\run_data.py  (orkestrerare: REGEN->PREP->UPLOAD)",
            r"  REGEN : C:\Projekt\Business_Analytics\regenerate_transaction_parquet_chunked.py  (BA-venv, pyodbc)",
            r"  PREP  : tools\replicate_dataprep.py  (DuckDB, py-3.11)",
            r"  UPLOAD: blob.upload_inputs()  -> Blob 'input'",
        ],
        "env_python": "REGEN: BA-venv (pyodbc) | PREP/UPLOAD: global py-3.11",
        "inputs": ["DW (via VPN, port 40615)", "ELLER befintlig transaction_data.parquet"],
        "outputs": [
            r"transaction_data.parquet  (Pipeline\...\Sweden_Elasticity_Data_Prep_SQL\parquet\)",
            "vecko-CSV:er (per familj)",
            "parquet -> Blob 'input'-container (overlever lokala datorn, LB.66)",
        ],
        "env_vars": ["BCG_END_DATE (styr prep-fonstret via _inject_dates, LB.22)",
                     "PRICINGMODEL_AUTH=key"],
        "traps": [
            ("LB.58", "DW nas bara via VPN, INTE fran VM -- darfor kor FORE lokalt"),
            ("LB.62", "parquet-namnskarv: regen skriver _GROWING, 00_read.sql laser utan suffix -> run_data tvingar --out + --overwrite"),
            ("LB.1",  "DW-token fornyas var 4:e timme"),
            ("LB.22", "PREP tar INTE --end som flagga; fonstret kommer via env BCG_END_DATE"),
        ],
        "run_cmd": r"py -3.11 orchestration\runners\run_data.py --end <YYYY-MM-DD>",
    },
    # ---------------- MOTOR (VM) ----------------
    "cluster_model": {
        "script": [r"orchestration\runners\run_cluster_model.py  ->  VM: launcher.py (5 steg)"],
        "env_python": "VM: ~/bcg/cluster/.venv/bin/python (3.11.9, Ray 2.41) -- AGER venv:en",
        "inputs": ["VM: ~/bcg/cluster/data/<MEASURED>.csv (vaxande)",
                   "VM: control_files/ fran tidigare korning"],
        "outputs": ["VM: ~/bcg/cluster/output/model/output_summary.xlsx",
                    "lokalt: ...\\2. Product Cluster Level Models\\output\\azure_run_model\\",
                    "Blob: output/<date>/cluster/"],
        "env_vars": ["BCG_START_DATE", "BCG_END_DATE  (G7-injektion via export i setsid-launch)"],
        "expected_keys": 4180,
        "traps": [
            ("LB.78", "G7-datumlas: constants.py MASTE vara env-overbar (annars kapas vaxande data tyst)"),
            ("LB.18", "feature_selection two-pass: pass 1 skapar control_file + kraschar by design -> runnern relaunchar EN gang"),
            ("LB.44", "launcherns steg 5 (data_prep_after_model_output) kraschar ALLTID pa Linux (xlwings) -- benignt, steg 1-4 = SUCCESS"),
            ("LB.79", "feature_selection skapar INTE sina automl-mappar -> OSError om output/model/* rensats. INGEN runner mkdir:ar dem an (oppen lucka)"),
            ("LB.80", "poll-lognen LOST i kod: pgrep + 'not running and not fresh' -> sann dod-detektering (alla tre runners)"),
        ],
        "run_cmd": r"py -3.11 orchestration\runners\run_cluster_model.py",
    },
    "site_model": {
        "script": [r"orchestration\runners\run_site_model.py  ->  VM: launcher.py (5 steg)"],
        "env_python": "VM: LANAR cluster:s venv (~/bcg/cluster/.venv) -- FAS 13",
        "inputs": ["VM: ~/bcg/site/data/0902_Sweden_weekly_model_data_site_level.csv"],
        "outputs": ["VM: ~/bcg/site/output/model/output_summary.xlsx",
                    "lokalt: ...\\3. Product Site Level Models\\output\\azure_run_model\\",
                    "Blob: output/<date>/site/"],
        "env_vars": ["BCG_START_DATE", "BCG_END_DATE"],
        "expected_keys": 6624,
        "traps": [
            ("LB.78", "G7-env-override (samma som cluster)"),
            ("LB.18", "two-pass (samma)"),
            ("LB.44", "step5 xlwings benign (samma)"),
            ("LB.79", "automl-mapp-lucka (samma)"),
        ],
        "run_cmd": r"py -3.11 orchestration\runners\run_site_model.py",
    },
    "bundle_model": {
        "script": [r"orchestration\runners\run_bundle_model.py  ->  VM: launcher.py (5 steg)"],
        "env_python": "VM: lanar cluster:s venv",
        "inputs": ["VM: ~/bcg/bundle/data/bundle_weekly_model_data_clinic_hospital.xlsx (XLSX, ej CSV -- FD.36)"],
        "outputs": ["VM: ~/bcg/bundle/output/model/output_summary.xlsx (~125 KEY)",
                    "lokalt: ...\\5. Bundle Clinic Models\\output\\azure_run_full\\model\\",
                    "Blob: output/<date>/bundle/"],
        "env_vars": ["BCG_START_DATE", "BCG_END_DATE"],
        "expected_keys": 125,
        "traps": [
            ("LB.78", "bundle constants.py var ALDRIG G7-patchad fore maj-sessionen -- nu env-overbar"),
            ("LB.81", "UPPSTROMS steg C (model-data-creation) anvander Ray -> KRASCHAR lokalt pa Windows. Steg C ar ett VM-steg. CSV->xlsx-brygga kravs (inget BCG-skript gor det)"),
            ("LB.18", "two-pass (samma)"),
            ("LB.44", "step5 xlwings benign (samma)"),
            ("LB.79", "automl-mapp-lucka (samma)"),
        ],
        "run_cmd": r"py -3.11 orchestration\runners\run_bundle_model.py",
    },
    # ---------------- EFTER (lokal) ----------------
    "site_step5": {
        "script": [r"data_prep_after_model_output.py  (i resp. familjs code\, xlwings/COM)"],
        "env_python": "lokal py-3.11 fran modell-roten (Windows-only; xlwings = Excel-COM)",
        "inputs": ["modellens output_summary.xlsx + namnomraden i Excel-mall"],
        "outputs": ["blended/steg5-output (Excel)"],
        "env_vars": [],
        "traps": [
            ("LB.44", "xlwings/COM kan EJ kora pa Linux -- darfor lokal EFTER-fas, aldrig pa VM"),
            ("LB.45/LB.53", "write_df_preserve_named_range fangar KeyError men xlwings kastar com_error -> except Exception; filen ar redan sparad vid icke-noll exit"),
        ],
        "run_cmd": "TODO: verifiera exakt kommando mot kallan (ingen egen runner i inventeringen 2026-06-25)",
        "todo": "Ingen dedikerad runner hittad. Korning verkar manuell (py -3.11 data_prep_after_model_output.py fran modell-roten). Belagg mot kalla nasta steg-5-session.",
    },
    "step6": {
        "script": [r"orchestration\runners\run_after.py  (orkestrerare: PULL->STEG6->STEG7->PUSH)",
                   r"  STEG6: verify_tool\run\run_step6.py  (Fall_Back_Logic-vav, F1-F7)"],
        "env_python": "lokal py-3.11",
        "inputs": [
            "Blob-PULL (download_outputs): 5 Step-6-inputs",
            "  2 LIVE  ur 'output'/<date>  (cluster + site output_summary)",
            "  3 FROZEN ur 'pipeline'      (cluster-steg5 FD.15, bundle FD.11, vav-vikter FD.14)",
        ],
        "outputs": ["Final_Fallback_Data_*.xlsx  -> Blob (samma datummapp)"],
        "env_vars": ["PRICINGMODEL_AUTH=key"],
        "traps": [
            ("LB.52", "step6 forvantar pre-splittad KEY (ItemCode-kolumn); vaxande output har bara KEY -> KeyError nedstroms"),
            ("LB.53", "xlwings com_error om mallens namnomrade saknas (samma klass som LB.45)"),
            ("LB.77", "de 3 frusna lasen (FD.11/14/15) RAPPORTERAS i slutloggen, doljs ej"),
        ],
        "run_cmd": r"py -3.11 orchestration\runners\run_after.py --date-folder <YYYY-MM-DD>",
        "todo": "Finkorniga in-filer for run_step6 ej belagda mot kalla (vag a). Blockerad tills Site+Bundle output_summary finns.",
    },
    "build_r12": {
        "script": [r"orchestration\runners\run_after.py  (samma orkestrerare, STEG7)",
                   r"  STEG7: verify_tool\run\build_r12_for_model.py  (auto-tx)"],
        "env_python": "lokal py-3.11",
        "inputs": ["Final_Fallback_Data_* (fran step6)", "tx-CSV (PULL:ad)"],
        "outputs": ["Model_Feed_*.xlsx  -> Blob"],
        "env_vars": ["PRICINGMODEL_AUTH=key"],
        "traps": [],
        "run_cmd": r"py -3.11 orchestration\runners\run_after.py --date-folder <YYYY-MM-DD>  (kor STEG6+STEG7)",
        "todo": "Finkorniga in-filer ej belagda mot kalla (vag a).",
    },
}

# ---------------------------------------------------------------------------
# Resultatuppsamling (samma kontrakt som bundle_chain_validator)
# ---------------------------------------------------------------------------
ROWS: list[tuple[str, str, str]] = []  # (status, check, detalj)


def rec(status: str, check: str, detalj: str = "") -> None:
    """status: PASS / FAIL / REVIEW / INFO."""
    ROWS.append((status, check, detalj))
    mark = {"PASS": "[PASS]", "FAIL": "[FAIL]", "REVIEW": "[REVIEW]", "INFO": "[INFO]"}.get(status, status)
    line = f"{mark:9} {check}"
    if detalj:
        line += f"  --  {detalj}"
    print(line, flush=True)


def section(title: str) -> None:
    print("\n" + "=" * 78, flush=True)
    print(title, flush=True)
    print("=" * 78, flush=True)
    ROWS.append(("", f"=== {title} ===", ""))


# ---------------------------------------------------------------------------
# Hjalpare
# ---------------------------------------------------------------------------
def read_text(p: Path) -> str:
    for enc in ("utf-8", "cp1252", "utf-16"):
        try:
            return p.read_text(encoding=enc)
        except (UnicodeDecodeError, FileNotFoundError):
            continue
    return ""


def file_info(p: Path) -> str:
    if not p.exists():
        return "SAKNAS"
    st = p.stat()
    mb = st.st_size / 1024 / 1024
    mt = datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
    if mb >= 1:
        return f"{mb:.1f} MB, {mt}"
    return f"{st.st_size:,} B, {mt}"


def grep_count(text: str, pattern: str, flags=re.I) -> int:
    return len(re.findall(pattern, text, flags))


def extract_func(text: str, name: str) -> str:
    """Returnera normaliserad kropp av toppniva-funktionen `name` (def name ...
    fram till nasta toppniva-def/class eller en #===-banner). Whitespace-
    normaliserad sa CRLF/indentering inte stor jamforelsen (LB.35-klassen)."""
    out: list[str] = []
    capturing = False
    for ln in text.splitlines():
        if not capturing:
            if re.match(rf"^def {re.escape(name)}\b", ln):
                capturing = True
                out.append(ln.rstrip())
            continue
        if re.match(r"^(def |class |# =====)", ln):
            break
        out.append(ln.rstrip())
    return "\n".join(l.strip() for l in out if l.strip())


def newest_output_summary(root: Path) -> Path | None:
    if not root.exists():
        return None
    hits = list(root.glob("**/output_summary.xlsx"))
    if not hits:
        return None
    return max(hits, key=lambda p: p.stat().st_mtime)


# ---------------------------------------------------------------------------
# Harled faser + tredelning ur statuskontraktet (sanningskallan)
# ---------------------------------------------------------------------------
def load_pipeline_phases():
    """Importera default_pipeline ur run_status (ren stdlib -> ska alltid ga pa
    en klon). Returnerar (phases | None). Att detta misslyckas ar i sig en
    strukturell defekt -- vi rapporterar den i validate_flow_in_sync."""
    sys.path.insert(0, str(SHARED))
    sys.path.insert(0, str(INFRA))
    try:
        from run_status import default_pipeline  # type: ignore
        return default_pipeline(run_id="all_chain_validator_probe").phases
    except Exception as e:  # noqa: BLE001
        rec("FAIL", "kan importera run_status.default_pipeline",
            f"{type(e).__name__}: {e} -- karnkontraktet saknas/trasigt (replikering bruten)")
        return None


def derive_parts(phases) -> dict[str, str]:
    """Harled FORE/MOTOR/EFTER ur PhaseLocation + ordning. MOTOR = vm-faser;
    lokala faser FORE forsta vm-fasen = FORE; lokala faser efter = EFTER.
    Deklareras inte -- harleds (kandidat-karnprincipen i praktik)."""
    locs = [(p.key, p.location.value) for p in phases]
    vm_idx = [i for i, (_, l) in enumerate(locs) if l == "vm"]
    first_vm = min(vm_idx) if vm_idx else None
    parts: dict[str, str] = {}
    for i, (k, l) in enumerate(locs):
        if l == "vm":
            parts[k] = "MOTOR"
        elif first_vm is not None and i < first_vm:
            parts[k] = "FORE"
        else:
            parts[k] = "EFTER"
    return parts


# ---------------------------------------------------------------------------
# VALIDERINGAR
# ---------------------------------------------------------------------------
def validate_flow_in_sync(phases) -> dict[str, str]:
    """Kandidat-karnprincipen, kodifierad: FLOW_DETAIL maste vara i synk med
    default_pipeline. Drift = FAIL (det ar feature)."""
    section("1. FLODESMODELL <-> STATUSKONTRAKT (harled, deklarera ej tva ganger)")
    parts: dict[str, str] = {}
    if phases is None:
        rec("REVIEW", "synk-koll hoppas over", "kunde ej lasa default_pipeline (se ovan)")
        return parts
    pipe_keys = [p.key for p in phases]
    flow_keys = list(FLOW_DETAIL.keys())
    parts = derive_parts(phases)

    rec("INFO", "faser i default_pipeline", f"{len(pipe_keys)}: {', '.join(pipe_keys)}")
    missing_in_flow = [k for k in pipe_keys if k not in flow_keys]
    extra_in_flow = [k for k in flow_keys if k not in pipe_keys]
    if missing_in_flow:
        rec("FAIL", "FLOW_DETAIL saknar fas(er) ur default_pipeline",
            f"{', '.join(missing_in_flow)} -- lagg till i FLOW_DETAIL (drift)")
    if extra_in_flow:
        rec("FAIL", "FLOW_DETAIL har fas(er) som ej finns i default_pipeline",
            f"{', '.join(extra_in_flow)} -- borttagen fas? (drift)")
    if not missing_in_flow and not extra_in_flow:
        rec("PASS", "FLOW_DETAIL i synk med default_pipeline", f"{len(pipe_keys)} faser matchar")

    # tredelning harledd
    for part in ("FORE", "MOTOR", "EFTER"):
        ks = [k for k in pipe_keys if parts.get(k) == part]
        rec("INFO", f"del {part} (harledd ur PhaseLocation)", ", ".join(ks) or "(tom)")
    return parts


def validate_replication_contract(phases) -> None:
    """Replikeringskontrakt: finns varje skript/lank FLOW deklarerar? Det har ar
    'GitHub-klonen har allt'-kollen. Kor utan VM/azure."""
    section("2. REPLIKERINGSKONTRAKT -- finns varje deklarerad lank?")
    core = [
        (RUN_STATUS_PY, "kontrakt: run_status.py"),
        (BLOB_PY, "infra: blob.py"),
        (AZURE_VM_PY, "infra: azure_vm.py"),
        (RUN_DATA_PY, "FORE-orkestrerare: run_data.py"),
        (RUN_AFTER_PY, "EFTER-orkestrerare: run_after.py"),
    ]
    for fr in FAMILY_RUNNERS.values():
        core.append((fr, f"MOTOR-runner: {fr.name}"))
    for p, label in core:
        info = file_info(p)
        rec("PASS" if info != "SAKNAS" else "FAIL", label, info)

    # belagda hjalpskript ur FLOW (de med fast sokvag i repot)
    referenced = [
        (REPO / "tools" / "replicate_dataprep.py", "FORE: replicate_dataprep.py"),
        (VERIFY / "run" / "run_step6.py", "EFTER: run_step6.py"),
        (VERIFY / "run" / "build_r12_for_model.py", "EFTER: build_r12_for_model.py"),
        (VERIFY / "output_rationality" / "run_all_rationality.py", "validering: run_all_rationality.py"),
    ]
    for p, label in referenced:
        info = file_info(p)
        rec("PASS" if info != "SAKNAS" else "REVIEW", label, info)


def validate_runner_sync() -> None:
    """De tre MOTOR-runnarna ar kloner -> defekt i en = defekt i alla. Bevisa
    att hardningen (poll-dod, G7, two-pass, step5-benign) finns i ALLA tre, och
    att poll-kroppen ar identisk."""
    section("3. RUNNER-SYNK -- ar de tre MOTOR-runnarna kloner (horisontell sanning)?")
    texts = {fam: read_text(p) for fam, p in FAMILY_RUNNERS.items()}

    markers = {
        "poll-dod (pgrep)":        r"pgrep -f launcher\.py",
        "dod-detektering":         r"not running and not fresh",
        "G7-injektion":            r"export BCG_START_DATE=\{start_date\} BCG_END_DATE=\{end_date\}",
        "two-pass-relaunch":       r"TWO_PASS_SIG",
        "step5-benign (LB.44)":    r"BENIGN_STEP5",
        "utfallsstyrd dealloc":    r"_handle_outcome",
    }
    for label, pat in markers.items():
        present = {fam: bool(re.search(pat, t)) for fam, t in texts.items()}
        missing = [fam for fam, ok in present.items() if not ok]
        if not missing:
            rec("PASS", f"alla 3 runners: {label}", "narvarande i cluster/site/bundle")
        else:
            rec("REVIEW", f"{label} SAKNAS i: {', '.join(missing)}",
                "kloner har divergerat -- horisontell validering bruten")

    # poll-kroppen byte-(normaliserat)-identisk?
    bodies = {fam: extract_func(t, "poll_until_done") for fam, t in texts.items()}
    uniq = set(b for b in bodies.values() if b)
    if len(uniq) == 1 and all(bodies.values()):
        rec("PASS", "poll_until_done identisk i alla 3", "LB.80-fixen ar i samtliga (kloner)")
    elif not all(bodies.values()):
        miss = [f for f, b in bodies.items() if not b]
        rec("REVIEW", "poll_until_done ej extraherbar", f"saknas/anonym i: {', '.join(miss)}")
    else:
        rec("REVIEW", "poll_until_done SKILJER mellan familjer",
            "kropparna har divergerat -- granska varfor (LB.80 kanske ej i alla)")


def validate_automl_gap() -> None:
    """Den enda av de fem fallorna som annu ar oppen: LB.79. Ingen runner skapar
    automl-mapparna. Flagga den explicit (additiv rad i preflight_remote)."""
    section("4. AUTOML-MAPP-LUCKAN (LB.79 -- enda oppna fallan i runnern)")
    found_any = False
    for fam, p in FAMILY_RUNNERS.items():
        txt = read_text(p)
        hits = grep_count(txt, r"automl|model_objects")
        if hits:
            found_any = True
            rec("PASS", f"{fam}: automl/model_objects-mkdir", f"{hits} traffar")
        else:
            rec("REVIEW", f"{fam}: automl-mapp EJ hanterad",
                "ingen mkdir for output/model/automl|model_objects -> OSError om output rensats (LB.79)")
    if not found_any:
        rec("INFO", "atgardsforslag (additiv)",
            "lagg i preflight_remote: mkdir -p output/model/automl/{details,results} output/model/model_objects")


def validate_expected_keys() -> None:
    """Mat, gissa inte: jamfor varje runners EXPECTED_KEYS mot KEY-antalet i
    senaste lokala output_summary. Kraver openpyxl/pandas -- degraderas annars."""
    section("5. EXPECTED_KEYS <-> SENASTE OUTPUT (mat, gissa inte)")
    try:
        import openpyxl  # noqa: F401
        have_xlsx = True
    except ImportError:
        have_xlsx = False
        rec("INFO", "openpyxl saknas", "hoppar KEY-antalskoll (kor i global py-3.11)")

    for fam, runner in FAMILY_RUNNERS.items():
        txt = read_text(runner)
        m = re.search(r"EXPECTED_KEYS\s*=\s*(\d+)", txt)
        exp = int(m.group(1)) if m else None
        rec("INFO", f"{fam}: EXPECTED_KEYS (deklarerad)", str(exp) if exp is not None else "ej hittad")

        out = newest_output_summary(FAMILY_OUTPUT_ROOTS[fam])
        if out is None:
            rec("REVIEW", f"{fam}: senaste output_summary", "ingen hittad under output\\ (ej kord an?)")
            continue
        rec("INFO", f"{fam}: senaste output_summary", f"{file_info(out)}  ({out.parent.name}\\)")
        if not have_xlsx or exp is None:
            continue
        try:
            import openpyxl
            wb = openpyxl.load_workbook(out, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            hdr = [str(c) for c in rows[0]] if rows else []
            if "KEY" in hdr:
                ki = hdr.index("KEY")
                keys = len({r[ki] for r in rows[1:] if r[ki] is not None})
            else:
                keys = len(rows) - 1
            verdict = "PASS" if keys == exp else "REVIEW"
            rec(verdict, f"{fam}: KEY-antal {keys} vs EXPECTED {exp}",
                "matchar" if keys == exp else "skiljer -- vaxande data driftar (kan vara vantat, IB.6/IB.11)")
        except Exception as e:  # noqa: BLE001
            rec("INFO", f"{fam}: KEY-antal ej last", f"{type(e).__name__}: {e}")


def validate_status_contract() -> None:
    """Statuskontraktets halsa: finalize() finns, succeed() ar dodkod (definierad
    men oanvand), containrar deklarerade pa ett stalle."""
    section("6. STATUSKONTRAKT & BLOB-KONTRAKT")
    rs_txt = read_text(RUN_STATUS_PY)
    if "def finalize(" in rs_txt:
        rec("PASS", "run_status.finalize() finns", "harleder run-tillstand ur faserna")
    else:
        rec("REVIEW", "run_status.finalize() saknas", "vantad efter heartbeat-refaktorn")

    # succeed() dodkod? definierad men ingen runner anropar .succeed(
    succeed_defined = "def succeed(" in rs_txt
    callers = []
    for label, p in [("run_data", RUN_DATA_PY), ("run_after", RUN_AFTER_PY),
                     *[(f, fp) for f, fp in FAMILY_RUNNERS.items()]]:
        if re.search(r"\.succeed\(", read_text(p)):
            callers.append(label)
    if succeed_defined and not callers:
        rec("REVIEW", "run_status.succeed() ar dodkod",
            "definierad men ingen runner anropar den -- foraldrad av finalize(), stada (LB.29-klass)")
    elif succeed_defined:
        rec("INFO", "succeed() anvands av", ", ".join(callers))

    # Blob-containrar (Z.0-fakta, ett stalle)
    blob_txt = read_text(BLOB_PY)
    for c in ["runstatus", "output", "input", "pipeline"]:
        present = (f'"{c}"' in blob_txt) or (f"'{c}'" in blob_txt)
        rec("PASS" if present else "REVIEW", f"Blob-container '{c}' deklarerad",
            "i blob.py" if present else "ej funnen i blob.py")
    # download_outputs Step-6-inputs (run_after drar dessa)
    n_inputs = grep_count(blob_txt, r'"label":')
    rec("INFO", "Blob download_outputs Step-6-inputs", f"{n_inputs} deklarerade (vantat: 5 -- 2 LIVE + 3 FROZEN)")


def validate_workspace_shadows() -> None:
    """Replikeringsrisk: stale dubletter i workspace\\ skuggar kanon i
    orchestration\\. En klon far inte plocka fel blob.py."""
    section("7. SKUGG-KOPIOR (replikeringsrisk -- kanon vs arkiv)")
    ws = REPO / "workspace"
    shadows = {
        "blob.py": INFRA / "blob.py",
        "run_status.py": SHARED / "run_status.py",
        "azure_vm.py": INFRA / "azure_vm.py",
        "run_site_model.py": RUNNERS / "run_site_model.py",
    }
    any_shadow = False
    for name, canon in shadows.items():
        wp = ws / name
        if wp.exists():
            any_shadow = True
            ws_kb = wp.stat().st_size / 1024
            cn_kb = canon.stat().st_size / 1024 if canon.exists() else 0
            rec("REVIEW", f"workspace\\{name} skuggar kanon",
                f"workspace {ws_kb:.1f} KB vs orchestration {cn_kb:.1f} KB -- kanon = orchestration\\, workspace\\ = arkiv")
    if not any_shadow:
        rec("PASS", "inga skugg-kopior i workspace\\", "kanon entydig")
    rec("INFO", "kanon for replikering", "kod = orchestration\\ + verify_tool\\ + tools\\ ; data = Blob (input/output/pipeline)")


def validate_vm_side(parts: dict[str, str]) -> None:
    """MOTOR-bandet, levande. Kors bara med --vm (kraver VM uppe). Speglar
    bundle_chain_validator.validate_vm_side men tvars familjer."""
    section("8. MOTOR-SIDAN (--vm) -- levande VM-kontroller")

    def ssh(cmd: str, timeout=60) -> str:
        full = ["ssh", f"{VM_USER}@{VM_HOST}", cmd]
        try:
            cp = subprocess.run(full, capture_output=True, text=True, timeout=timeout)
            return (cp.stdout.strip() or cp.stderr.strip())
        except Exception as e:  # noqa: BLE001
            return f"SSH-FEL: {e}"

    probe = ssh("echo VM_OK")
    if "VM_OK" not in probe:
        rec("REVIEW", "VM nabar", f"ssh svarade ej ({probe[:60]}) -- starta VM, eller kor utan --vm")
        return
    rec("PASS", "VM nabar", "ssh OK")

    ray_vm = ssh(f"{VM_VENV_PY} -c 'import ray,pandas; print(ray.__version__,pandas.__version__)'")
    rec("INFO", "VM venv Ray+pandas", ray_vm)

    motor_families = [k for k, v in parts.items() if v == "MOTOR"] or ["cluster", "site", "bundle"]
    name_map = {"cluster_model": "cluster", "site_model": "site", "bundle_model": "bundle"}
    for key in motor_families:
        fam = name_map.get(key, key)
        base = f"/home/azureuser/bcg/{fam}"
        # LB.79: finns automl-mappstrukturen (eller maste den skapas fore D)?
        automl = ssh(f"test -d {base}/output/model/automl && echo JA || echo NEJ")
        rec("INFO" if "JA" in automl else "REVIEW", f"{fam}: VM automl-mapp finns?",
            automl + ("" if "JA" in automl else " -- mkdir kravs fore korning (LB.79)"))
        # input pa plats?
        inp = ssh(f"ls -t {base}/data/* 2>/dev/null | head -1 || echo SAKNAS")
        rec("INFO", f"{fam}: VM senaste input", inp)


# ---------------------------------------------------------------------------
# Utdata: fristaende flodeskarta (markdown) -- AI-kontext-artefakten
# ---------------------------------------------------------------------------
def phase_label(p) -> str:
    """Lasbart fasnamn utan att gissa faltnamnet: prova kandidater, fall tillbaka pa key."""
    for attr in ("title", "name", "label", "display_name", "desc", "description"):
        v = getattr(p, attr, None)
        if isinstance(v, str) and v:
            return v
    return getattr(p, "key", "?")

def render_flow_md(phases, parts: dict[str, str]) -> Path | None:
    """Rendera en KOD-FRI flodeskarta som ar fullstandig i sig sjalv: bifoga den
    till en framtida AI-session istallet for kallkod. Harledd ur koden vid
    korning, lasbar utan den."""
    if phases is None:
        return None
    title_map = {p.key: phase_label(p) for p in phases}
    L: list[str] = []
    L.append("# BCG Pricing -- modellkedjans flode (FORE -> MOTOR -> EFTER)")
    L.append("")
    L.append(f"_Genererad {datetime.datetime.now():%Y-%m-%d %H:%M} av all_chain_validator.py, "
             f"harledd ur run_status.default_pipeline. Utvecklare: Jens Palmo._")
    L.append("")
    L.append("Kedjan kor i tre delar pa tva platser: **FORE** och **EFTER** lokalt "
             "(DW kraver VPN, xlwings kraver Windows), **MOTOR** pa Azure-VM (Ray kraver VM-RAM).")
    L.append("")
    # oversiktstabell
    L.append("## Oversikt")
    L.append("")
    L.append("| Del | Fas | Miljo | Huvudskript |")
    L.append("|-----|-----|-------|-------------|")
    for p in phases:
        d = FLOW_DETAIL.get(p.key, {})
        scripts = d.get("script", ["?"])
        first = scripts[0].split("(")[0].strip() if scripts else "?"
        L.append(f"| {parts.get(p.key,'?')} | {p.key} | {p.location.value} | {first} |")
    L.append("")

    # per del
    for part in ("FORE", "MOTOR", "EFTER"):
        keys = [p.key for p in phases if parts.get(p.key) == part]
        if not keys:
            continue
        loc = "lokal dator" if part in ("FORE", "EFTER") else "Azure-VM"
        L.append(f"## {part}  ({loc})")
        L.append("")
        for k in keys:
            d = FLOW_DETAIL.get(k, {})
            L.append(f"### {k} -- {title_map.get(k,'')}")
            L.append("")
            for s in d.get("script", []):
                L.append(f"- **Skript:** `{s}`" if not s.startswith("  ") else f"    - `{s.strip()}`")
            if d.get("env_python"):
                L.append(f"- **Miljo:** {d['env_python']}")
            if d.get("inputs"):
                L.append(f"- **In:** {'; '.join(d['inputs'])}")
            if d.get("outputs"):
                L.append(f"- **Ut:** {'; '.join(d['outputs'])}")
            if d.get("env_vars"):
                L.append(f"- **Env:** {', '.join(d['env_vars'])}")
            if d.get("expected_keys") is not None:
                L.append(f"- **EXPECTED_KEYS:** {d['expected_keys']}")
            if d.get("traps"):
                L.append("- **Kanda fallor:**")
                for ref, txt in d["traps"]:
                    L.append(f"    - `{ref}` {txt}")
            if d.get("todo"):
                L.append(f"- **TODO (ej belagt mot kalla):** {d['todo']}")
            L.append(f"- **Kor:** `{d.get('run_cmd','?')}`")
            L.append("")

    # fallsammanfattning + replikering
    L.append("## Kanda fallor -- sammanfattning")
    L.append("")
    L.append("| LB | Fas(er) | Kort |")
    L.append("|----|---------|------|")
    seen: dict[str, tuple[set, str]] = {}
    for k, d in FLOW_DETAIL.items():
        for ref, txt in d.get("traps", []):
            key = ref.split("/")[0]
            if key not in seen:
                seen[key] = (set(), txt)
            seen[key][0].add(k)
    for ref in sorted(seen):
        fams, txt = seen[ref]
        L.append(f"| {ref} | {', '.join(sorted(fams))} | {txt} |")
    L.append("")
    L.append("## Replikering (GitHub-klon)")
    L.append("")
    L.append("- **Kanon (kod):** `orchestration\\` + `verify_tool\\` + `tools\\`.")
    L.append("- **Arkiv (kor EJ):** `workspace\\`, `_ATT_RADERA\\` -- stale dubletter, stadas.")
    L.append("- **Data (overlever lokala datorn):** Blob-containrar `input`, `output`, `pipeline`, `runstatus`.")
    L.append("- **Statuskontrakt:** `run_status.py` -- EN statusfil per datafonster (run_id), "
             "alla 7 faser i samma fil (FORE+MOTOR+EFTER).")
    L.append("")

    out_dir = REPO / "workspace" / "flow_map"
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        out = out_dir / f"BCG_flow_map_{STAMP}.md"
        out.write_text("\n".join(L), encoding="utf-8")
        return out
    except Exception as e:  # noqa: BLE001
        print(f"[flow-md] kunde ej skriva: {e}")
        return None


# ---------------------------------------------------------------------------
# Utdata: Excel-receipt (samma stil som bundle_chain_validator -- Jens standard)
# ---------------------------------------------------------------------------
def write_receipt() -> Path | None:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        print("\n[receipt] openpyxl saknas -- hoppar over Excel-receipt.")
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "all_chain"
    mono = Font(name="Consolas", size=10)
    bold = Font(name="Consolas", size=10, bold=True)
    hdr = ["all_chain_validator", f"kord {datetime.datetime.now():%Y-%m-%d %H:%M:%S}",
           "Jens Palmo / Evidensia"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=i, column=1, value=h)
        c.font = bold
    r = len(hdr) + 2
    ws.cell(row=r, column=1, value="STATUS").font = bold
    ws.cell(row=r, column=2, value="CHECK").font = bold
    ws.cell(row=r, column=3, value="DETALJ").font = bold
    r += 1
    for status, check, detalj in ROWS:
        ws.cell(row=r, column=1, value=status).font = mono
        ws.cell(row=r, column=2, value=check).font = (bold if check.startswith("===") else mono)
        ws.cell(row=r, column=3, value=detalj).font = mono
        r += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 96
    out_dir = REPO / "workspace" / "validation_receipts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"all_chain_validator_{STAMP}.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(
        description="End-to-end statisk validering av hela modellkedjan (FORE/MOTOR/EFTER).")
    ap.add_argument("--vm", action="store_true", help="Kor aven levande MOTOR-kontroller (kraver VM uppe).")
    ap.add_argument("--no-receipt", action="store_true", help="Skriv inte Excel-receipt.")
    ap.add_argument("--no-flow-md", action="store_true", help="Rendera inte flodeskartan (markdown).")
    args = ap.parse_args()

    print("=" * 78)
    print("ALL CHAIN VALIDATOR  --  hela modellkedjan, FORE -> MOTOR -> EFTER")
    print(f"  {datetime.datetime.now():%Y-%m-%d %H:%M:%S}   repo={REPO}")
    print("=" * 78)

    phases = load_pipeline_phases()
    parts = validate_flow_in_sync(phases)
    validate_replication_contract(phases)
    validate_runner_sync()
    validate_automl_gap()
    validate_expected_keys()
    validate_status_contract()
    validate_workspace_shadows()
    if args.vm:
        validate_vm_side(parts)

    # flodeskarta (AI-kontext) -- skrivs aven om --vm ej anvants
    flow_path = None
    if not args.no_flow_md:
        flow_path = render_flow_md(phases, parts)

    # sammanfattning
    section("SAMMANFATTNING")
    n_pass = sum(1 for s, _, _ in ROWS if s == "PASS")
    n_fail = sum(1 for s, _, _ in ROWS if s == "FAIL")
    n_rev = sum(1 for s, _, _ in ROWS if s == "REVIEW")
    print(f"  PASS={n_pass}   FAIL={n_fail}   REVIEW={n_rev}")
    if n_fail:
        print("  -> FAIL finns: strukturell drift eller bruten lank. Atgarda fore korning.")
    else:
        print("  -> Inga FAIL. REVIEW = vantade designnoter + oppna kandidater (las dem).")

    if flow_path:
        print(f"\n[flow-md] flodeskarta (AI-kontext): {flow_path}")
    if not args.no_receipt:
        rp = write_receipt()
        if rp:
            print(f"[receipt] {rp}")

    return 1 if n_fail else 0


if __name__ == "__main__":
    sys.exit(main())
