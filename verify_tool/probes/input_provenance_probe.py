#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
input_provenance_probe.py -- the missing axis: is the RIGHT MATERIAL in place BEFORE the weave?
================================================================================================
Why this exists (read this first)
    On 2026-07-03 the EFTER chain (FD.37) was proven end-to-end -- but only after FIVE runs
    and THREE silent contaminations, all one root cause: an input file the code READ was not
    the one the WINDOW required, and nothing measured it until step6 had already woven on the
    wrong material. A pipeline is only as fresh as its STALEST input; a green status phase
    proves the PROCESS ran, not that the right MATERIAL went in. We have validated process for
    months (all_chain_validator, dry_run_pipeline, provenance/rationality receipts) but never
    input freshness per window. This probe is that missing axis.

    Had it run BEFORE the first run_after, it would have printed three RED rows in 20 seconds:
      output_summary_ready (cluster)  LOCAL 5889B (stub!)          -> RED
      output_summary (site)           BLOB 06-17 (pre-maj)         -> RED
      model_summary (site)            LOCAL 06-10, missing in Blob -> RED (VM disk: ~/bcg/site/)
    Instead of five runs + three contaminations: read three reds -> repair three files -> one run.

What it is / is NOT
    It is the FORE-stage freshness/placement check, callable from dry_run_e2e --stage fore.
    It is NOT a fifth probe generation: it adds the axis dry_run_e2e lacks (MOTOR-drift +
    leak scan it already has). It touches no other probe; consolidation is BB.1.

Three axes per key file (a file is GREEN only if all three hold)
    1. LOCAL  : exists at the EXACT path the code reads (from blob._AFTER_INPUTS dest / Constant.py),
                NOT in an azure_run_* mirror or an archive; AND passes a size-sanity floor
                (a 5,889-byte xlsx cannot carry thousands of KEY -> stub).
    2. BLOB   : exists in Blob at the run_id path (survives Jens's laptop -- the robustness goal).
    3. WINDOW : timestamp lands in the target window (maj, not left-over april).
    GREEN = all three. YELLOW = axis 3 ambiguous (deterministically regenerable, e.g.
    ivc_sweden_price.csv -- same bytes regardless of window, old date but fresh content).
    RED = any axis fails.

Registry is DERIVED, never re-declared (LB.85): read from blob._AFTER_INPUTS at runtime.
If that import fails, fall back to a documented static mirror (clearly marked).

Auth: reuses blob.py env handling (PRICINGMODEL_AUTH=key) and its _read_account_key
(az keys list --resource-group -- LB.88: sub-wide lookup needs rights outside scoped PIM).
Never verifies via print-to-read; every gate is code.

Usage
    py -3.11 verify_tool\\probes\\input_provenance_probe.py --window 2022-07-01_2026-05-31 --date-folder 2026-06-17
    py -3.11 verify_tool\\probes\\input_provenance_probe.py --window <id> --date-folder <d> --no-blob   (local axis only)
    py -3.11 verify_tool\\probes\\input_provenance_probe.py --window <id> --date-folder <d> --window-start 2026-05

Lessons honored: LB.85 (derive), LB.86 (stub size-sanity), LB.88 (key lookup + RG),
LB.89 (per-input window verification), LB.90 (VM disk is last resort -- hinted, not started),
R7 (receipts over exit codes), A.9b (read source paths, don't guess).

Developer : Jens Palmö (Senior Business Analyst, Evidensia). Author: Claude advisor.
Created   : 2026-07-03
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(os.environ.get("BCG_REPO", r"C:\Projekt\BCG"))
if not REPO.exists():
    REPO = Path.cwd()
RECEIPT_DIR = Path(os.environ.get("INPUT_PROV_RECEIPT_DIR",
                                  str(REPO / "workspace" / "validation_receipts")))

sys.path.insert(0, str(REPO / "orchestration" / "infrastructure"))
sys.path.insert(0, str(REPO / "orchestration" / "shared"))

RESOURCE_GROUP = os.environ.get("PRICINGMODEL_RG", "ev-openai-swce-rg-test")

# --------------------------------------------------------------------------- file-class policy
# Size floors (bytes): a file below this on the LOCAL axis is a stub, not real output.
# Tuned from 2026-07-03 measurements (site summary ~561KB, cluster ready ~368KB, model_summary ~52MB).
SIZE_FLOOR = {
    "cluster output_summary": 100_000,   # real ~325KB; the stub was 5,889 B
    "site output_summary":    100_000,   # real ~561KB
    "model_summary":          1_000_000, # real ~52MB
    "output_summary_ready":   100_000,   # real ~368KB (KEY-split); the stub was 5,889 B
    "_frozen":                1_000,      # frozen facit files: existence-level floor only
    "_regen":                 1_000,      # regenerable (regular price): floor only, window N/A
}
# Files whose content is deterministic from static raw data -> window axis is N/A (YELLOW not RED).
REGENERABLE_HINTS = ("ivc_sweden_price", "regular price")
# Frozen facit -> window axis N/A by design (FD.11/14/15).
FROZEN_HINTS = ("frozen_facit", "00_frozen", "final_model_cluster_granularity_Ivce",
                "Complete_Product_Data", "output_summary_bundle")

RESULTS: "list[dict]" = []


def log(status: str, label: str, local: str, blob: str, window: str, note: str = "") -> None:
    RESULTS.append({"status": status, "label": label, "local": local,
                    "blob": blob, "window": window, "note": note})
    tag = {"GREEN": "[GRÖN]", "YELLOW": "[GUL ]", "RED": "[RÖD ]",
           "GATE": "[GATE]", "INFO": "[INFO]"}.get(status, f"[{status}]")
    print(f"{tag} {label:<32} L:{local:<20} B:{blob:<18} W:{window:<12} {note}", flush=True)


def classify(label: str, blob_path: str) -> str:
    low = (label + " " + blob_path).lower()
    if any(h.lower() in low for h in FROZEN_HINTS):
        return "_frozen"
    if any(h.lower() in low for h in REGENERABLE_HINTS):
        return "_regen"
    for key in SIZE_FLOOR:
        if key.startswith("_"):
            continue
        if key.lower() in label.lower():
            return key
    return ""  # unknown -> floor 0, existence only


def floor_for(cls: str) -> int:
    return SIZE_FLOOR.get(cls, 0)


def load_registry() -> "list[dict]":
    """Derive the input registry from blob._AFTER_INPUTS (LB.85: never re-declare paths)."""
    try:
        import blob  # type: ignore
        reg = []
        for item in blob._AFTER_INPUTS:
            reg.append({"label": item["label"],
                        "container": item["container"],
                        "blob": item["blob"],
                        "dest": item["dest"]})
        log("INFO", f"registry: {len(reg)} inputs derived from blob._AFTER_INPUTS", "-", "-", "-")
        return reg
    except Exception as e:
        log("INFO", "registry: blob._AFTER_INPUTS import FAILED -- static mirror",
            "-", "-", "-", f"{type(e).__name__}: {e}")
        # Documented static mirror (must match blob._AFTER_INPUTS; update if that changes).
        base = r"Pipeline\02. Elasticity"
        return [
            {"label": "cluster output_summary (LIVE)", "container": "output",
             "blob": "{date}/cluster/model/output_summary.xlsx",
             "dest": rf"{base}\2. Product Cluster Level Models\_archive_growing_2026-04-27_v2_pg4fix\output_summary.xlsx"},
            {"label": "site output_summary (LIVE)", "container": "output",
             "blob": "{date}/output_summary.xlsx",
             "dest": rf"{base}\3. Product Site Level Models\output\model\output_summary.xlsx"},
        ]


def blob_meta(container: str, blob_name: str, key: str) -> "tuple[bool,int,str]":
    """Return (exists, bytes, lastModified-date). Uses az CLI to avoid SDK auth divergence."""
    import subprocess, json
    cmd = ("az storage blob show "
           f"--account-name evbcgpricinginput --container-name {container} "
           f'--name "{blob_name}" --account-key {key} '
           "--query \"{len:properties.contentLength, mod:properties.lastModified}\" -o json")
    try:
        p = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=45)
        if p.returncode != 0:
            return (False, 0, "-")
        d = json.loads(p.stdout)
        return (True, int(d.get("len") or 0), str(d.get("mod") or "-")[:10])
    except Exception:
        return (False, 0, "-")


def in_window(date_str: str, window_start: str) -> "bool | None":
    """True if date >= window_start (YYYY-MM). None if undecidable."""
    if not date_str or date_str == "-":
        return None
    try:
        return date_str[:7] >= window_start[:7]
    except Exception:
        return None


def main() -> int:
    ap = argparse.ArgumentParser(description="Input freshness/placement probe (3 axes, pre-run).")
    ap.add_argument("--window", required=True, help="run_id, e.g. 2022-07-01_2026-05-31")
    ap.add_argument("--date-folder", required=True, help="Blob date folder, e.g. 2026-06-17")
    ap.add_argument("--window-start", default=None,
                    help="YYYY-MM the window's data should reach into (default: derive from run_id end month)")
    ap.add_argument("--no-blob", action="store_true", help="local axis only (offline)")
    args = ap.parse_args()

    # window-start default: the END month of the run_id (maj = 2026-05 for ..._2026-05-31)
    if args.window_start:
        wstart = args.window_start
    else:
        end = args.window.split("_")[-1]      # 2026-05-31
        wstart = end[:7]                       # 2026-05

    print(f"=== input_provenance_probe | window {args.window} | date-folder {args.date_folder} "
          f"| freshness threshold >= {wstart} ===", flush=True)

    key = None
    if not args.no_blob:
        import subprocess
        cmd = ("az storage account keys list --account-name evbcgpricinginput "
               f"--resource-group {RESOURCE_GROUP} --query \"[0].value\" -o tsv")
        r = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=45)
        key = (r.stdout or "").strip()
        if not key:
            log("GATE", "Blob key unavailable -- degrading to --no-blob", "-", "-", "-",
                "PIM active? right subscription? (LB.88) Local axis still runs.")
            args.no_blob = True

    registry = load_registry()

    for item in registry:
        label = item["label"]
        cls = classify(label, item["blob"])
        floor = floor_for(cls)
        dest = REPO / item["dest"] if not os.path.isabs(item["dest"]) else Path(item["dest"])

        # ---- LOCAL axis ----
        if dest.exists():
            sz = dest.stat().st_size
            mt = datetime.fromtimestamp(dest.stat().st_mtime).strftime("%Y-%m-%d")
            if sz < floor:
                local_ax, local_txt = "RED", f"{sz}B STUB!"
            else:
                local_ax, local_txt = "OK", f"{sz//1024}KB {mt}"
        else:
            local_ax, local_txt, mt = "RED", "MISSING", "-"

        # ---- BLOB axis ----
        if args.no_blob:
            blob_ax, blob_txt, bmod = "SKIP", "skipped", "-"
        else:
            bname = item["blob"].format(date=args.date_folder)
            bexists, bsz, bmod = blob_meta(item["container"], bname, key)
            if bexists and bsz >= floor:
                blob_ax, blob_txt = "OK", f"{bsz//1024}KB {bmod}"
            elif bexists:
                blob_ax, blob_txt = "RED", f"{bsz}B small"
            else:
                blob_ax, blob_txt = "RED", "MISSING"

        # ---- WINDOW axis ----
        if cls in ("_frozen", "_regen"):
            win_ax, win_txt = "NA", ("frozen" if cls == "_frozen" else "regen")
        else:
            # judge on the freshest evidence we have (blob mod preferred, else local mtime)
            ev = bmod if (not args.no_blob and bmod != "-") else mt
            iw = in_window(ev, wstart)
            if iw is True:
                win_ax, win_txt = "OK", f"maj({ev[:7]})"
            elif iw is False:
                win_ax, win_txt = "RED", f"STALE({ev[:7]})"
            else:
                win_ax, win_txt = "RED", "undecidable"

        # ---- VERDICT ----
        axes = [local_ax, blob_ax, win_ax]
        if "RED" in axes:
            verdict = "RED"
        elif cls in ("_frozen", "_regen") and local_ax == "OK":
            verdict = "YELLOW" if cls == "_regen" else "GREEN"
        elif all(a in ("OK", "SKIP", "NA") for a in axes):
            verdict = "GREEN"
        else:
            verdict = "YELLOW"

        note = ""
        if local_ax == "RED" and blob_ax == "RED":
            note = "VM-disk? ssh ls ~/bcg/<fam>/output/ (LB.90)"
        elif local_txt.endswith("STUB!"):
            note = "stub in the channel the code reads (LB.86)"
        elif win_ax == "RED":
            note = "left-over pre-window file -- repair before weave (LB.89)"

        log(verdict, label, local_txt, blob_txt, win_txt, note)

    # ---- receipt ----
    reds = [r for r in RESULTS if r["status"] == "RED"]
    yellows = [r for r in RESULTS if r["status"] == "YELLOW"]
    greens = [r for r in RESULTS if r["status"] == "GREEN"]
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        path = RECEIPT_DIR / f"input_provenance_{ts}.xlsx"
        wb = Workbook(); ws = wb.active; ws.title = "input_provenance"
        for k, v in [("Receipt", "input_provenance_probe (3-axis freshness/placement)"),
                     ("Generated (UTC)", ts), ("Window", args.window),
                     ("Date folder", args.date_folder), ("Freshness >=", wstart),
                     ("Verdict", f"GREEN={len(greens)} YELLOW={len(yellows)} RED={len(reds)}"),
                     ("Developer", "Jens Palmö (Senior Business Analyst, Evidensia)")]:
            ws.append([k, v]); ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.append([])
        ws.append(["Verdict", "File", "Local", "Blob", "Window", "Note"])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        for r in RESULTS:
            if r["status"] in ("INFO", "GATE"):
                continue
            ws.append([r["status"], r["label"], r["local"], r["blob"], r["window"], r["note"]])
        for col, w in zip("ABCDEF", (9, 34, 22, 20, 14, 46)):
            ws.column_dimensions[col].width = w
        wb.save(path)
        print(f"[RECEIPT] {path}")
    except Exception as e:
        print(f"[WARN] receipt not written ({type(e).__name__}: {e})")

    print(f"\n=== SUMMARY: GREEN={len(greens)} YELLOW={len(yellows)} RED={len(reds)} ===")
    if reds:
        print("RED inputs would contaminate the weave -- repair BEFORE run_after:")
        for r in reds:
            print(f"  - {r['label']}: {r['note'] or 'see axes above'}")
    return 1 if reds else 0


if __name__ == "__main__":
    sys.exit(main())
