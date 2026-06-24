#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bundle_chain_validator.py  --  end-to-end statisk validering av HELA bundle-kedjan
=================================================================================
Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia). Forfattare: Claude.

SYFTE
-----
EN sond som kartlagger och validerar hela bundle-kedjan i ett svep -- sa vi slutar
rycka i dorrar en i taget. Validerar struktur, dataflode, alla oppna hypoteser, och
hur bundle-steget hanger ihop med foregaende (data_prep) och kommande (modell, steg6)
steg. Kor LOKALT, ingen Ray, ingen VM, ingen DW -- ren statisk + filinspektion.

Detta ar mallen for motsvarande site_chain_validator / cluster_chain_validator
(schema-systemet). Forkroppsligar P.5 (sond fore lager-for-lager) + R7 (mat, gissa ej).

KEDJAN SOM VALIDERAS (bundle, maj-fonstret)
-------------------------------------------
  [data_prep, FOREGAENDE]  Sweden_masterdata.csv (DW-export, maj)
    -> A  convert_masterdata_to_parquet.py        LOKAL   -> sweden_master_data.parquet
    -> B  run_bundle_dataprep.py (00/01/02 SQL)   LOKAL   -> 4 CSV (Raw_Data m.fl.)
    -> C  2.Sweden_..._Model_Data_Creation.py     **VM**  -> Bundle_Clinic_Data.csv (10-kol)
          (+ CSV->xlsx-brygga)                     **VM**  -> bundle_weekly_model_data_clinic_hospital.xlsx
    -> D  run_bundle_model.py (mapp 5, Ray)        VM      -> output_summary.xlsx (~125 KEY)
    -> E  run_all_rationality.py                   LOKAL   -> receipts
  [steg 6, KOMMANDE]  Fall_Back_Logic anvander output_summary

KRITISK LARDOM SONDEN KODIFIERAR (bevisad 2026-06-24 + FAS 18)
--------------------------------------------------------------
Steg C (model-data-creation) anvander Ray (@ray.remote build_bundle_for_type).
Ray KRASCHAR lokalt pa Windows 31 GB ("Windows fatal exception: access violation"
i remote_function.py) -- bevisat tva ggr. => steg C ar ett VM-steg, EJ lokalt.
Den lokala bundle_weekly_model...xlsx (2025-10-03) ar BCG-original, ALDRIG omskriven
lokalt; den vaxande xlsx:en byggdes pa VM. Sonden flaggar om nagon tror C kan kora lokalt.

KOR (global py-3.11, fran repo-roten)
-------------------------------------
    cd "C:\\Projekt\\BCG"
    py -3.11 verify_tool\\probes\\bundle_chain_validator.py

Options:
    --no-receipt    Skriv inte Excel-receipt (bara konsol).
    --vm            Kor aven VM-sidans kontroller (kraver VM uppe + ssh). Default: av.
"""
from __future__ import annotations

import argparse
import datetime
import os
import re
import subprocess
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Sokvagar (Z.0-fakta -- enda stallet de bor)
# ---------------------------------------------------------------------------
REPO = Path(r"C:\Projekt\BCG")
ELAST = REPO / "Pipeline" / "02. Elasticity"

# data_prep (foregaende steg) -- maj-kallan
MASTERDATA_CSV = ELAST / "Sweden_Elasticity_Data_Prep_SQL" / "output" / "Sweden_masterdata.csv"

# Steg A -- convert
CONVERT_SCRIPT = REPO / "tools" / "convert_masterdata_to_parquet.py"
BUNDLE_PARQUET = ELAST / "4. Bundle Clinic Data Prep" / "Sweden_Bundling_Data_Prep" / "parquet" / "sweden_master_data.parquet"

# Steg B -- SQL-dataprep
DATAPREP_RUNNER = REPO / "verify_tool" / "run" / "run_bundle_dataprep.py"
BUNDLING_BASE = ELAST / "4. Bundle Clinic Data Prep" / "Sweden_Bundling_Data_Prep"
SQL_DIR = BUNDLING_BASE / "scripts"
SQL_OUT = BUNDLING_BASE / "output"
B_OUTPUTS = [
    "Raw_Data_Clinic_Hospital.csv",
    "Sweden_Clinic_Hospital_FTE_Data.csv",
    "bundlegroup_bundle_mapping.csv",
    "Bundle_Clinic_Data.csv",
]

# Steg C -- model-data-creation (VM-steg)
MDC_DIR = ELAST / "4. Bundle Clinic Data Prep" / "1.Data_Pre_Processing" / "code"
MDC_SCRIPT = MDC_DIR / "2.Sweden_Bundle_Clinic_Model_Data_Creation.py"
MDC_UTILS = MDC_DIR / "bundle_utils.py"
MDC_CONFIG = MDC_DIR / "src" / "config.yml"
MDC_DATA = ELAST / "4. Bundle Clinic Data Prep" / "1.Data_Pre_Processing" / "data"
MDC_INPUT_CSV = MDC_DATA / "Raw_Data_Clinic_Hospital.csv"
MDC_OUTPUT_CSV = MDC_DATA / "Bundle_Clinic_Data.csv"
MODEL_XLSX = MDC_DATA / "bundle_weekly_model_data_clinic_hospital.xlsx"

# Steg D -- modell
MODEL_RUNNER = REPO / "orchestration" / "runners" / "run_bundle_model.py"
UPLOAD_TOOL = REPO / "orchestration" / "tools" / "upload_input_to_vm.py"
MODEL_LOCAL_OUT = ELAST / "5. Bundle Clinic Models" / "output" / "azure_run_model" / "output_summary.xlsx"

# Steg E -- validering
RATIONALITY = REPO / "verify_tool" / "output_rationality" / "run_all_rationality.py"

# VM-sokvagar (for --vm)
VM_USER = "azureuser"
VM_HOST = os.environ.get("BCG_VM_HOST", "172.18.148.4")
VM_INPUT_XLSX = "/home/azureuser/bcg/bundle/data/bundle_weekly_model_data_clinic_hospital.xlsx"
VM_OUTPUT = "/home/azureuser/bcg/bundle/output/model/output_summary.xlsx"
VM_VENV_PY = "/home/azureuser/bcg/cluster/.venv/bin/python"

STAMP = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# ---------------------------------------------------------------------------
# Resultatuppsamling
# ---------------------------------------------------------------------------
ROWS: list[tuple[str, str, str]] = []  # (status, check, detalj)


def rec(status: str, check: str, detalj: str = "") -> None:
    """status: PASS / FAIL / REVIEW / INFO."""
    ROWS.append((status, check, detalj))
    mark = {"PASS": "[PASS]", "FAIL": "[FAIL]", "REVIEW": "[REVIEW]", "INFO": "[INFO]"}.get(status, status)
    line = f"{mark:9} {check}"
    if detalj:
        line += f"  --  {detalj}"
    print(line, flush=True)


def section(title: str) -> None:
    print("\n" + "=" * 78, flush=True)
    print(title, flush=True)
    print("=" * 78, flush=True)
    ROWS.append(("", f"=== {title} ===", ""))


# ---------------------------------------------------------------------------
# Hjalpare
# ---------------------------------------------------------------------------
def read_text(p: Path) -> str:
    for enc in ("utf-8", "cp1252", "utf-16"):
        try:
            return p.read_text(encoding=enc)
        except (UnicodeDecodeError, FileNotFoundError):
            continue
    return ""


def file_info(p: Path) -> str:
    if not p.exists():
        return "SAKNAS"
    st = p.stat()
    mb = st.st_size / 1024 / 1024
    mt = datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")
    if mb >= 1:
        return f"{mb:.1f} MB, {mt}"
    return f"{st.st_size:,} B, {mt}"


def sniff_encoding(p: Path) -> str:
    """Las forsta bytena: BOM-detektion."""
    if not p.exists():
        return "SAKNAS"
    with open(p, "rb") as fh:
        head = fh.read(4)
    if head.startswith(b"\xff\xfe") or head.startswith(b"\xfe\xff"):
        return "UTF-16 (BOM)"
    if head.startswith(b"\xef\xbb\xbf"):
        return "UTF-8 (BOM)"
    return "UTF-8/ascii (ingen BOM)"


def csv_header(p: Path) -> list[str]:
    txt = read_text(p)
    if not txt:
        return []
    first = txt.splitlines()[0] if txt.splitlines() else ""
    return [c.strip() for c in first.split(",")]


def xlsx_info(p: Path) -> tuple[int, list[str], list[str]]:
    """(rader, header, sheetnames) -- kraver openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return (-1, [], [])
    if not p.exists():
        return (-1, [], [])
    try:
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c) for c in rows[0]] if rows else []
        return (len(rows), hdr, wb.sheetnames)
    except Exception as e:
        return (-2, [str(e)], [])


def grep(p: Path, pattern: str, flags=re.I) -> list[tuple[int, str]]:
    txt = read_text(p)
    hits = []
    for i, line in enumerate(txt.splitlines(), 1):
        if re.search(pattern, line, flags):
            hits.append((i, line.strip()))
    return hits


# ---------------------------------------------------------------------------
# VALIDERINGAR
# ---------------------------------------------------------------------------
def validate_files_exist() -> None:
    section("1. FILINVENTERING -- finns varje lank i kedjan?")
    checks = [
        (MASTERDATA_CSV, "data_prep maj-kalla (Sweden_masterdata.csv)", True),
        (CONVERT_SCRIPT, "A: convert_masterdata_to_parquet.py", True),
        (BUNDLE_PARQUET, "A-output: sweden_master_data.parquet", True),
        (DATAPREP_RUNNER, "B: run_bundle_dataprep.py", True),
        (MDC_SCRIPT, "C: model-data-creation-skript", True),
        (MDC_UTILS, "C: bundle_utils.py", True),
        (MDC_CONFIG, "C: config.yml", True),
        (MODEL_RUNNER, "D: run_bundle_model.py", True),
        (UPLOAD_TOOL, "D: upload_input_to_vm.py", True),
        (RATIONALITY, "E: run_all_rationality.py", True),
    ]
    for p, label, required in checks:
        info = file_info(p)
        if info == "SAKNAS":
            rec("FAIL" if required else "REVIEW", label, "SAKNAS")
        else:
            rec("PASS", label, info)


def validate_step_a() -> None:
    section("2. STEG A -- parquet (har bundle-parqueten maj?)")
    info = file_info(BUNDLE_PARQUET)
    rec("INFO", "bundle-parquet", info)
    # backup-disciplin (LB.24)
    backups = list(BUNDLE_PARQUET.parent.glob("*.april_backup_*")) + \
              list(BUNDLE_PARQUET.parent.glob("*frozen*"))
    if backups:
        rec("PASS", "LB.24 parquet-backup finns", f"{len(backups)} st (t.ex. {backups[0].name})")
    else:
        rec("REVIEW", "LB.24 parquet-backup", "ingen backup hittad -- om parqueten regenererats utan backup ar facit borta")
    # datumtackning via duckdb (om tillganglig)
    try:
        import duckdb
        if BUNDLE_PARQUET.exists():
            psql = str(BUNDLE_PARQUET).replace("\\", "/")
            r = duckdb.connect().execute(
                f"SELECT COUNT(*), MIN(CAST(InvoiceDate AS DATE)), MAX(CAST(InvoiceDate AS DATE)) "
                f"FROM read_parquet('{psql}')"
            ).fetchone()
            rows, mn, mx = r
            rec("INFO", "parquet rader", f"{rows:,}")
            rec("INFO", "parquet datumspann", f"{mn} -> {mx}")
            mxs = str(mx)
            if mxs >= "2026-05-31":
                rec("PASS", "parquet har maj", f"max={mx} (>= 2026-05-31)")
            elif mxs >= "2026-04-30":
                rec("REVIEW", "parquet ar april", f"max={mx} -- maj saknas, kor om steg A")
            else:
                rec("FAIL", "parquet for gammal", f"max={mx}")
    except ImportError:
        rec("INFO", "duckdb saknas", "hoppar over parquet-datumkoll (kor i global py-3.11)")


def validate_step_b() -> None:
    section("3. STEG B -- SQL-dataprep (4 CSV producerade, fonster?)")
    # runner-disciplin: parameterlos, chdir, duckdb-Python (ej exe)
    runner_txt = read_text(DATAPREP_RUNNER)
    if "os.chdir" in runner_txt:
        rec("PASS", "B-runner LB.20 chdir", "satter working dir (relativa SQL-makron)")
    else:
        rec("REVIEW", "B-runner LB.20 chdir", "ingen os.chdir hittad")
    if "duckdb.connect" in runner_txt and "duckdb.exe" not in runner_txt:
        rec("PASS", "B-runner LB.2 duckdb-Python", "anvander Python-API, ej blockerad exe")
    # 4 output-filer
    for fn in B_OUTPUTS:
        p = SQL_OUT / fn
        rec("PASS" if p.exists() else "FAIL", f"B-output {fn}", file_info(p))
    # fonster i Raw_Data
    try:
        import duckdb
        raw = SQL_OUT / "Raw_Data_Clinic_Hospital.csv"
        if raw.exists():
            psql = str(raw).replace("\\", "/")
            r = duckdb.connect().execute(
                f"SELECT COUNT(*), MIN(week_starting_monday), MAX(week_starting_monday) "
                f"FROM read_csv('{psql}', header=true, all_varchar=true)"
            ).fetchone()
            rec("INFO", "Raw_Data rader", f"{r[0]:,}")
            rec("INFO", "Raw_Data veckospann", f"{r[1]} -> {r[2]}")
            if r[0] > 0 and str(r[2]) >= "2026-05":
                rec("PASS", "B-output har maj-veckor", f"max={r[2]}")
            else:
                rec("REVIEW", "B-output fonster", f"max={r[2]}, rader={r[0]}")
    except ImportError:
        pass


def validate_step_c_bridge() -> None:
    section("4. STEG C -- model-data-creation + xlsx-brygga (HYPOTESERNA)")

    # --- H1: Ray-beroende => VM-steg, EJ lokalt ---
    utils_txt = read_text(MDC_UTILS)
    ray_hits = grep(MDC_UTILS, r"@ray\.remote|ray\.init|\.remote\(")
    if ray_hits:
        rec("PASS", "H1 steg C anvander Ray", f"{len(ray_hits)} traffar (build_bundle_for_type m.fl.)")
        rec("FAIL", "H1 steg C KAN EJ koras lokalt", "Ray kraschar pa Windows 31 GB (access violation) -- bevisat 2x. KOR PA VM.")
    else:
        rec("REVIEW", "H1 Ray", "ingen Ray hittad -- ovantat, model-data-creation antas Ray-parallell")
    # num_cpus / object_store (visar att det dimensionerats for stor maskin)
    cpu_hits = grep(MDC_UTILS, r"num_cpus|object_store_memory")
    for ln, txt in cpu_hits:
        rec("INFO", f"Ray-config bundle_utils L{ln}", txt)

    # --- H2: xlsx-bryggan -- skapar nagot skript xlsx:en? ---
    to_excel_mdc = grep(MDC_SCRIPT, r"to_excel|ExcelWriter|\.xlsx")
    to_excel_utils = grep(MDC_UTILS, r"to_excel|ExcelWriter|\.xlsx")
    to_csv_mdc = grep(MDC_SCRIPT, r"to_csv")
    if to_excel_mdc or to_excel_utils:
        rec("PASS", "H2 xlsx skapas av skript", f"to_excel hittad (mdc={len(to_excel_mdc)}, utils={len(to_excel_utils)})")
    else:
        rec("REVIEW", "H2 xlsx-brygga SAKNAS som skript",
            "inget to_excel i mapp 4 -- skriptet skriver bara CSV (to_csv). xlsx byggs PA VM eller via separat steg.")
    if to_csv_mdc:
        for ln, txt in to_csv_mdc:
            rec("INFO", f"C skriver CSV (L{ln})", txt)

    # --- H3: config output vs vad modellen laser ---
    cfg_txt = read_text(MDC_CONFIG)
    out_match = re.search(r"output_data:\s*['\"]?([^'\"#\n]+)", cfg_txt)
    cfg_out = out_match.group(1).strip() if out_match else "?"
    rec("INFO", "config.yml output_data", cfg_out)
    model_reads = grep(MODEL_RUNNER, r"REMOTE_INPUT")
    reads_xlsx = any("xlsx" in t.lower() for _, t in model_reads)
    if "Bundle_Clinic_Data.csv" in cfg_out and reads_xlsx:
        rec("REVIEW", "H3 config/modell-glapp",
            f"config skriver '{cfg_out}' men modellen laser xlsx -- CSV->xlsx-brygga MASTE finnas (VM-steg)")

    # --- H4: lokal xlsx ar BCG-original (aldrig omskriven lokalt) ---
    xinfo = file_info(MODEL_XLSX)
    rec("INFO", "lokal model-xlsx", xinfo)
    if MODEL_XLSX.exists():
        mt = datetime.datetime.fromtimestamp(MODEL_XLSX.stat().st_mtime)
        if mt.year == 2025:
            rec("PASS", "H4 lokal xlsx = BCG-original",
                f"daterad {mt:%Y-%m-%d} -- ALDRIG omskriven lokalt. Bekraftar steg C kors pa VM (xlsx byggs dar).")
        else:
            rec("INFO", "H4 lokal xlsx omskriven", f"daterad {mt:%Y-%m-%d}")
        # header-jamforelse: matchar model-data-creation-output?
        rows, hdr, sheets = xlsx_info(MODEL_XLSX)
        if rows > 0:
            rec("INFO", "lokal xlsx struktur", f"{rows} rader, sheets={sheets}")
            expected = {"Clusters", "week_starting_monday", "Bundle_description", "Bundle_code",
                        "Bundle_visits", "basket_price", "basket_revenue",
                        "bundle_visits_per_site", "num_of_sites", "FTE_Interpolated"}
            if expected.issubset(set(hdr)):
                rec("PASS", "xlsx-header = model-data-creation-output",
                    "10 forvantade kolumner -> xlsx ar C:s output sparad som xlsx")

    # --- H5: namnkonflikt Bundle_Clinic_Data.csv (B vs C skriver samma namn) ---
    b_bcd = SQL_OUT / "Bundle_Clinic_Data.csv"
    c_bcd = MDC_OUTPUT_CSV
    b_hdr = csv_header(b_bcd)[:3]
    c_hdr = csv_header(c_bcd)[:3]
    rec("INFO", "B:s Bundle_Clinic_Data.csv header", str(b_hdr))
    rec("INFO", "C:s Bundle_Clinic_Data.csv header", str(c_hdr))
    if b_hdr and c_hdr and b_hdr != c_hdr:
        rec("REVIEW", "H5 namnkonflikt Bundle_Clinic_Data.csv",
            "B (exploded membership) och C (10-kol model-data) delar filnamn, olika struktur -- C skriver over B. Ofarligt (B-versionen oanvand nedstroms) men forvirrande.")
    elif b_hdr and c_hdr and b_hdr == c_hdr:
        rec("INFO", "Bundle_Clinic_Data.csv", "B och C har samma header (C har skrivit over)")

    # --- H6: tyst tomnings-bugg fixad (LB.75/FD.36)? ---
    fix_hits = grep(MDC_UTILS, r"FD\.36|datetime.*divergens|to_datetime.*slutmerge|Additiv 2026-06-17")
    if fix_hits:
        rec("PASS", "H6 tyst-tomning-fix narvarande", f"{len(fix_hits)} FD.36-additiv markering(ar) i bundle_utils")
    else:
        rec("REVIEW", "H6 tyst-tomning-fix", "ingen FD.36-markering hittad -- verifiera att process_bundles_with_fte-fixen finns")

    # --- C input pa plats (steg B -> C koppling) ---
    rec("PASS" if MDC_INPUT_CSV.exists() else "FAIL",
        "C-input (Raw_Data i mapp 4/data)", file_info(MDC_INPUT_CSV))
    if MDC_INPUT_CSV.exists() and (SQL_OUT / "Raw_Data_Clinic_Hospital.csv").exists():
        same = MDC_INPUT_CSV.stat().st_size == (SQL_OUT / "Raw_Data_Clinic_Hospital.csv").stat().st_size
        rec("PASS" if same else "REVIEW", "B->C koppling (Raw_Data kopierad)",
            "samma storlek som B-output" if same else "storlek skiljer -- kopiera B-output till mapp 4/data")


def validate_step_d() -> None:
    section("5. STEG D -- modell (run_bundle_model kopplingar)")
    txt = read_text(MODEL_RUNNER)
    for key, label in [
        (r'REMOTE_INPUT\s*=\s*["\']([^"\']+)', "D REMOTE_INPUT"),
        (r'REMOTE_OUTPUT\s*=\s*["\']([^"\']+)', "D REMOTE_OUTPUT"),
        (r'REMOTE_PYTHON\s*=\s*["\']([^"\']+)', "D REMOTE_PYTHON"),
        (r'PHASE_KEY\s*=\s*["\']([^"\']+)', "D PHASE_KEY"),
    ]:
        m = re.search(key, txt)
        rec("INFO" if m else "REVIEW", label, m.group(1) if m else "ej hittad")
    # benign step5 (LB.44) hanteras?
    if "data_prep_after_model_output" in txt:
        rec("PASS", "D benign Step5 (LB.44)", "runnern kanner till xlwings-kraschen som vantad")
    # launch-hardning (dagens fix)
    if re.search(r"retries\s*=\s*[1-9]", txt) or "pgrep" in txt:
        rec("PASS", "D launch-hardning", "retry/pgrep-spar (commit e4f0515-klassen)")
    # G7 datumfonster
    if re.search(r"BCG_START_DATE|BCG_END_DATE|start.date|end.date", txt, re.I):
        rec("PASS", "D G7 datumfonster", "tar start/end (env eller arg)")


def validate_step_e_and_step6() -> None:
    section("6. STEG E + STEG 6 -- validering & koppling framat")
    rec("PASS" if RATIONALITY.exists() else "FAIL", "E rationality-svit", file_info(RATIONALITY))
    rec("INFO" if MODEL_LOCAL_OUT.exists() else "REVIEW",
        "D-output lokalt (fran forra korning)", file_info(MODEL_LOCAL_OUT))
    # steg 6 koppling
    step6 = ELAST / "6. Fall Back Logic"
    if step6.exists():
        rec("INFO", "steg 6 (Fall Back Logic) finns", "anvander output_summary fran alla familjer")
        rec("REVIEW", "steg 6 beroende", "kraver Site + Bundle output_summary.xlsx -- blockerad tills bundle klar")


def validate_vm_side() -> None:
    section("7. VM-SIDAN (--vm) -- VM-kontroller")

    def ssh(cmd: str, timeout=60) -> str:
        full = ["ssh", f"{VM_USER}@{VM_HOST}", cmd]
        try:
            cp = subprocess.run(full, capture_output=True, text=True, timeout=timeout)
            return cp.stdout.strip() or cp.stderr.strip()
        except Exception as e:
            return f"SSH-FEL: {e}"

    # ar VM nabar?
    probe = ssh("echo VM_OK")
    if "VM_OK" not in probe:
        rec("REVIEW", "VM nabar", f"ssh svarade ej ({probe[:60]}) -- starta VM, eller hoppa --vm")
        return
    rec("PASS", "VM nabar", "ssh OK")
    # finns model-data-creation pa VM?
    mdc_vm = ssh("find ~/bcg -name '*Model_Data_Creation*' 2>/dev/null")
    rec("INFO" if mdc_vm else "REVIEW", "VM har model-data-creation?", mdc_vm or "hittas ej -- maste laddas upp for steg C")
    # bundle_utils pa VM?
    bu_vm = ssh("find ~/bcg -name 'bundle_utils.py' 2>/dev/null")
    rec("INFO" if bu_vm else "REVIEW", "VM har bundle_utils?", bu_vm or "saknas")
    # Ray i VM-venv?
    ray_vm = ssh(f"{VM_VENV_PY} -c 'import ray; print(ray.__version__)'")
    rec("INFO", "VM venv Ray-version", ray_vm)
    # model-input xlsx pa VM?
    xlsx_vm = ssh(f"stat -c '%y %s' {VM_INPUT_XLSX} 2>/dev/null || echo SAKNAS")
    rec("INFO", "VM model-input xlsx", xlsx_vm)
    # gammal output?
    out_vm = ssh(f"stat -c '%y' {VM_OUTPUT} 2>/dev/null || echo INGEN")
    rec("INFO", "VM tidigare output", out_vm)


# ---------------------------------------------------------------------------
# Receipt
# ---------------------------------------------------------------------------
def write_receipt() -> Path | None:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        print("\n[receipt] openpyxl saknas -- hoppar over Excel-receipt.")
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "bundle_chain"
    mono = Font(name="Consolas", size=10)
    bold = Font(name="Consolas", size=10, bold=True)
    hdr = ["bundle_chain_validator", f"kord {datetime.datetime.now():%Y-%m-%d %H:%M:%S}", "Jens Palmo / Evidensia"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=i, column=1, value=h); c.font = bold
    r = len(hdr) + 2
    ws.cell(row=r, column=1, value="STATUS").font = bold
    ws.cell(row=r, column=2, value="CHECK").font = bold
    ws.cell(row=r, column=3, value="DETALJ").font = bold
    r += 1
    for status, check, detalj in ROWS:
        ws.cell(row=r, column=1, value=status).font = mono
        ws.cell(row=r, column=2, value=check).font = (bold if check.startswith("===") else mono)
        ws.cell(row=r, column=3, value=detalj).font = mono
        r += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 90
    out_dir = REPO / "workspace" / "validation_receipts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"bundle_chain_validator_{STAMP}.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description="End-to-end statisk validering av bundle-kedjan.")
    ap.add_argument("--no-receipt", action="store_true")
    ap.add_argument("--vm", action="store_true", help="Kor aven VM-sidans kontroller (kraver VM uppe).")
    args = ap.parse_args()

    print("=" * 78)
    print("BUNDLE CHAIN VALIDATOR  --  end-to-end statisk kedjevalidering")
    print(f"  {datetime.datetime.now():%Y-%m-%d %H:%M:%S}   repo={REPO}")
    print("=" * 78)

    validate_files_exist()
    validate_step_a()
    validate_step_b()
    validate_step_c_bridge()
    validate_step_d()
    validate_step_e_and_step6()
    if args.vm:
        validate_vm_side()

    # sammanfattning
    section("SAMMANFATTNING")
    n_pass = sum(1 for s, _, _ in ROWS if s == "PASS")
    n_fail = sum(1 for s, _, _ in ROWS if s == "FAIL")
    n_rev = sum(1 for s, _, _ in ROWS if s == "REVIEW")
    print(f"  PASS={n_pass}   FAIL={n_fail}   REVIEW={n_rev}")
    if n_fail:
        print("  -> FAIL finns: lank(ar) brutna eller fel miljo. Atgarda fore korning.")
    else:
        print("  -> Inga FAIL. REVIEW = vantade designnoter (las dem).")

    if not args.no_receipt:
        rp = write_receipt()
        if rp:
            print(f"\n[receipt] {rp}")

    return 1 if n_fail else 0


if __name__ == "__main__":
    sys.exit(main())
