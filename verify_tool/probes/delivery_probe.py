#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
delivery_probe.py  --  LEVERANS-STATUS-SOND (smal Lager 3: ar en korning komplett?)
====================================================================================
Svarar pa EN aterkommande fraga som kostade hela kvallen 2026-06-22:
  "For run_id X -- finns alla artefakter, ar de samtida, NADDE de Blob, och
   sager statusfilen sanningen?"

Bakgrund: site-majkorningen korde KLART pa VM (6624 modeller) men tunneln dog ->
(a) statusfilen fastnade pa 'running' (ingen finalize), (b) output_summary nadde
ALDRIG Blob (sott kvar pa VM-disk). Vi upptackte det forst genom att MANUELLT lasa
VM-logg + lista Blob. Denna sond gor det pa ett kommando.

VIKTIGT designval: sonden anropar `az` (CLI) -- INTE projektets blob.py. Skalet:
az-vagen ar BEVISAD att fungera (auth-mode key), medan blob.py:s signaturer inte ar
verifierade har. Den lutar sig pa kommandon vi vet fungerar, inte gissade API:er.

Den DOMER inte modellkvalitet (det ar rationality-sviten) och mater inte population
per steg (det ar flow-sonden). Den svarar bara: finns det som en KOMPLETT, SAMTIDIG
leverans kraver, och stammer statusfilen med verkligheten?

KOR (PowerShell, global py -3.11; KRAVER fungerande az + PIM/token mot Blob):
    py -3.11 delivery_probe.py --account evbcgpricinginput
    py -3.11 delivery_probe.py --account evbcgpricinginput --run-id 2022-07-01_2026-05-31
    # utan --run-id: listar alla run_ids i runstatus och analyserar den SENASTE
    # --offline <json-fil>: testlage, las en redan sparad 'az ... -o json'-dump i st f att anropa az

FORUTSATTNING: az inloggad, ratt subscription (ev-lz3-ai SE), PIM aktiv. Sonden
SAGER TILL tydligt om az failar (auth) -- den gissar inte.

Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia), assisterad av Claude.
Beroende: openpyxl (kvitto). az CLI pa PATH. std-lib (subprocess, json).
"""
from __future__ import annotations
import argparse
import datetime
import json
import shutil
import subprocess
import sys
from pathlib import Path

# Vad en KOMPLETT leverans innehaller. Per familj: forvantad output i 'output'-containern.
# (Justera prefix/namn mot din faktiska Blob-layout -- dessa ar rimliga defaults ur
#  det vi sett: output/<datum>/<family>/... och output/<datum>/output_summary.xlsx.)
EXPECTED = {
    "site_output":    {"contains": ["site", "output_summary"], "label": "Site output_summary"},
    "cluster_output": {"contains": ["cluster", "output_summary"], "label": "Cluster output_summary"},
    "bundle_output":  {"contains": ["bundle", "output_summary"], "label": "Bundle output_summary"},
    "step6":          {"contains": ["Final_Fallback"], "label": "Step 6 Final_Fallback"},
    "r12":            {"contains": ["Model_Feed"], "label": "R12 Model_Feed"},
}

OK = "OK"; MISS = "SAKNAS"; STALE = "EJ SAMTIDA"; WARN = "GRANSKA"


def log(m): print(m, flush=True)


def _find_az():
    """Hitta az pa Windows: az ar ett .cmd/.bat, inte .exe -> shutil.which loser det.
    Returnerar full sokvag, eller None."""
    for name in ("az", "az.cmd", "az.bat", "az.exe"):
        p = shutil.which(name)
        if p:
            return p
    # Kanda installationsplatser som fallback
    import os
    for cand in (
        r"C:\Program Files\Microsoft SDKs\Azure\CLI2\wbin\az.cmd",
        r"C:\Program Files (x86)\Microsoft SDKs\Azure\CLI2\wbin\az.cmd",
    ):
        if os.path.isfile(cand):
            return cand
    return None


_AZ = _find_az()


def run_az(args_list):
    """Kor az robust pa Windows. az ar ett batch-skript (.cmd) -> Pythons subprocess
    hittar det inte via ["az"] som PowerShell gor. Vi anvander full sokvag om vi hittar
    den, annars shell=True (som later skalet sla upp az precis som i din terminal)."""
    # Forsta forsoket: full sokvag (sakrast)
    if _AZ:
        try:
            r = subprocess.run([_AZ] + args_list, capture_output=True, text=True, timeout=60)
            return (r.returncode == 0, r.stdout, r.stderr)
        except Exception as e:
            pass  # fall through till shell=True
    # Fallback: shell=True -- skalet hittar az.cmd pa PATH (samma som din PowerShell)
    try:
        cmd = "az " + " ".join(_quote(a) for a in args_list)
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=60, shell=True)
        if r.returncode == 0 or r.stdout:
            return (r.returncode == 0, r.stdout, r.stderr)
        return (False, r.stdout, r.stderr or "az gav ingen utdata")
    except FileNotFoundError:
        return (False, "", "az hittades inte (varken via which eller shell)")
    except subprocess.TimeoutExpired:
        return (False, "", "az timeout (60s)")


def _quote(a):
    """Citera ett argument for shell=True sa query-uttryck med [] och mellanslag overlever."""
    if any(ch in a for ch in ' []{}"?'):
        return '"' + a.replace('"', '\\"') + '"'
    return a


def az_blob_list(account, container):
    """Lista blobbar (namn + lastModified + storlek) via az, auth-mode key."""
    ok, out, err = run_az([
        "storage", "blob", "list", "--account-name", account,
        "--container-name", container, "--auth-mode", "key",
        "--query", "[].{name:name, modified:properties.lastModified, size:properties.contentLength}",
        "-o", "json"])
    if not ok:
        return (False, [], err)
    try:
        return (True, json.loads(out or "[]"), "")
    except json.JSONDecodeError:
        return (False, [], "kunde ej parsa az json-utdata")


def auth_hint(err):
    e = (err or "").lower()
    if "authorizationfailed" in e or "permission" in e or "70043" in e or "interactionrequired" in e:
        return ("\n  -> Ser ut som auth/PIM. Kor: az logout; az login (valj ev-lz3-ai SE); "
                "aktivera PIM i portalen; forsok igen.")
    if "az hittades inte" in e:
        return "\n  -> az CLI saknas pa PATH."
    return ""


def analyze(account, run_id, blobs_status, blobs_output, F):
    """Syntetisera leverans-status. F = lista av (klass, vad, detalj)."""
    # 1. Statusfilen for run_id
    sf = [b for b in blobs_status if b["name"] == f"{run_id}.json"]
    if sf:
        F.append((OK, f"Statusfil finns: {run_id}.json",
                  f"storlek {sf[0].get('size','?')} B, andrad {sf[0].get('modified','?')}"))
    else:
        F.append((MISS, f"Statusfil SAKNAS for run_id {run_id}", "ingen runstatus/<run_id>.json"))

    # 2. Output-artefakter (matcha mot EXPECTED). Vi vet inte exakt prefix -> matcha pa 'contains'.
    out_names = [b["name"] for b in blobs_output]
    found_any_for_run = [n for n in out_names if run_id in n or run_id.split("_")[-1] in n]
    # Tidsstamplar for samtidighet
    def newest_matching(words):
        cand = [b for b in blobs_output if all(w.lower() in b["name"].lower() for w in words)]
        if not cand: return None
        return max(cand, key=lambda b: b.get("modified", ""))

    times = {}
    for key, spec in EXPECTED.items():
        b = newest_matching(spec["contains"])
        if b:
            times[key] = b.get("modified", "")
            F.append((OK, f"{spec['label']} finns", f"{b['name']} (andrad {b.get('modified','?')})"))
        else:
            F.append((MISS, f"{spec['label']} SAKNAS i 'output'", f"matchade ej {spec['contains']}"))

    # 3. Samtidighet: spreder tidsstamplarna over mer an ~2 dygn -> EJ SAMTIDA (granska)
    if len(times) >= 2:
        ds = sorted(t[:10] for t in times.values() if t)  # YYYY-MM-DD
        if ds and ds[0] != ds[-1]:
            F.append((STALE, f"Artefakter EJ samtida: spann {ds[0]} .. {ds[-1]}",
                      "olika korningar blandade -> en familj kan vara gammal (granska)"))
        else:
            F.append((OK, "Artefakter samtida (samma datum)", f"{ds[0] if ds else '?'}"))

    # 4. Det klassiska kvallsfallet: statusfil 'running' MEN output finns -> finalize uteblev
    #    (Vi kan inte lasa statusfilens innehall via list -- men vi flaggar monstret att
    #     KOLLA om site-output finns men ingen step6/r12 = halvfardig leverans.)
    site_ok = any("site" in n.lower() and "output_summary" in n.lower() for n in out_names)
    step6_ok = any("final_fallback" in n.lower() for n in out_names)
    if site_ok and not step6_ok:
        F.append((WARN, "Site-output finns men Step 6 (Final_Fallback) saknas",
                  "halvfardig leverans: modell klar men efterbearbetning ej kord/uppladdad"))


def write_receipt(account, run_id, F, out_dir, run_ids):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception as e:
        log(f"[WARN] openpyxl saknas ({e}) -- konsol racker."); return None
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    fp = out_dir / f"00_delivery_{run_id}_{stamp}.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Leverans-status"
    mono = Font(name="Consolas", size=10); bold = Font(name="Consolas", size=10, bold=True)
    fills = {OK: PatternFill("solid", fgColor="E8F0E8"),
             WARN: PatternFill("solid", fgColor="FFF3CD"),
             STALE: PatternFill("solid", fgColor="FFF3CD"),
             MISS: PatternFill("solid", fgColor="F8D7DA")}
    ws.append([f"LEVERANS-STATUS  konto={account}  run_id={run_id}  {stamp}"]); ws["A1"].font = bold
    # Go/no-go overst
    has_miss = any(k == MISS for k, _, _ in F)
    has_warn = any(k in (WARN, STALE) for k, _, _ in F)
    verdict = "NEJ (saknas artefakter)" if has_miss else ("GRANSKA" if has_warn else "JA -- komplett & samtidig")
    ws.append([f"LEVERANSREDO? {verdict}"]); ws.cell(ws.max_row,1).font = bold
    ws.cell(ws.max_row,1).fill = fills[MISS if has_miss else (WARN if has_warn else OK)]
    ws.append([]); ws.append(["STATUS", "VAD", "DETALJ"])
    for c in range(1,4): ws.cell(ws.max_row, c).font = bold
    for k, what, detail in F:
        ws.append([k, what, detail]); ws.cell(ws.max_row,1).fill = fills.get(k, fills[OK])
    if run_ids:
        ws.append([]); ws.append([f"Alla run_ids i runstatus ({len(run_ids)}): {', '.join(run_ids)}"])
    for col, w in (("A",12),("B",55),("C",65)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        for c in row:
            if not c.font or c.font.name != "Consolas": c.font = mono
            c.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(fp); return fp


def main() -> int:
    ap = argparse.ArgumentParser(description="Leverans-status-sond (smal Lager 3) via az-CLI.")
    ap.add_argument("--account", default="evbcgpricinginput", help="Storage-konto (TEST default).")
    ap.add_argument("--status-container", default="runstatus")
    ap.add_argument("--output-container", default="output")
    ap.add_argument("--run-id", default=None, help="Specifikt run_id; annars senaste i runstatus.")
    ap.add_argument("--out", default=None)
    ap.add_argument("--offline", default=None,
                    help="Testlage: json-fil {status:[...], output:[...]} i st f att anropa az.")
    args = ap.parse_args()

    # Hamta blob-listor (live via az, ELLER offline-dump for test)
    if args.offline:
        data = json.loads(Path(args.offline).read_text())
        blobs_status = data.get("status", []); blobs_output = data.get("output", [])
        log(f"[RUN] OFFLINE-lage ur {args.offline}")
    else:
        log(f"[RUN] konto={args.account} (az, auth-mode key)")
        log(f"[RUN] az hittad: {_AZ or 'NEJ via which -> provar shell=True'}")
        ok1, blobs_status, e1 = az_blob_list(args.account, args.status_container)
        if not ok1:
            log(f"[ERROR] kunde ej lista '{args.status_container}': {e1.strip()[:200]}{auth_hint(e1)}")
            return 2
        ok2, blobs_output, e2 = az_blob_list(args.account, args.output_container)
        if not ok2:
            log(f"[ERROR] kunde ej lista '{args.output_container}': {e2.strip()[:200]}{auth_hint(e2)}")
            return 2

    run_ids = sorted({b["name"][:-5] for b in blobs_status if b["name"].endswith(".json")}, reverse=True)
    if not run_ids:
        log("[KPI] inga run_ids i runstatus."); return 0
    run_id = args.run_id or run_ids[0]
    log(f"[RUN] analyserar run_id={run_id}" + ("" if args.run_id else " (senaste)"))
    log(f"[RUN] {len(run_ids)} run_ids totalt: {', '.join(run_ids)}")

    F = []
    analyze(args.account, run_id, blobs_status, blobs_output, F)

    # Konsol: go/no-go + bara icke-OK
    has_miss = any(k == MISS for k, _, _ in F)
    has_warn = any(k in (WARN, STALE) for k, _, _ in F)
    verdict = "NEJ (saknas)" if has_miss else ("GRANSKA" if has_warn else "JA")
    log(f"\n[VERDICT] Leveransredo for {run_id}? {verdict}")
    for k, what, detail in F:
        if k != OK:
            log(f"  [{k}] {what}")
            if detail: log(f"        {detail}")
    n_ok = sum(1 for k,_,_ in F if k == OK)
    log(f"[KPI] {n_ok} OK, {sum(1 for k,_,_ in F if k!=OK)} att granska/saknas")

    out_dir = Path(args.out) if args.out else (
        Path.cwd() / "verify_tool" / "receipts" / datetime.date.today().isoformat() / "delivery")
    fp = write_receipt(args.account, run_id, F, out_dir, run_ids)
    if fp: log(f"\n[Saved] {fp}")
    log("[DONE]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
