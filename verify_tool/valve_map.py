"""
valve_map.py  --  ventilkarta: var sitter ventilerna, hur mycket slapper de, VAD rann ut
=========================================================================================
Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia Djursjukvard AB).
Forfattare: Claude advisor.

VAD DETTA AR (din idé, inte en larmgrind)
-----------------------------------------
Inte conservation ("lacker det?"). Detta ar en FLODESMATARE vid varje VENTIL. Varje
skarv har avsiktliga ventiler -- stallen dar rader/entiteter LAMNAR roret MED AVSIKT
(Fee-filter, signifikanstrosklar, inner joins, min-sites). Du VET att de finns och
SKA finnas. Det som saknats: en exakt karta over VAR varje ventil sitter, HUR MYCKET
den slapper, och VAD som rann ut -- sa du kan titta pa utflodet och bekrafta "ja, ratt
sorts vatten lattades av har". Baslinjen blir guld den dag en ventil borjar slappa
FEL sorts vatten (lacka dar det inte ska lacka).

Detta ar FLOW ACCOUNTING / DROP ATTRIBUTION: varje rad som lamnar roret tillskrivs
en NAMNGIVEN ventil med ett SKAL. Det gor bortfallet GRANSKNINGSBART.

VENTILERNA I STEP 6-VAVEN (belagda rad-for-rad i Fall_Back_Logic.py)
--------------------------------------------------------------------
  V1  Fee-filter        (rad ~102/630): service == 'Fee' lattas av
  V2  service-join      (rad ~252): inner join pa ProductKey -> produkter utan
                                    service-match faller ur (Natrium Catalyst-ventilen)
  V3  signifikansport   (rad ~377/408): significant == 1 kravs (RSQ>=.5, PVALUE<=.20,
                                    -10 < elast < 0) -- HUVUDVENTILEN, lattar av allt
                                    insignifikant fran fallback-bidraget
  V4  min-sites-port    (rad ~658): SigSites_Sum >= 10 for site-elasticitet

VARJE VENTIL MATS SOM: in -> ut -> avlättat (antal + andel) + ett PROV pa vad som
rann ut (de forsta N entiteterna + skalet), sa du SER vattnet, inte bara mangden.

AERLIG TACKNING (kalla fore pastaende)
--------------------------------------
Detta matet ventilerna i kod jag HAR (Fall_Back_Logic.py + Constant.py). Ventiler i
replicate_dataprep.py (skarv 2) och modellstegens signifikansfiltrering (skarv 3, pa
VM) sitter i kod jag inte sett -- de ar listade i VALVE_REGISTRY som EJ_MATT med exakt
var de borde matas, redo att byggas nar respektive kalla laddas upp. Ingen gissning.

SCHEMA-VARNING (belagt 2026-06-26): parquetens RADATA-kolumner ar DW-namn (ID_Item,
ID_Customer, ID_Patient), INTE de nedstroms-namn Constant.py anvander (ItemCode...).
Omdopningen sker i vaven. valve_map matar pa de namn som galler DAR ventilen sitter.

KOR (global py-3.11, repo-roten)
--------------------------------
    py -3.11 verify_tool\\valve_map.py                 # mat vaven + skriv karta (md) + protokoll (xlsx)
    py -3.11 verify_tool\\valve_map.py --no-receipt    # bara karta + stdout
    py -3.11 verify_tool\\valve_map.py --registry      # lista alla kanda ventiler (matta + ej matta)
"""
from __future__ import annotations

import argparse
import datetime
import sys
from dataclasses import dataclass, field
from pathlib import Path

REPO = Path(r"C:\Projekt\BCG")
ELAST = REPO / "Pipeline" / "02. Elasticity"
FBL = ELAST / "6. Fall Back Logic"
VERIFY = REPO / "verify_tool"
STAMP = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# Step 6 input-kallor (samma som Constant.py / run_step6 placerar)
CLUSTER_MODEL = ELAST / "2. Product Cluster Level Models" / "output" / "output_summary_ready.xlsx"
CLUSTER_GRAN = ELAST / "2. Product Cluster Level Models" / "output" / "final_model_cluster_granularity.xlsx"
SITE_MODEL = ELAST / "3. Product Site Level Models" / "output" / "model" / "output_summary.xlsx"
BUNDLE_MODEL = ELAST / "5. Bundle Clinic Models" / "output" / "model" / "output_summary.xlsx"
ALL_PRODUCT = FBL / "input_data" / "Complete_Product_Data.xlsx"


# ---------------------------------------------------------------------------
# VENTILREGISTER -- alla kanda ventiler i hela roret. matt=True byggs nu;
# matt=False ar belagd plats redo att byggas nar kallan finns.
# ---------------------------------------------------------------------------
@dataclass
class Valve:
    vid: str
    skarv: str
    where: str          # var i koden (fil:rad-region)
    rule: str           # villkoret som slapper ut
    purpose: str        # VARFOR ventilen finns (avsikten)
    measured: bool = False
    confidence: str = "EXAKT"   # EXAKT | APPROX | EJ -- far en osaker siffra aldrig
                                # lasas som baslinje. APPROX = matningen replikerar
                                # vavens logik UNGEFARLIGT; kalibrera mot vaven fore
                                # den litas pa (se kalibreringsblock nederst i filen).


VALVE_REGISTRY: list[Valve] = [
    Valve("V1", "step6", "Fall_Back_Logic.py ~102/630", "service == 'Fee'",
          "Fee-tjanster ar ej priselastiska produkter -- ska ej ha elasticitet",
          measured=True, confidence="EXAKT"),
    Valve("V2", "step6", "Fall_Back_Logic.py ~252", "inner join pa ProductKey (service_map)",
          "produkter utan service-match kan ej fallback-vavas (t.ex. icke-prissatt Internal)",
          measured=True, confidence="APPROX"),   # in=682 misstankt lagt; nyckelmatchning osaker -- kalibrera
    Valve("V3", "step6", "Fall_Back_Logic.py ~377/408", "significant == 1 (RSQ>=.5 & PVALUE<=.20 & -10<elast<0)",
          "bara statistiskt sakra modeller far bidra till fallback-elasticiteten",
          measured=True, confidence="EXAKT"),
    Valve("V4", "step6", "Fall_Back_Logic.py ~658", "SigSites_Sum >= 10",
          "site-elasticitet kraver minst 10 signifikanta sites for robusthet",
          measured=True, confidence="APPROX"),   # 99.6% extremt -- aggregering approx; kalibrera mot vaven
    # --- ej matta (belagd plats, kraver kalla) ---
    Valve("P1", "dataprep", "replicate_dataprep.py (ej sedd)", "datumfonster-filter (_inject_dates)",
          "bara rader inom BCG_START..BCG_END -- G7-fonstret", measured=False, confidence="EJ"),
    Valve("P2", "dataprep", "replicate_dataprep.py (ej sedd)", "aggregering produkt x vecka",
          "transaktionsrader -> modellgranularitet (radantal krymper avsiktligt)", measured=False, confidence="EJ"),
    Valve("M1", "model_vm", "feature_selection/model.py (VM, ej sedd)", "min-observationer for regression",
          "produkter med for fa datapunkter far ingen elasticitet", measured=False, confidence="EJ"),
]


@dataclass
class ValveReading:
    vid: str
    rows_in: int = 0
    rows_out: int = 0
    sample_out: list = field(default_factory=list)   # prov pa vad som rann ut
    note: str = ""

    @property
    def released(self) -> int:
        return self.rows_in - self.rows_out

    @property
    def release_pct(self) -> float:
        return (self.released / self.rows_in * 100) if self.rows_in else 0.0


READINGS: list[ValveReading] = []


def _p(msg=""):
    print(msg, flush=True)


def measure_step6_valves() -> list[ValveReading]:
    """Mat V1-V4 mot de faktiska Step 6-input-filerna. Replikerar ventilernas
    villkor EXAKT som Fall_Back_Logic gor -- men matet bara, andrar inget."""
    import pandas as pd
    out: list[ValveReading] = []

    # --- V1: Fee-filter pa df_all_product ---
    if ALL_PRODUCT.exists():
        df = pd.read_excel(ALL_PRODUCT)
        r = ValveReading("V1", rows_in=len(df))
        if "ProductGroupL4Name" in df.columns:
            cat = df["ProductGroupL4Name"].astype(str).str.lower().str.strip()
            fee_mask = cat == "fee"
            kept = df[~fee_mask]
            r.rows_out = len(kept)
            fee_rows = df[fee_mask]
            if "ProductGroupL4Name" in fee_rows.columns:
                r.sample_out = [{"ProductGroupL4Name": str(v), "n": int(c)}
                                for v, c in fee_rows["ProductGroupL4Name"].value_counts().head(5).items()]
            r.note = "Fee-tjanster avlättade fore vav"
        else:
            r.rows_out = len(df)
            r.note = "ingen ProductGroupL4Name-kolumn (kunde ej mata)"
        out.append(r)

    # --- V2: service-join (inner) -- produkter utan service-match ---
    # Approx: hur manga distinkta ProductKeys i cluster-modellen som SAKNAS i
    # df_all_product:s service-map (= skulle falla ur inner join).
    if CLUSTER_MODEL.exists() and ALL_PRODUCT.exists():
        dfc = pd.read_excel(CLUSTER_MODEL)
        dfp = pd.read_excel(ALL_PRODUCT)
        r = ValveReading("V2")
        # cluster-modellens nyckel: ItemCode (ur KEY-split) eller ItemCode-kolumn
        if "ItemCode" in dfc.columns:
            ckeys = set(dfc["ItemCode"].dropna().astype(str))
        elif "KEY" in dfc.columns:
            ckeys = set(dfc["KEY"].astype(str).str.rsplit("-", n=1).str[-1])
        else:
            ckeys = set()
        pkeys = set(dfp["ItemCode"].dropna().astype(str)) if "ItemCode" in dfp.columns else set()
        r.rows_in = len(ckeys)
        matched = ckeys & pkeys
        r.rows_out = len(matched)
        unmatched = list(ckeys - pkeys)[:10]
        r.sample_out = [{"ProductKey_utan_servicematch": k} for k in unmatched]
        r.note = "distinkta cluster-ProductKeys; utflode = saknar service-map-match (inner join droppar)"
        out.append(r)

    # --- V3: signifikansport pa cluster-modellen (huvudventilen) ---
    if CLUSTER_MODEL.exists():
        dfc = pd.read_excel(CLUSTER_MODEL)
        r = ValveReading("V3", rows_in=len(dfc))
        need = {"RSQ", "PVALUE_Regular_Price_fwbw_max_6", "ELASTICITY_Regular_Price_fwbw_max_6"}
        # tillat aven redan omdopta namn
        rsq = "RSQ" if "RSQ" in dfc.columns else None
        pval = ("PVALUE_Regular_Price_fwbw_max_6" if "PVALUE_Regular_Price_fwbw_max_6" in dfc.columns
                else ("PVALUE_PRICE" if "PVALUE_PRICE" in dfc.columns else None))
        elas = ("ELASTICITY_Regular_Price_fwbw_max_6" if "ELASTICITY_Regular_Price_fwbw_max_6" in dfc.columns
                else ("ELASTICITY_PRICE" if "ELASTICITY_PRICE" in dfc.columns else None))
        if rsq and pval and elas:
            R = pd.to_numeric(dfc[rsq], errors="coerce").round(2)
            P = pd.to_numeric(dfc[pval], errors="coerce").round(2)
            E = pd.to_numeric(dfc[elas], errors="coerce")
            sig = (R >= 0.5) & (P <= 0.20) & (E < 0) & (E > -10)
            r.rows_out = int(sig.sum())
            # vad rann ut: bryt ner VARFOR de var insignifikanta
            insig = dfc[~sig]
            reasons = {
                "RSQ < 0.5": int((R < 0.5).sum()),
                "PVALUE > 0.20": int((P > 0.20).sum()),
                "elast >= 0 (fel tecken)": int((E >= 0).sum()),
                "elast <= -10 (orimlig)": int((E <= -10).sum()),
            }
            r.sample_out = [{"skal": k, "antal_rader": v} for k, v in reasons.items() if v > 0]
            r.note = "insignifikanta modeller avlättade (bidrar ej till fallback) -- huvudventilen"
        else:
            r.rows_out = len(dfc)
            r.note = f"saknar kolumn for signifikanstest (rsq={rsq}, pval={pval}, elas={elas})"
        out.append(r)

    # --- V4: min-sites-port pa site-modellen ---
    if SITE_MODEL.exists():
        dfs = pd.read_excel(SITE_MODEL)
        r = ValveReading("V4", rows_in=len(dfs))
        # V4 verkar pa AGGREGERAD niva (SigSites_Sum>=10 per ProductKey). Approx:
        # mat hur manga distinkta produkter som har >=10 signifikanta sites.
        # Site-modellen har KEY (Cluster-ItemCode); signifikans via samma trosklar.
        need_cols = {"RSQ", "ELASTICITY_Regular_Price_fwbw_max_6"}
        pval = ("PVALUE_Regular_Price_fwbw_max_6" if "PVALUE_Regular_Price_fwbw_max_6" in dfs.columns
                else ("PVALUE_PRICE" if "PVALUE_PRICE" in dfs.columns else None))
        elas = ("ELASTICITY_Regular_Price_fwbw_max_6" if "ELASTICITY_Regular_Price_fwbw_max_6" in dfs.columns
                else ("ELASTICITY_PRICE" if "ELASTICITY_PRICE" in dfs.columns else None))
        if "KEY" in dfs.columns and "RSQ" in dfs.columns and pval and elas:
            R = pd.to_numeric(dfs["RSQ"], errors="coerce").round(2)
            P = pd.to_numeric(dfs[pval], errors="coerce").round(2)
            E = pd.to_numeric(dfs[elas], errors="coerce")
            sig = ((R >= 0.5) & (P <= 0.20) & (E < 0) & (E > -10)).astype(int)
            pk = dfs["KEY"].astype(str).str.rsplit("-", n=1).str[-1]
            per_product = sig.groupby(pk).sum()
            r.rows_in = int(per_product.shape[0])           # distinkta produkter
            passed = per_product[per_product >= 10]
            r.rows_out = int(passed.shape[0])
            below = per_product[(per_product > 0) & (per_product < 10)]
            r.sample_out = [{"produkter_med_1_till_9_sig_sites": int(below.shape[0])},
                            {"produkter_med_0_sig_sites": int((per_product == 0).sum())}]
            r.note = "produkter med >=10 signifikanta sites slapps; faerre avlättas (robusthetskrav)"
        else:
            r.rows_out = len(dfs)
            r.note = "saknar kolumner for min-sites-matning"
        out.append(r)

    return out


def render_valve_map(readings: list[ValveReading]) -> Path | None:
    """Visuell rorkarta (markdown) -- VAR ventilerna sitter + hur mycket de slapper."""
    by_id = {r.vid: r for r in readings}
    L: list[str] = []
    L.append("# Ventilkarta -- Step 6-vaven (flodesmatning per ventil)")
    L.append("")
    L.append(f"_Genererad {datetime.datetime.now():%Y-%m-%d %H:%M} av valve_map.py. "
             f"Utvecklare: Jens Palmo. Matet pa faktiska Step 6-input-filer._")
    L.append("")
    L.append("Varje ventil ar ett AVSIKTLIGT stalle dar rader lamnar roret. Detta visar "
             "var de sitter, hur mycket de slapper, och vad som rann ut. En baslinje: "
             "nar en ventils utflode AVVIKER fran detta har roret borjat bete sig annorlunda.")
    L.append("")
    L.append("```")
    L.append("  PARQUET (radata, DW-namn: ID_Item...)")
    L.append("     |")
    L.append("     |  [ skarv 1-2: dataprep-ventiler -- EJ MATTA, se register ]")
    L.append("     v")
    L.append("  MODELL-OUTPUT per familj (cluster/site/bundle)")
    L.append("     |")
    L.append("     v")
    for vid in ("V1", "V2", "V3", "V4"):
        v = next((x for x in VALVE_REGISTRY if x.vid == vid), None)
        r = by_id.get(vid)
        if not v:
            continue
        cflag = "" if v.confidence == "EXAKT" else f" [{v.confidence}]"
        if r and r.rows_in:
            bar_len = int(r.release_pct / 5)  # 1 tecken per 5%
            bar = "#" * bar_len + "-" * (20 - bar_len)
            L.append(f"  +--[ {vid}{cflag} ]--{v.rule[:42]}")
            L.append(f"  |   in={r.rows_in:>7,}  ut={r.rows_out:>7,}  "
                     f"avlättat={r.released:>6,} ({r.release_pct:4.1f}%)")
            L.append(f"  |   slapper: [{bar}]")
        else:
            L.append(f"  +--[ {vid}{cflag} ]--{v.rule[:42]}  (ej matt denna korning)")
        L.append("  |")
    L.append("  v")
    L.append("  FINAL_FALLBACK (vavd elasticitet per ProductKey)")
    L.append("```")
    L.append("")

    # Detalj per ventil
    L.append("## Vad rann ut ur varje ventil")
    L.append("")
    for vid in ("V1", "V2", "V3", "V4"):
        v = next((x for x in VALVE_REGISTRY if x.vid == vid), None)
        r = by_id.get(vid)
        if not v:
            continue
        conf_badge = {"EXAKT": "**[VERIFIERAD]**",
                      "APPROX": "**[APPROXIMATION -- kalibrera fore baslinje]**",
                      "EJ": "**[EJ MATT]**"}.get(v.confidence, "")
        L.append(f"### {vid} -- {v.purpose}  {conf_badge}")
        L.append(f"- **Var:** `{v.where}`")
        L.append(f"- **Regel:** `{v.rule}`")
        if v.confidence == "APPROX":
            L.append(f"- **OBS:** denna siffra REPLIKERAR vavens logik ungefarligt och "
                     f"kan vara fel. Lita INTE pa den som baslinje forran den kalibrerats "
                     f"mot vavens faktiska tal (se kalibreringsblock i valve_map.py-huvudet).")
        if r and r.rows_in:
            L.append(f"- **Floede:** in {r.rows_in:,} -> ut {r.rows_out:,} "
                     f"(**avlättat {r.released:,}, {r.release_pct:.1f}%**)")
            if r.sample_out:
                L.append(f"- **Utflode (prov):**")
                for s in r.sample_out:
                    L.append(f"    - {s}")
            L.append(f"- _{r.note}_")
        else:
            L.append("- _ej matt denna korning_")
        L.append("")

    # Ej matta ventiler
    L.append("## Ej matta ventiler (belagd plats, kraver kalla)")
    L.append("")
    L.append("| Ventil | Skarv | Var | Regel | Avsikt |")
    L.append("|--------|-------|-----|-------|--------|")
    for v in VALVE_REGISTRY:
        if not v.measured:
            L.append(f"| {v.vid} | {v.skarv} | {v.where} | {v.rule} | {v.purpose} |")
    L.append("")
    L.append("_Ladda upp replicate_dataprep.py (skarv 1-2) + ett modellsteg (skarv 3) "
             "for att mata dessa exakt -- ingen gissning._")

    out_dir = VERIFY / "valve_map"
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        out = out_dir / f"valve_map_{STAMP}.md"
        out.write_text("\n".join(L), encoding="utf-8")
        return out
    except Exception as e:  # noqa: BLE001
        _p(f"[karta] kunde ej skriva: {e}")
        return None


def write_receipt(readings: list[ValveReading]) -> Path | None:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        _p("[receipt] openpyxl saknas -- hoppar Excel-protokoll.")
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "valve_map"
    bold = Font(name="Consolas", size=10, bold=True)
    mono = Font(name="Consolas", size=10)
    hdr = ["valve_map", f"kord {datetime.datetime.now():%Y-%m-%d %H:%M:%S}", "Jens Palmo / Evidensia"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=i, column=1, value=h).font = bold
    r0 = len(hdr) + 2
    cols = ["Ventil", "Skarv", "Var", "Regel", "in", "ut", "avlättat", "andel%", "utflode_prov", "avsikt"]
    for j, c in enumerate(cols, 1):
        ws.cell(row=r0, column=j, value=c).font = bold
    reg = {v.vid: v for v in VALVE_REGISTRY}
    rr = r0 + 1
    for rd in readings:
        v = reg.get(rd.vid)
        vals = [rd.vid, v.skarv if v else "", v.where if v else "", v.rule if v else "",
                rd.rows_in, rd.rows_out, rd.released, round(rd.release_pct, 1),
                "; ".join(str(s) for s in rd.sample_out), v.purpose if v else ""]
        for j, val in enumerate(vals, 1):
            ws.cell(row=rr, column=j, value=val).font = mono
        rr += 1
    for col, w in zip("ABCDEFGHIJ", [8, 10, 34, 44, 9, 9, 9, 8, 50, 50]):
        ws.column_dimensions[col].width = w
    out_dir = VERIFY / "valve_map"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"valve_map_{STAMP}.xlsx"
    wb.save(out)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Ventilkarta: var sitter ventilerna, hur mycket slapper de, vad rann ut.")
    ap.add_argument("--no-receipt", action="store_true", help="Skriv inte Excel-protokoll.")
    ap.add_argument("--registry", action="store_true", help="Lista alla kanda ventiler och avsluta.")
    args = ap.parse_args()

    print("=" * 72)
    print("VALVE MAP  --  var sitter ventilerna, hur mycket slapper de, vad rann ut?")
    print("=" * 72)

    if args.registry:
        for v in VALVE_REGISTRY:
            mark = "MATT " if v.measured else "ej   "
            print(f"  [{mark}] {v.vid} ({v.skarv}): {v.rule}")
            print(f"            var: {v.where}")
            print(f"            avsikt: {v.purpose}")
        return 0

    try:
        readings = measure_step6_valves()
    except Exception as e:  # noqa: BLE001
        print(f"  matfel: {type(e).__name__}: {e}")
        return 1
    READINGS.extend(readings)

    print("\n-- Ventilflode (Step 6-vaven) --")
    for r in readings:
        v = next((x for x in VALVE_REGISTRY if x.vid == r.vid), None)
        conf = v.confidence if v else "?"
        flag = "" if conf == "EXAKT" else f"   <-- {conf}: kalibrera fore baslinje (se filhuvud)"
        print(f"\n  [{r.vid}] ({conf}) {v.rule if v else ''}{flag}")
        print(f"        in={r.rows_in:,}  ut={r.rows_out:,}  "
              f"avlättat={r.released:,} ({r.release_pct:.1f}%)")
        if r.sample_out:
            print(f"        utflode: {r.sample_out}")
        print(f"        {r.note}")

    karta = render_valve_map(readings)
    if karta:
        print(f"\n[karta]   {karta}")
    if not args.no_receipt:
        rp = write_receipt(readings)
        if rp:
            print(f"[receipt] {rp}")

    print("\n" + "-" * 72)
    print("Ventilkartan ar en BASLINJE. Nar en ventils utflode avviker fran detta")
    print("pa oforandrad data -> roret har borjat lacka dar det inte ska. Felsoknings-")
    print("startpunkt: vilken ventil andrade sig.")
    return 0


if __name__ == "__main__":
    sys.exit(main())


# ===========================================================================
# KALIBRERINGSBLOCK -- gor V2/V4 EXAKTA mot vavens faktiska tal
# ===========================================================================
# V2 och V4 ar APPROXIMATIONER: de replikerar vavens logik ungefarligt utifran
# input-filerna, men vaven gor merge/aggregering pa satt som kan skilja. Forsta
# korningen gav misstankta tal (V2 in=682 oforklarligt lagt; V4 99.6% extremt).
#
# For att kalibrera: klistra in dessa print-satser TEMPORART i Fall_Back_Logic.py
# (de ANDRAR INGEN LOGIK -- bara mater), kor vaven en gang, och jamfor mot vad
# valve_map rapporterar. Ta bort dem efterat (eller lat dem sta -- de ar ofarliga).
#
# --- V2: service-join (Fall_Back_Logic.read_blended_model_data, ~rad 252) ---
# FORE raden  `dfcluster = dfcluster.merge(service_map, on=ProductKey)`:
#     print(f"[V2-kal] dfcluster fore join: {len(dfcluster)} rader, "
#           f"{dfcluster[ProductKey].nunique()} distinkta ProductKey")
#     print(f"[V2-kal] service_map: {len(service_map)} rader, "
#           f"{service_map[ProductKey].nunique()} distinkta ProductKey")
# EFTER raden:
#     print(f"[V2-kal] dfcluster efter join: {len(dfcluster)} rader, "
#           f"{dfcluster[ProductKey].nunique()} distinkta ProductKey "
#           f"-> avlättat {len(service_map)-len(dfcluster)} (om positivt)")
#
# --- V4: min-sites (Fall_Back_Logic main, ~rad 656-658) ---
# EFTER  `dfsite_temp = dfsite.groupby(ProductKey).agg(...)`:
#     print(f"[V4-kal] dfsite_temp: {len(dfsite_temp)} distinkta ProductKey")
#     print(f"[V4-kal] med SigSites_Sum>=10: "
#           f"{(dfsite_temp['SigSites_Sum']>=10).sum()} produkter")
#     print(f"[V4-kal] fordelning SigSites_Sum:\n"
#           f"{dfsite_temp['SigSites_Sum'].value_counts().sort_index().head(15)}")
#
# Nar du har vavens faktiska tal: justera measure_step6_valves() sa V2/V4 matchar,
# och andra confidence fran APPROX till EXAKT i VALVE_REGISTRY. Da ar hela kartan
# en palitlig baslinje -- spike-to-harden: approximationen var stallningen, det
# kalibrerade matet ar byggnaden.
# ===========================================================================
