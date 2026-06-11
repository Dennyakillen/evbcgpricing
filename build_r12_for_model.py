"""
build_r12_for_model.py  --  fyll BCG-modellens blå flikar med växt data
========================================================================
Producerar de tre drivande kolumnerna för BCG-prismodellens FACT_CodeClinic-flik,
på växt data med samma R12-fönsterlängd som BCG men framflyttat slutdatum:

  FACT_Quant_25  = SUM(SoldQuantity) per ItemCode x Cluster, senaste 12 mån
  FACT_Sales_25  = SUM(SalesTotal)   per ItemCode x Cluster, senaste 12 mån  (brutto, IB.5)
  FACT_Elasticity / _R2 / _pValue = ur Step 6-väven (Final_Fallback_Data)

INGA beräkningar i modellen ändras -- detta producerar bara DATA att klistra in i
den blå fliken. Modellens egna formler (Calculations, Pricing Model) räknar om
omsättningseffekten utifrån de inklistrade värdena + prisantagandet.

R12-FÖNSTER (kärnan i uppgiften):
  BCG:s FACT_Quant_25 var R12 t.o.m. deras sista slutdatum (2025-06).
  Vår motsvarighet = R12 t.o.m. senaste kompletta månad i DW (t.ex. 2026-04).
  Samma 12-månaders längd, framflyttat -- så talen är jämförbara i karaktär,
  bara på färskare fönster.

KÄLLOR (vad skriptet läser):
  1. Transaktionsdata med ItemCode, Cluster, månad/vecka, SoldQuantity, SalesTotal.
     Detta är samma DW-extrakt som export_b4b_for_model.py bygger (växande CSV).
  2. Final_Fallback_Data_<datum>.xlsx (Step 6-output) för elasticiteten.

OBS GRAIN: modellens KEY = ItemCode x Cluster (IB.8). Transaktionsdatan kan ha
finare grain (site) -- vi aggregerar upp till kod x kluster.

Run (PowerShell):
    cd "C:\\Projekt\\BCG"
    py -3.11 build_r12_for_model.py
    py -3.11 build_r12_for_model.py --tx "<sokvag till transaction-csv>" --end 2026-04

Utdata: output_model_feed\\FACT_CodeClinic_feed_<datum>.xlsx  (klistras i bla fliken)

Developer: Jens Palmö (Senior Business Analyst, Evidensia Djursjukvård AB)
Author: Claude advisor, 2026-06-11.
"""
import argparse
import glob
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

BCG_ROOT = Path(r"C:\Projekt\BCG")
FBL = BCG_ROOT / "Pipeline" / "02. Elasticity" / "6. Fall Back Logic"
OUT = BCG_ROOT / "output_model_feed"

# Kandidat-sokvagar for transaktionsdatan (vaxande extrakt). Override med --tx.
TX_CANDIDATES = [
    BCG_ROOT / "Pipeline" / "01. Data Prep" / "output" / "*growing*.csv",
    BCG_ROOT / "Pipeline" / "01. Data Prep" / "output" / "*P_C*.csv",
    BCG_ROOT / "output_b4b" / "*.csv",
]

# Kolumnnamn vi forsoker kanna igen (DW-extraktet kan variera nagot)
COL_ITEMCODE = ["ItemCode", "FACT_ItemCode", "Item_Code", "Code"]
COL_CLUSTER = ["Cluster", "Cluster_Granularity", "New_Cluster", "FACT_Cluster", "Clusters"]
COL_SITE = ["SiteCode", "Cluster", "FACT_ID_Department", "Site", "DepartmentID", "ID_Department"]
COL_QTY = ["SoldQuantity", "QuantitySold", "SoldQuantity(SalesTotal>0)", "Quant", "Quantity"]
COL_SALES = ["SalesTotal", "TotalNet", "Sales", "Net"]
COL_DATE = ["week_starting_monday", "WeekStarting", "Week", "Date", "Month", "InvoiceMonth"]


def _pick(df, candidates, label):
    for c in candidates:
        if c in df.columns:
            return c
    raise KeyError(f"Hittar ingen {label}-kolumn. Provade {candidates}. "
                   f"Faktiska kolumner: {list(df.columns)[:20]}")


def _find_tx(override):
    if override:
        return Path(override)
    for pat in TX_CANDIDATES:
        hits = glob.glob(str(pat))
        if hits:
            return Path(max(hits, key=os.path.getmtime))
    return None


def _find_fallback(override):
    if override:
        return Path(override)
    cands = glob.glob(str(FBL / "output_data" / "Final_Fallback_Data*.xlsx")) + \
            glob.glob(str(FBL / "Final_Fallback_Data*.xlsx"))
    return Path(max(cands, key=os.path.getmtime)) if cands else None


def build_r12(tx_path, end_month, fallback_path):
    print("=" * 68)
    print("BUILD R12 FOR MODEL  -  fyll blå flikar med växt data")
    print("=" * 68)

    # ---- 1) transaktionsdata -> R12 per kod x site ----
    print(f"[LÄS] transaktionsdata: {tx_path.name}")
    tx = pd.read_csv(tx_path, sep=None, engine="python")
    ic = _pick(tx, COL_ITEMCODE, "ItemCode")
    site = _pick(tx, COL_SITE, "SiteCode/Cluster(site-ID)")
    qty = _pick(tx, COL_QTY, "SoldQuantity")
    sal = _pick(tx, COL_SALES, "SalesTotal")
    dt = _pick(tx, COL_DATE, "datum")
    print(f"  kolumner: kod={ic}  site={site}  volym={qty}  oms={sal}  datum={dt}")

    tx[dt] = pd.to_datetime(tx[dt], errors="coerce")
    tx[qty] = pd.to_numeric(tx[qty], errors="coerce")
    tx[sal] = pd.to_numeric(tx[sal], errors="coerce")
    tx[ic] = tx[ic].astype(str).str.strip()
    tx[site] = tx[site].astype(str).str.strip()
    tx = tx.dropna(subset=[dt])

    if end_month:
        end = pd.Timestamp(end_month) + pd.offsets.MonthEnd(0)
        print(f"[R12] slutmånad (angiven): {end.date()}")
    else:
        # senaste KOMPLETTA månad i datan = elasticitetens slutdatum (samma extrakt matar båda)
        last = tx[dt].max()
        end = (last.replace(day=1) - pd.Timedelta(days=1))  # sista dagen föregående månad
        if last.day >= 28:  # om datan når månadens slut, använd den månaden
            end = last + pd.offsets.MonthEnd(0)
        print(f"[R12] slutmånad (auto = elasticitetens slutdatum): {end.date()}  (senaste i data: {last.date()})")
    start = (end - pd.DateOffset(months=12)).normalize() + pd.Timedelta(days=1)
    print(f"[R12] fönster: {start.date()} → {end.date()}  (12 mån; samma längd som BCG, framflyttat)")

    win = tx[(tx[dt] >= start) & (tx[dt] <= end)].copy()
    print(f"  rader i fönstret: {len(win):,} av {len(tx):,}")

    r12 = win.groupby([ic, site], dropna=False).agg(
        FACT_Quant_25=(qty, "sum"),
        FACT_Sales_25=(sal, "sum"),
    ).reset_index().rename(columns={ic: "FACT_ItemCode", site: "FACT_SiteCode"})
    print(f"  R12 per kod×site: {len(r12):,} rader")

    # ---- 2) elasticitet ur Step 6 (per ItemCode x SiteCode) ----
    print(f"[LÄS] elasticitet: {fallback_path.name}")
    fb = pd.read_excel(fallback_path)
    pk_col = "ProductKey" if "ProductKey" in fb.columns else None
    site_col = "SiteCode" if "SiteCode" in fb.columns else None
    fe_col = "final_elasticity" if "final_elasticity" in fb.columns else None
    clus_col = "Clusters" if "Clusters" in fb.columns else None
    if not (pk_col and site_col and fe_col):
        print("  VARNING: saknar ProductKey/SiteCode/final_elasticity — hoppar elasticitet.")
        elas = pd.DataFrame(columns=["FACT_ItemCode", "FACT_SiteCode", "FACT_Elasticity"])
    else:
        keep = [c for c in [pk_col, site_col, clus_col, fe_col, "RSQ", "PVALUE_PRICE"] if c and c in fb.columns]
        ren = {pk_col: "FACT_ItemCode", site_col: "FACT_SiteCode", fe_col: "FACT_Elasticity"}
        if clus_col:
            ren[clus_col] = "FACT_Cluster"
        if "RSQ" in keep:
            ren["RSQ"] = "FACT_Elasticity_R2"
        if "PVALUE_PRICE" in keep:
            ren["PVALUE_PRICE"] = "FACT_Elasticity_pValue"
        elas = fb[keep].rename(columns=ren)
        elas["FACT_ItemCode"] = elas["FACT_ItemCode"].astype(str).str.strip()
        elas["FACT_SiteCode"] = elas["FACT_SiteCode"].astype(str).str.strip()
        # en rad per kod×site (Step 6 kan ha dubletter) -> ta första/median
        num = [c for c in ["FACT_Elasticity", "FACT_Elasticity_R2", "FACT_Elasticity_pValue"] if c in elas.columns]
        agg = {c: "median" for c in num}
        if "FACT_Cluster" in elas.columns:
            agg["FACT_Cluster"] = "first"
        elas = elas.groupby(["FACT_ItemCode", "FACT_SiteCode"], dropna=False).agg(agg).reset_index()
    print(f"  elasticitet per kod×site: {len(elas):,} rader")

    # ---- 3) join på ItemCode x SiteCode (Step 6 bär bryggan namn<->site) ----
    feed = r12.merge(elas, on=["FACT_ItemCode", "FACT_SiteCode"], how="left")
    feed["FACT_Elasticity"] = pd.to_numeric(feed.get("FACT_Elasticity"), errors="coerce")

    # ---- rapport ----
    n_with_elas = feed["FACT_Elasticity"].notna().sum()
    print()
    print("--- RESULTAT ---")
    print(f"  rader (kod×site)            : {len(feed):,}")
    print(f"  med elasticitet (matchad)   : {n_with_elas:,} ({100*n_with_elas/len(feed):.1f}%)")
    print(f"  saknar elasticitet          : {len(feed)-n_with_elas:,}  (ny kod/site ej i väven)")
    print(f"  Σ R12 volym                 : {feed['FACT_Quant_25'].sum():,.0f}")
    print(f"  Σ R12 omsättning (brutto)   : {feed['FACT_Sales_25'].sum():,.0f}")
    print()
    print("  Övriga blå-flik-kolumner (pris, konkurrens, FTE, band, _MANUAL) fylls från")
    print("  andra källor — se Model_Update_Guide. Modellens formelflikar rör du INTE.")

    # bygg FACT_CodeClinicKey (ItemCode-SiteCode, modellens nyckel) + ordna kolumner
    feed.insert(0, "FACT_CodeClinicKey", feed["FACT_ItemCode"] + "-" + feed["FACT_SiteCode"])
    col_order = ["FACT_CodeClinicKey", "FACT_ItemCode", "FACT_SiteCode"]
    if "FACT_Cluster" in feed.columns:
        col_order.append("FACT_Cluster")
    col_order += [c for c in ["FACT_Quant_25", "FACT_Sales_25", "FACT_Elasticity",
                              "FACT_Elasticity_R2", "FACT_Elasticity_pValue"] if c in feed.columns]
    feed = feed[col_order]

    # ---- DIM_Code: per ItemCode (R12 oms/volym totalt + hospital/clinic-split) ----
    # hospital/clinic-split kräver site-typ; härled från Step 6 SiteType om möjligt,
    # annars lämna split tom. R12 totalt per kod fyller vi alltid.
    code = win.groupby(ic, dropna=False).agg(
        DIM_Sales_2025=(sal, "sum"),
        DIM_Quant_2025=(qty, "sum"),
    ).reset_index().rename(columns={ic: "DIM_ItemCode"})
    code["DIM_ItemCode"] = code["DIM_ItemCode"].astype(str).str.strip()
    code["DIM_CodePrefix"] = code["DIM_ItemCode"].str.extract(r"^([A-Za-z]+)")
    # hospital/clinic-split via Step 6 SiteType om den finns
    if "SiteType" in fb.columns and site_col:
        st = fb[[pk_col, "SiteType"]].copy()
        st[pk_col] = st[pk_col].astype(str).str.strip()
        sitetype_map = dict(zip(st[pk_col], st["SiteType"]))  # grov: per kod
    # split kräver site×typ -> gör enkel: join win med en typ per site om tillgängligt (annars tom)
    code["DIM_Quant_2025_Hospitals"] = ""   # fylls om site-typ-källa finns (extern/DIM_Site)
    code["DIM_Quant_2025_Clinics"] = ""
    code["DIM_InvoiceGroupDescription"] = ""  # masterdata (service) — fyll från Step 6 service om vill
    if "service" in fb.columns:
        svc = fb[[pk_col, "service"]].copy()
        svc[pk_col] = svc[pk_col].astype(str).str.strip()
        svc_map = svc.dropna().drop_duplicates(pk_col).set_index(pk_col)["service"].to_dict()
        code["DIM_InvoiceGroupDescription"] = code["DIM_ItemCode"].map(svc_map).fillna("")
    code = code[["DIM_ItemCode", "DIM_CodePrefix", "DIM_InvoiceGroupDescription",
                 "DIM_Sales_2025", "DIM_Quant_2025",
                 "DIM_Quant_2025_Hospitals", "DIM_Quant_2025_Clinics"]]

    # ---- DIM_Site: per site-ID (kluster + R12-volym per site) ----
    sitedf = win.groupby(site, dropna=False).agg(
        _R12_qty=(qty, "sum"),
    ).reset_index().rename(columns={site: "DIM_ID_Department"})
    sitedf["DIM_ID_Department"] = sitedf["DIM_ID_Department"].astype(str).str.strip()
    # klusternamn per site från Step 6 (SiteCode -> Clusters)
    if site_col and clus_col:
        sc = fb[[site_col, clus_col]].copy()
        sc[site_col] = sc[site_col].astype(str).str.strip()
        clus_per_site = sc.dropna().drop_duplicates(site_col).set_index(site_col)[clus_col].to_dict()
        sitedf["DIM_Cluster"] = sitedf["DIM_ID_Department"].map(clus_per_site).fillna("")
    else:
        sitedf["DIM_Cluster"] = ""
    sitedf["DIM_SiteType"] = sitedf["DIM_Cluster"].apply(
        lambda c: "Hospital" if "jukhus" in str(c) or "ospital" in str(c) else ("Clinic" if c else ""))
    sitedf = sitedf[["DIM_ID_Department", "DIM_Cluster", "DIM_SiteType"]]

    return {"FACT_CodeClinic": feed, "DIM_Code": code, "DIM_Site": sitedf}


def _write_copypaste(sheets, out_path):
    """En Excel, en flik per destinationsflik i modellen (FACT_CodeClinic, DIM_Code,
    DIM_Site), hårdkodade värden (inga formler) -- ren copy-paste in i BCG-modellen.
    Tomma kolumner = fält som fylls från extern källa (markeras med gul rubrik)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    wb = Workbook()
    thin = Side(style="thin", color="BFBFBF")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill("solid", fgColor="1F487C")
    amber = PatternFill("solid", fgColor="FFF2CC")

    first = True
    for sheet_name, df in sheets.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=j, value=col)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10); c.border = bord
            # gul rubrik om kolumnen är tom (fylls externt)
            is_empty = df[col].astype(str).str.strip().eq("").all() if len(df) else True
            c.fill = amber if is_empty else navy
            if is_empty:
                c.font = Font(name="Arial", bold=True, color="9C5700", size=10)
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            for j, col in enumerate(df.columns, start=1):
                v = row[col]
                if isinstance(v, float):
                    v = round(v, 6)
                c = ws.cell(row=i, column=j, value=v)
                c.font = Font(name="Arial", size=10); c.border = bord
        for j, col in enumerate(df.columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = max(14, len(str(col)) + 2)
        ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tx", default=None, help="sökväg till transaction-CSV (annars auto)")
    ap.add_argument("--fallback", default=None, help="sökväg till Final_Fallback_Data (annars senaste)")
    ap.add_argument("--end", default=None, help="R12 slutmånad YYYY-MM (annars senaste i datan)")
    args = ap.parse_args()

    tx = _find_tx(args.tx)
    fb = _find_fallback(args.fallback)
    if not tx or not tx.exists():
        print("FEL: hittar ingen transaktions-CSV. Ange med --tx <sökväg>.")
        print(f"  provade: {[str(p) for p in TX_CANDIDATES]}")
        return 1
    if not fb or not fb.exists():
        print("FEL: hittar ingen Final_Fallback_Data. Kör run_step6.py först eller ange --fallback.")
        return 1

    sheets = build_r12(tx, args.end, fb)

    OUT.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d")
    out_path = OUT / f"Model_Feed_{stamp}.xlsx"
    _write_copypaste(sheets, out_path)
    print()
    print("--- FLIKAR I OUTPUT ---")
    for name, df in sheets.items():
        empty_cols = [c for c in df.columns if (df[c].astype(str).str.strip().eq("").all() if len(df) else True)]
        print(f"  {name:<18} {len(df):>7,} rader  ({len(df.columns)} kol; {len(empty_cols)} tomma=externa)")
    print()
    print(f"[KLART] sparat: {out_path}")
    print("  Tre flikar = modellens tre blå indataflikar. Gul rubrik = fyll från extern källa.")
    print("  Markera kolumner per flik, kopiera, klistra in i modellens motsvarande flik.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
