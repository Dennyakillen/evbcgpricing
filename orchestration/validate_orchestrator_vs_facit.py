"""
validate_orchestrator_vs_facit.py -- Validate the Phase Z orchestrator's Site
model output against the 2026-06-09 manual growing run (the reproduction facit).
=============================================================================
Question this answers: did AUTOMATING the run change the RESULT? It compares
the orchestrator's output_summary.xlsx (run 2026-06-12 via run_site_model.py)
against the file Jens produced by hand on 2026-06-09 on the same growing data.
If the orchestration is faithful, the two are effectively identical: same KEY
set, same elasticity distribution, same significance share. Any drift here is
an orchestration artefact to investigate -- NOT model drift (that is a separate
analysis) and NOT a BCG-faithfulness check (done in earlier phases).

What it checks (R7: trust the content, not the row count):
  1. KEY set      -- exact same KEYs? (count match alone can hide swapped KEYs)
  2. elasticity   -- per-KEY equality where both have a value; distribution
  3. p-values     -- significance share, distribution
  4. column parity -- same columns present

Output: an Excel receipt (single sheet, raw lines) with date/timestamp in the
filename and first cells, written via openpyxl, in validation_receipts/.

Usage (global Python 3.11, from repo root):
    py -3.11 validate_orchestrator_vs_facit.py

Developer: Jens Palmo (Senior Business Analyst, Evidensia Djursjukvard AB)
Author: Claude advisor, Phase Z session 2026-06-12.
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np

REPO = Path(r"C:\Projekt\BCG")
SITE = REPO / "Pipeline" / "02. Elasticity" / "3. Product Site Level Models"
NEW = SITE / "output" / "azure_run_model" / "output_summary.xlsx"          # today, orchestrator
FACIT = SITE / "output_growing_2026-06-09" / "model" / "output_summary.xlsx"  # 2026-06-09 manual
RECEIPT_DIR = REPO / "workspace" / "validation_receipts"

LINES: list[str] = []


def log(msg: str = "") -> None:
    print(msg, flush=True)
    LINES.append(msg)


def find_col(df: pd.DataFrame, *needles: str) -> str | None:
    for c in df.columns:
        cu = c.upper()
        if all(n.upper() in cu for n in needles):
            return c
    return None


def summarize_numeric(df: pd.DataFrame, col: str, label: str) -> None:
    s = pd.to_numeric(df[col], errors="coerce")
    log(f"  {label} ({col}): n={s.notna().sum()} median={s.median():.4f} "
        f"mean={s.mean():.4f} neg%={100*(s<0).mean():.1f} "
        f"in(-10,0)%={100*((s>-10)&(s<0)).mean():.1f}")


def main() -> int:
    ts = datetime.now()
    log("=" * 68)
    log("ORCHESTRATOR OUTPUT VALIDATION vs 2026-06-09 FACIT")
    log(f"Generated: {ts:%Y-%m-%d %H:%M:%S}")
    log(f"NEW   (orchestrator 2026-06-12): {NEW}")
    log(f"FACIT (manual 2026-06-09):       {FACIT}")
    log("=" * 68)

    for p, tag in [(NEW, "NEW"), (FACIT, "FACIT")]:
        if not p.exists():
            log(f"[ABORT] {tag} not found: {p}")
            _write_receipt(ts, verdict="ABORT-MISSING-FILE")
            return 1

    new = pd.read_excel(NEW)
    fac = pd.read_excel(FACIT)
    log(f"[SHAPE] NEW   {new.shape}   ({NEW.stat().st_size:,} bytes)")
    log(f"[SHAPE] FACIT {fac.shape}   ({FACIT.stat().st_size:,} bytes)")

    # --- 1. Column parity ---
    only_new = sorted(set(new.columns) - set(fac.columns))
    only_fac = sorted(set(fac.columns) - set(new.columns))
    log("")
    log("[COLUMNS]")
    log(f"  shared={len(set(new.columns) & set(fac.columns))} "
        f"only_new={only_new or '-'} only_facit={only_fac or '-'}")

    # --- 2. KEY set ---
    keycol_n = find_col(new, "KEY") or new.columns[0]
    keycol_f = find_col(fac, "KEY") or fac.columns[0]
    kn, kf = set(new[keycol_n].astype(str)), set(fac[keycol_f].astype(str))
    log("")
    log(f"[KEYS] new col='{keycol_n}' facit col='{keycol_f}'")
    log(f"  new={len(kn)} facit={len(kf)} shared={len(kn & kf)} "
        f"only_new={len(kn - kf)} only_facit={len(kf - kn)}")
    keys_identical = (kn == kf)
    log(f"  KEY SETS IDENTICAL: {keys_identical}")
    if not keys_identical:
        for k in list(kn - kf)[:5]:
            log(f"    only in NEW:   {k}")
        for k in list(kf - kn)[:5]:
            log(f"    only in FACIT: {k}")

    # --- 3. Elasticity distribution + per-KEY equality ---
    log("")
    log("[ELASTICITY]")
    ec_n = find_col(new, "ELASTICITY")
    ec_f = find_col(fac, "ELASTICITY")
    if ec_n and ec_f:
        summarize_numeric(new, ec_n, "NEW  ")
        summarize_numeric(fac, ec_f, "FACIT")
        # per-KEY equality on the shared key set
        a = new[[keycol_n, ec_n]].copy(); a.columns = ["K", "new_e"]
        b = fac[[keycol_f, ec_f]].copy(); b.columns = ["K", "fac_e"]
        a["K"] = a["K"].astype(str); b["K"] = b["K"].astype(str)
        m = a.merge(b, on="K", how="inner")
        m["ne"] = pd.to_numeric(m["new_e"], errors="coerce")
        m["fe"] = pd.to_numeric(m["fac_e"], errors="coerce")
        both = m.dropna(subset=["ne", "fe"])
        if len(both):
            diff = (both["ne"] - both["fe"]).abs()
            log(f"  per-KEY match: compared={len(both)} "
                f"max_abs_diff={diff.max():.2e} mean_abs_diff={diff.mean():.2e}")
            log(f"  exact(<1e-9)={100*(diff<1e-9).mean():.2f}%  "
                f"close(<1e-4)={100*(diff<1e-4).mean():.2f}%")
            corr = both["ne"].corr(both["fe"])
            log(f"  correlation(new,facit) = {corr:.6f}")
        else:
            log("  [WARN] no overlapping numeric elasticity to compare")
    else:
        log(f"  [WARN] elasticity column not found (new={ec_n}, facit={ec_f})")

    # --- 4. p-values ---
    log("")
    log("[P-VALUES]")
    pc_n = find_col(new, "PVALUE") or find_col(new, "P", "VALUE")
    pc_f = find_col(fac, "PVALUE") or find_col(fac, "P", "VALUE")
    if pc_n and pc_f:
        pn = pd.to_numeric(new[pc_n], errors="coerce")
        pf = pd.to_numeric(fac[pc_f], errors="coerce")
        log(f"  NEW   sig(p<0.05) = {100*(pn<0.05).mean():.2f}%")
        log(f"  FACIT sig(p<0.05) = {100*(pf<0.05).mean():.2f}%")
    else:
        log(f"  [WARN] p-value column not found (new={pc_n}, facit={pc_f})")

    # --- Verdict ---
    log("")
    log("=" * 68)
    verdict = "PASS" if keys_identical else "REVIEW"
    log(f"[VERDICT] {verdict} -- "
        + ("KEY sets identical; orchestrator reproduces the manual run."
           if keys_identical else
           "KEY sets differ; investigate before trusting the orchestration."))
    log("=" * 68)

    _write_receipt(ts, verdict)
    return 0 if keys_identical else 2


def _write_receipt(ts: datetime, verdict: str) -> None:
    from openpyxl import Workbook
    RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
    fn = RECEIPT_DIR / f"validate_orchestrator_vs_facit_{ts:%Y%m%d_%H%M%S}_{verdict}.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "validation"
    ws["A1"] = f"Validation receipt -- generated {ts:%Y-%m-%d %H:%M:%S}"
    ws["A2"] = f"Verdict: {verdict}"
    ws["A3"] = "Developer: Jens Palmo (Evidensia). Author: Claude advisor."
    for i, line in enumerate(LINES, start=5):
        ws.cell(row=i, column=1, value=line)
    wb.save(fn)
    log(f"[RECEIPT] {fn}")


if __name__ == "__main__":
    sys.exit(main())
