"""
app.py -- Phase Z status dashboard (READ-ONLY, local Flask app)
================================================================
Local web view of the orchestrator's Blob status files. Shows what is (or
was) running, per-phase progress, durations, notes, data-info, errors/hints,
output locations, AND per-family validation results (read from the local
verify_tool/receipts). Nothing is invented client-side; the page renders only
what the status file and the existing validation receipts contain.

READ-ONLY by construction: imports only read functions; reads Blob status and
local validation receipts. It cannot write status, upload, start/stop the VM,
or trigger runs. Download routes stream existing files -- still read-only.

Pattern: reuses the proven Blob layer (blob.py) -- no duplicated Azure code.
Validation drill-down reads the EXISTING master receipts the verify suites
already produced (verify_tool/receipts/<date>/...), no re-running.

Prerequisites (one-time):
    py -3.11 -m pip install flask
    az login --scope https://management.core.windows.net//.default

Run:
    py -3.11 orchestration\\webapp\\app.py            (serves http://127.0.0.1:5000)
    py -3.11 orchestration\\webapp\\app.py --check    (verify Blob reachable, exit)
    py -3.11 orchestration\\webapp\\app.py --port 5050

Developer: Jens Palmö (Senior Business Analyst)
Author: Claude advisor, Phase Z (grows iteratively).

FD.33-B1 (2026-07-03): window-aware receipts. The run_id IS the data window
(2022-07-01_2026-05-31). Receipt routes accept ?run=<run_id>; when it is
window-shaped the app reads receipts/<suite>/<window>/ in Blob (populated by
the FD.33 Etapp A migration) and falls back to the local date folders only
when Blob is unreachable or the run_id is legacy-shaped. Blob receipt paths
are marked with the '_blob/' prefix and served by the same /api/receipt route.
Still READ-ONLY by construction (list/download only).
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from pathlib import Path

# Same bootstrap as the runners.
ORCH = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ORCH / "shared"))
sys.path.insert(0, str(ORCH / "infrastructure"))

# Documented debt: AAD data role is ABAC-blocked -> account-key mode.
os.environ.setdefault("PRICINGMODEL_AUTH", "key")

from flask import Flask, jsonify, render_template, Response, abort, send_file, request  # noqa: E402
from azure.core.exceptions import ResourceNotFoundError  # noqa: E402
from blob import read_status, list_runs  # noqa: E402  (READ-ONLY imports)
from story_config import STORY, BAGE_SV, GROUPS, VALIDATORS, PROOF_CHAIN, FUNNEL, FACIT_WINDOW, OUTPUT_PURPOSE  # noqa: E402

log = logging.getLogger("webapp")
app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False  # svenska tecken ska inte ASCII-escapas

# Lokal sokvag till valideringskvittona (verify_tool/receipts/).
REPO_ROOT = ORCH.parent
RECEIPTS_DIR = REPO_ROOT / "verify_tool" / "receipts"

# Mappning: fas-key -> vilken master-receipt-typ validerar den fasen.
# (Etapp 2: extraction klar. cluster/site/step6 fylls pa i etapp 3.)
PHASE_RECEIPT = {
    "extraction":    {"glob": "00_master_summary_*.xlsx",    "subdir": "",            "suite": "Extraction"},
    "cluster_model": {"glob": "00_rationality_master_*.xlsx", "subdir": "rationality", "suite": "Rationality (output)"},
    "site_model":    {"glob": "00_rationality_master_*.xlsx", "subdir": "rationality", "suite": "Rationality (output)"},
    "bundle_model":  {"glob": "00_rationality_master_*.xlsx", "subdir": "rationality", "suite": "Rationality (output)"},
    "step6":         {"glob": "00_provenance_master_*.xlsx",  "subdir": "provenance",  "suite": "Provenance"},
    "build_r12":     {"glob": "00_provenance_master_*.xlsx",  "subdir": "provenance",  "suite": "Provenance (R12)"},
    "site_step5":    {"glob": "00_rationality_master_*.xlsx", "subdir": "rationality", "suite": "Rationality (Step 5)"},
}

# ------------------------------------------------------------------ FD.33-B1
# Fonster-medvetna kvitton. run_id AR fonstret sedan FD.33. Blob
# receipts/<suite>/<window>/ ar primar kalla; lokala datum-mappar ar FALLBACK
# (Blob onaabart eller legacy run_id). '_blob/'-prefix i kvitto-vagar =>
# /api/receipt serverar ur receipts-containern (JS orord).
RE_WINDOW = re.compile(r"^\d{4}-\d{2}-\d{2}_\d{4}-\d{2}-\d{2}$")
BLOB_MARK = "_blob/"
_SUITE_BY_SUBDIR = {"": "extraction", "rationality": "rationality", "provenance": "provenance"}


def _win(run_id):
    """run_id -> fonster-id om det har fonsterform, annars None (legacy)."""
    return run_id if run_id and RE_WINDOW.match(run_id) else None


def _blob_receipts(suite: str, window: str):
    """Blob-kvitton for suite/fonster (nyast forst) eller None om onaabart.
    None => lokal fallback. TOM lista => Blob naddes men fonstret saknar
    kvitton -- arligt tomt, vi visar ALDRIG fel fonsters kvitton i stallet."""
    try:
        from blob import list_receipts
        return list_receipts(suite, window)
    except Exception as e:
        log.warning("Blob-kvitton onaabara (%s) -- lokal fallback.", str(e)[:120])
        return None


def _blob_receipt_bytes(suite: str, window: str, name: str) -> bytes:
    from blob import _client, CONTAINER_RECEIPTS
    bc = _client().get_container_client(CONTAINER_RECEIPTS).get_blob_client(
        f"{suite}/{window}/{name}")
    return bc.download_blob().readall()


def _xlsx_lines_from_bytes(data: bytes) -> list:
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    return [row[0] for row in ws.iter_rows(values_only=True)]


@app.after_request
def _utf8(resp):
    """Deklarera UTF-8 pa HTML sa webblasaren inte gissar teckenkodning."""
    ct = resp.headers.get("Content-Type", "")
    if ct.startswith("text/html") and "charset" not in ct:
        resp.headers["Content-Type"] = "text/html; charset=utf-8"
    return resp


@app.get("/")
def index():
    return render_template("dashboard.html")


@app.get("/api/story")
def api_story():
    """Statisk facit-referens + texter + grupper + validator-förklaringar (story_config.py)."""
    return jsonify({"story": STORY, "groups": GROUPS, "bage": BAGE_SV,
                    "validators": VALIDATORS, "proof_chain": PROOF_CHAIN,
                    "funnel": FUNNEL, "output_purpose": OUTPUT_PURPOSE,
                    "data_window": _data_window(_win(request.args.get("run", "")))})


@app.get("/api/runs")
def api_runs():
    """All run_ids that have a status file, newest first."""
    try:
        ids = sorted(list_runs(), reverse=True)
        return jsonify({"runs": ids})
    except Exception as e:
        log.exception("list_runs failed")
        return jsonify({"error": _friendly(e)}), 502


@app.get("/api/status/<run_id>")
def api_status(run_id: str):
    """The status file verbatim."""
    try:
        rs = read_status(run_id)
        return app.response_class(rs.to_json(), mimetype="application/json")
    except ResourceNotFoundError:
        return jsonify({"error": f"No status file for run '{run_id}'."}), 404
    except Exception as e:
        log.exception("read_status failed")
        return jsonify({"error": _friendly(e)}), 502


@app.get("/api/download/<path:blob_path>")
def api_download(blob_path: str):
    """Ladda ner en output-fil ur Blob via den lokala vyn (FD.23). Las-operation."""
    try:
        container, _, name = blob_path.partition("/")
        if not name:
            abort(400)
        from blob import _client
        bc = _client().get_blob_client(container=container, blob=name)
        data = bc.download_blob().readall()
        fname = name.split("/")[-1]
        return Response(
            data, mimetype="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except ResourceNotFoundError:
        return jsonify({"error": f"Filen finns inte i Blob: {blob_path}"}), 404
    except Exception as e:
        log.exception("blob download failed")
        return jsonify({"error": _friendly(e)}), 502


# ---------------------------------------------------------------------
# Etapp 2: validerings-drill-down. Laser de BEFINTLIGA master-kvittona
# som verify-sviterna redan producerat (verify_tool/receipts/). Ingen
# omkorning -- bara las PASS/FAIL som redan star dar.
# ---------------------------------------------------------------------
def _validator_receipt(phase_key: str, validator_name: str):
    """Hitta senaste enskilda validator-kvitto (ex 02_outliers_*.xlsx) for en
    validator inom en fas svit. Returnerar (Path, rel_path) eller (None, None)."""
    cfg = PHASE_RECEIPT.get(phase_key)
    if not cfg or not RECEIPTS_DIR.exists():
        return None, None
    # Validator-filer borjar med NN_<name>_<datum>.xlsx i samma subdir som mastret.
    candidates = []
    for date_dir in RECEIPTS_DIR.iterdir():
        if not date_dir.is_dir():
            continue
        search_dir = date_dir / cfg["subdir"] if cfg["subdir"] else date_dir
        if not search_dir.exists():
            continue
        for f in search_dir.glob(f"*_{validator_name}_*.xlsx"):
            if f.name.startswith("00_"):
                continue
            candidates.append((f.stat().st_mtime, f))
    if not candidates:
        return None, None
    candidates.sort(reverse=True)
    path = candidates[0][1]
    return path, str(path.relative_to(REPO_ROOT)).replace("\\", "/")


def _key_kpi(path: Path, validator_name: str) -> str:
    """Lokal fil-vag (fallback). Parsern delas med Blob-vagen (LB.85)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        lines = [str(r[0]).strip() for r in ws.iter_rows(values_only=True) if r[0] is not None]
    except Exception:
        return ""
    return _key_kpi_lines(lines)


def _key_kpi_lines(lines) -> str:
    """Plocka EN key-insight-rad ur ett validator-kvitto (top-management-niva).
    Vi laser de gate-rader som slutar med '-> PASS/REVIEW' och valjer den mest
    informativa. Ingen ny berakning."""
    # Gate-rader: "... -> PASS/REVIEW". Ta den forsta REVIEW (mest relevant) annars forsta PASS.
    gates = [l for l in lines if "->" in l and any(g in l for g in ("PASS", "REVIEW", "FAIL"))]
    review = [g for g in gates if "REVIEW" in g or "FAIL" in g]
    pick = (review[0] if review else (gates[0] if gates else ""))
    # Stada bort ledande markorer
    return pick.replace("  ", " ").strip()


@app.get("/api/receipts_list/<phase_key>")
def api_receipts_list(phase_key: str):
    """Lista de enskilda valideringskvittona för en fas (drill 3). Ingen
    PASS/REVIEW-dom -- bara filnamn + exportväg, så användaren kan hämta
    vilken granskning som helst. SENASTE per validator-typ (inga dubbletter
    från flera körningar samma dag, LB.69-anda) + datum/tid i namnet så det
    syns vilken körning det är."""
    import re as _re
    from datetime import datetime as _dt
    cfg = PHASE_RECEIPT.get(phase_key)
    if not cfg:
        return jsonify({"files": []})

    # FD.33-B1: fonster-vald korning -> DET fonstrets kvitton ur Blob.
    win = _win(request.args.get("run", ""))
    if win:
        suite = _SUITE_BY_SUBDIR.get(cfg["subdir"], cfg["subdir"] or "extraction")
        rows = _blob_receipts(suite, win)
        if rows is not None:
            by_type = {}
            for r in rows:  # nyast forst -> forsta traff per typ vinner
                name = r["name"]
                if name.startswith("00_") or not name.lower().endswith(".xlsx"):
                    continue
                stem = name.rsplit(".", 1)[0]
                vtype = _re.sub(r'^\d+_', '', stem)
                vtype = _re.sub(r'_\d{4}-\d{2}-\d{2}.*$', '', vtype)
                if vtype in by_type:
                    continue
                m = _re.search(r'(\d{4}-\d{2}-\d{2})_(\d{2})(\d{2})(\d{2})', stem)
                stamp = f"{m.group(1)} {m.group(2)}:{m.group(3)}" if m else str(r.get("modified", ""))[:16]
                label = vtype.replace("_", " ")
                if stamp:
                    label = f"{label} · {stamp}"
                by_type[vtype] = {"name": label,
                                  "file": f"{BLOB_MARK}{suite}/{win}/{name}"}
            return jsonify({"files": [by_type[k] for k in sorted(by_type)],
                            "source": "blob", "window": win})
        # Blob onaabart -> lokal fallback nedan (senaste-vinner, period-blind)

    if not RECEIPTS_DIR.exists():
        return jsonify({"files": []})
    # Hitta senaste datum-mapp som har den här svitens kvitton
    best_dir, best_mtime = None, -1
    for date_dir in RECEIPTS_DIR.iterdir():
        if not date_dir.is_dir():
            continue
        search_dir = date_dir / cfg["subdir"] if cfg["subdir"] else date_dir
        if not search_dir.exists():
            continue
        masters = list(search_dir.glob(cfg["glob"]))
        if masters:
            mt = max(m.stat().st_mtime for m in masters)
            if mt > best_mtime:
                best_mtime, best_dir = mt, search_dir
    if best_dir is None:
        return jsonify({"files": []})

    # Gruppera per validator-TYP (namn utan datum), behåll bara senaste per typ.
    # Dubbletterna kom av att samma validator kördes flera ggr samma dag.
    by_type = {}  # typ -> (mtime, path, datumtext)
    for f in sorted(best_dir.glob("*.xlsx")):
        if f.name.startswith("00_"):
            continue  # master hanteras separat
        vtype = _re.sub(r'^\d+_', '', f.stem)           # "02_outliers_2026-..." -> "outliers_2026-..."
        vtype = _re.sub(r'_\d{4}-\d{2}-\d{2}.*$', '', vtype)  # -> "outliers"
        # Plocka ut datum+tid ur filnamnet för visning (YYYY-MM-DD_HHMMSS)
        m = _re.search(r'(\d{4}-\d{2}-\d{2})_(\d{2})(\d{2})(\d{2})', f.stem)
        if m:
            stamp = f"{m.group(1)} {m.group(2)}:{m.group(3)}"
        else:
            stamp = ""
        mt = f.stat().st_mtime
        if vtype not in by_type or mt > by_type[vtype][0]:
            by_type[vtype] = (mt, f, stamp)

    files = []
    for vtype, (mt, f, stamp) in sorted(by_type.items()):
        label = vtype.replace("_", " ")
        if stamp:
            label = f"{label} · {stamp}"
        files.append({"name": label,
                      "file": str(f.relative_to(REPO_ROOT)).replace("\\", "/")})
    return jsonify({"files": files})


@app.get("/api/info")
def api_info():
    """About-fliken (FD.33-B2): README/arkitektur kondenserat till karnan --
    en kort utbildning av det bakomliggande. Innehallet ags av info_config.py."""
    try:
        from info_config import INFO
        return jsonify(INFO)
    except Exception as e:
        return jsonify({"title": "About", "sections": [
            {"h": "Not available", "body": f"info_config.py missing or broken: {str(e)[:200]}"}]})


@app.get("/api/proof_chain")
def api_proof_chain():
    """Bit-för-bit-bevisen mot fryst facit (statiskt, ur story_config)."""
    return jsonify(PROOF_CHAIN)


def _data_window(win=None) -> dict:
    """Läs det VÄXANDE datafönstret dynamiskt ur senaste extraction-kvitto
    ('Date window: A -> B'). Parqueten/dataprepen är källan och gäller tvärs
    alla familjer (samma data). BCG-facit är fryst och känt (konstant).
    Uppdateras av sig själv vid ny körning -- ingen hårdkodning. Keep it simple."""
    import re, openpyxl
    facit = FACIT_WINDOW  # enda agaren: story_config (FD.33-B1)
    if win:  # fonstret AR adressen (run_id) -- harled, parsea inte kvitton (LB.85)
        return {"facit": facit, "now": win.replace("_", " → ")}
    now = None
    path, _ = _latest_receipt("extraction")
    if path is not None:
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[0] and "Date window" in str(row[0]):
                    m = re.search(r'(\d{4}-\d{2}-\d{2})\s*->\s*(\d{4}-\d{2}-\d{2})', str(row[0]))
                    if m:
                        now = f"{m.group(1)} → {m.group(2)}"
                    break
        except Exception:
            pass
    return {"facit": facit, "now": now}


def _latest_receipt(phase_key: str):
    """Hitta senaste master-kvittot for en fas. Returnerar (Path, datum-str)
    eller (None, None). Soker i alla datum-mappar, valjer nyaste filen."""
    cfg = PHASE_RECEIPT.get(phase_key)
    if not cfg or not RECEIPTS_DIR.exists():
        return None, None
    candidates = []
    for date_dir in RECEIPTS_DIR.iterdir():
        if not date_dir.is_dir():
            continue
        search_dir = date_dir / cfg["subdir"] if cfg["subdir"] else date_dir
        if not search_dir.exists():
            continue
        for f in search_dir.glob(cfg["glob"]):
            candidates.append((f.stat().st_mtime, f, date_dir.name))
    if not candidates:
        return None, None
    candidates.sort(reverse=True)
    _, path, date_str = candidates[0]
    return path, date_str


def _parse_master_receipt(path: Path) -> dict:
    """Lokal fil-vag (fallback). Parsern delas med Blob-vagen (LB.85)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    return _parse_master_lines([row[0] for row in ws.iter_rows(values_only=True)])


def _parse_master_lines(lines) -> dict:
    """Las ett master-kvitto (rad-lista, en stdout-rad per cell). Extraherar
    overall, summary-counts och per-validator-status. Ingen ny berakning --
    bara las det som redan star dar."""
    validators, summary, overall, cur = [], {}, None, None
    for raw in lines:
        if raw is None:
            continue
        line = str(raw)
        m = re.search(r'\[(\d+)/(\d+)\]\s+Running:\s+(\S+\.py)', line)
        if m:
            cur = {"name": m.group(3).replace("validate_", "").replace(".py", ""), "status": None}
            validators.append(cur)
            continue
        m = re.search(r'Status:\s+([A-Z]+)', line)
        if m and cur is not None:
            cur["status"] = m.group(1)
            cur = None
            continue
        m = re.search(r'OVERALL:\s+(\w+)', line)
        if m:
            overall = m.group(1)
            continue
        m = re.match(r'\s+(PASS|REVIEW|FAIL|INFO|SKIP|OTHER|TOTAL):\s+(\d+)', line)
        if m:
            summary[m.group(1)] = int(m.group(2))
    return {"overall": overall, "summary": summary, "validators": validators}


@app.get("/api/validation/<phase_key>")
def api_validation(phase_key: str):
    """Validerings-sammanfattning for en fas, ur senaste lokala master-kvitto.
    Returnerar overall/summary/validators + receipt-filnamn for nedladdning."""
    cfg = PHASE_RECEIPT.get(phase_key)
    if not cfg:
        return jsonify({"available": False, "reason": "Ingen valideringssvit kopplad till denna fas."})
    # FD.33-B1: fonster-vald korning -> master + validator-KPI:er fran Blob.
    win = _win(request.args.get("run", ""))
    if win:
        import fnmatch
        suite = _SUITE_BY_SUBDIR.get(cfg["subdir"], cfg["subdir"] or "extraction")
        rows = _blob_receipts(suite, win)
        if rows is not None:
            masters = [r for r in rows if fnmatch.fnmatch(r["name"], cfg["glob"])]
            if not masters:
                return jsonify({"available": False, "window": win,
                                "reason": f"Inget {suite}-masterkvitto for fonstret {win} i Blob (kommer med Etapp B-korningar)."})
            m0 = masters[0]  # nyast forst
            try:
                parsed = _parse_master_lines(
                    _xlsx_lines_from_bytes(_blob_receipt_bytes(suite, win, m0["name"])))
                for v in parsed.get("validators", []):
                    cand = [r for r in rows
                            if fnmatch.fnmatch(r["name"], f"*_{v['name']}_*.xlsx")
                            and not r["name"].startswith("00_")]
                    if cand:
                        try:
                            vlines = _xlsx_lines_from_bytes(
                                _blob_receipt_bytes(suite, win, cand[0]["name"]))
                            v["kpi"] = _key_kpi_lines(
                                [str(x).strip() for x in vlines if x is not None])
                        except Exception:
                            pass
                        v["receipt_file"] = f"{BLOB_MARK}{suite}/{win}/{cand[0]['name']}"
                parsed.update({"available": True, "suite": cfg["suite"],
                               "date": f"{win} · {str(m0.get('modified',''))[:16]}",
                               "window": win, "source": "blob",
                               "receipt_file": f"{BLOB_MARK}{suite}/{win}/{m0['name']}"})
                return jsonify(parsed)
            except Exception as e:
                log.exception("blob master parse failed")
                return jsonify({"available": False, "reason": _friendly(e)})
        # Blob onaabart -> lokal fallback nedan

    path, date_str = _latest_receipt(phase_key)
    if path is None:
        return jsonify({"available": False, "reason": "Inget valideringskvitto hittat lokalt."})
    try:
        parsed = _parse_master_receipt(path)
        # Berika varje validator med key-KPI + eget kvitto (etapp 3, top-mgmt-niva)
        for v in parsed.get("validators", []):
            vpath, vrel = _validator_receipt(phase_key, v["name"])
            if vpath is not None:
                v["kpi"] = _key_kpi(vpath, v["name"])
                v["receipt_file"] = vrel
        parsed.update({
            "available": True,
            "suite": cfg["suite"],
            "date": date_str,
            "receipt_file": str(path.relative_to(REPO_ROOT)).replace("\\", "/"),
        })
        return jsonify(parsed)
    except Exception as e:
        log.exception("parse receipt failed")
        return jsonify({"available": False, "reason": _friendly(e)})


@app.get("/api/receipt/<path:rel_path>")
def api_receipt(rel_path: str):
    """Ladda ner ett valideringskvitto (Excel) lokalt fran verify_tool/receipts.
    Sakerhet: tillater bara filer UNDER verify_tool/ (ingen path-traversal)."""
    # FD.33-B1: '_blob/'-prefix = kvitto i receipts-containern (fonster-vagen).
    if rel_path.startswith(BLOB_MARK):
        name = rel_path[len(BLOB_MARK):]
        if ".." in name or name.count("/") < 2:
            abort(404)
        try:
            from blob import _client, CONTAINER_RECEIPTS
            data = _client().get_container_client(CONTAINER_RECEIPTS) \
                            .get_blob_client(name).download_blob().readall()
        except Exception:
            abort(404)
        fname = name.split("/")[-1]
        return Response(data, mimetype="application/octet-stream",
                        headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    target = (REPO_ROOT / rel_path).resolve()
    vt = (REPO_ROOT / "verify_tool").resolve()
    if not str(target).startswith(str(vt)) or not target.exists():
        abort(404)
    return send_file(str(target), as_attachment=True, download_name=target.name)


def _friendly(e: Exception) -> str:
    msg = str(e)
    if "token" in msg.lower() or "401" in msg or "Kunde inte lasa kontonyckeln" in msg:
        return ("Could not authenticate to Blob. Run: az login --scope "
                "https://management.core.windows.net//.default and reload.")
    if "403" in msg or "AuthorizationPermissionMismatch" in msg:
        return ("403 from Blob: data-plane role missing (ABAC). The app runs in "
                "account-key mode by default -- check PRICINGMODEL_AUTH=key.")
    return msg[:300]


def main() -> int:
    ap = argparse.ArgumentParser(description="Phase Z read-only status dashboard (local).")
    ap.add_argument("--port", type=int, default=5000)
    ap.add_argument("--check", action="store_true",
                    help="Verify Blob is reachable and list runs, then exit.")
    args = ap.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    if args.check:
        print("Checking Blob access (first call fetches the account key, ~3-5 s) ...")
        ids = sorted(list_runs(), reverse=True)
        print(f"OK -- {len(ids)} run(s) with a status file:")
        for i in ids[:10]:
            print(f"  {i}")
        print(f"Receipts dir: {RECEIPTS_DIR} (exists={RECEIPTS_DIR.exists()})")
        return 0

    print(f"Phase Z dashboard (read-only): http://127.0.0.1:{args.port}")
    print("First page load is slow (~3-5 s) while the account key is fetched.")
    app.run(host="127.0.0.1", port=args.port, debug=False)
    return 0


if __name__ == "__main__":
    sys.exit(main())
