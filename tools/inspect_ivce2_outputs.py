# inspect_ivce2_outputs.py -- READ-ONLY: vad ar 10/6-filerna Steg 5 skulle skriva over?
import openpyxl
from pathlib import Path

ROOT = Path(r"C:\Projekt\BCG\Pipeline\02. Elasticity\3. Product Site Level Models\output")
TARGETS = [
    "output_summary_ready_Ivce_2.xlsx",
    "model_result_summary_ready_Ivce_2.xlsx",
    "significant_variable_summary_Ivce_2.xlsx",
]
for name in TARGETS:
    p = ROOT / name
    print("=" * 60)
    print(name)
    if not p.exists():
        print("  (finns ej)")
        continue
    kb = round(p.stat().st_size / 1024, 1)
    import datetime
    mod = datetime.datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    print(f"  {kb} KB, andrad {mod}")
    try:
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb.active
        print(f"  ark: {wb.sheetnames}")
        print(f"  rader (aktiv): {ws.max_row}, kol: {ws.max_column}")
        # forsta dataraden for att se KEY-format/population-hint
        hdr = [c.value for c in next(ws.iter_rows(max_row=1))][:6]
        print(f"  rubriker: {hdr}")
    except Exception as e:
        print(f"  LASFEL: {e}")
