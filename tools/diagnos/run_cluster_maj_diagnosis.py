#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_cluster_maj_diagnosis.py  --  ORKESTRERARE: kör sond 1-4 i ordning
=======================================================================
Kör hela cluster-maj-diagnos-sviten (sond 1-4) i LOGISK ordning, fångar
varje sonds utfall, och skriver ETT samlat Excel-kvitto. Syftet: reda ut
hela 2026-06-23-blockeringen (KeyError 'No_of_Sites' i feature_selection)
på EN körning i stället för att snubbla diagnos för diagnos (vilket tog
hela den dagen och bytte spår fem gånger).

ORDNING (varje bygger på föregående):
  1. CSV vs config        -- VAD saknas i config col_type? (symptom)
  2. april vs maj CSV     -- HUR skiljer CSV-strukturerna? (löser motsägelse)
  3. data_prep-härkomst   -- VARFÖR skiljer de? (rotorsak, lokalt)
  4. nedströms-konsekvens -- räcker config-fix, eller biter det senare?

EFTER sviten vet man: vad, hur, varför, och om en config-fix räcker.
DÅ -- och först då -- skrivs fixen (config additivt i REPOT, ej bara VM),
committas, tillämpas på VM, och cluster relanchas.

FÖRUTSÄTTNING: ssh till VM igång (sond 1,2,4), token/PIM. Sond 3 är LOKAL
(läser repo) och behöver ingen VM. Kör därför sond 3 även om VM är nere.

KÖR (global py -3.11):
    cd <mappen med sonderna>
    py -3.11 run_cluster_maj_diagnosis.py
    py -3.11 run_cluster_maj_diagnosis.py --skip-vm   # bara sond 3 (lokal), om VM nere

Utvecklare: Jens Palmö (Senior Business Analyst, Evidensia). Författare: Claude-rådgivare.
Beroende: de fyra probe_*.py i samma mapp; openpyxl (kvitto); ssh för 1/2/4.
"""
from __future__ import annotations
import argparse
import datetime
import subprocess
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
PROBES = [
    ("1", "probe_1_csv_vs_config.py", "CSV vs config col_type", True),
    ("2", "probe_2_april_vs_maj_csv.py", "april-CSV vs maj-CSV strukturdiff", True),
    ("3", "probe_3_dataprep_provenance.py", "data_prep-härkomst (VARFÖR skiljer de)", False),
    ("4", "probe_4_downstream_impact.py", "nedströms-konsekvens", True),
]


def run_probe(fname: str) -> tuple[int, str]:
    """Kör en sond, fånga stdout+stderr och returkod."""
    fp = HERE / fname
    if not fp.exists():
        return (127, f"[SAKNAS] {fp}")
    r = subprocess.run([sys.executable, str(fp)], capture_output=True, text=True, timeout=300)
    return (r.returncode, (r.stdout or "") + (("\n[STDERR]\n" + r.stderr) if r.stderr.strip() else ""))


def write_receipt(results: list[tuple[str, str, str, int, str]], out_dir: Path) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception as e:
        print(f"[VARNING] openpyxl saknas ({e}) -- konsol räcker.")
        return None
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    fp = out_dir / f"00_cluster_maj_diagnosis_{stamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnos"
    mono = Font(name="Consolas", size=9)
    bold = Font(name="Consolas", size=10, bold=True)
    fills = {0: PatternFill("solid", fgColor="E8F0E8"),  # PASS grön
             2: PatternFill("solid", fgColor="FFF3CD"),  # REVIEW gul
             127: PatternFill("solid", fgColor="F8D7DA")}  # SAKNAS röd

    ws.append([f"CLUSTER-MAJ DIAGNOS (sond 1-4)  {stamp}"])
    ws["A1"].font = bold
    ws.append(["Reder ut KeyError 'No_of_Sites' i feature_selection (block 2026-06-23)"])
    ws.append([])
    ws.append(["SOND", "VAD", "DOM(rc)"])
    for c in range(1, 4):
        ws.cell(ws.max_row, c).font = bold
    for num, fname, label, rc, _ in results:
        verdict = {0: "PASS", 2: "REVIEW", 127: "SAKNAS"}.get(rc, f"rc={rc}")
        ws.append([f"Sond {num}", label, verdict])
        ws.cell(ws.max_row, 3).fill = fills.get(rc, fills[2])
    ws.append([])

    # Full stdout per sond, en rad per textrad (Consolas, data_type s)
    for num, fname, label, rc, out in results:
        ws.append([])
        ws.append([f"===== SOND {num}: {label}  ({fname}) ====="])
        ws.cell(ws.max_row, 1).font = bold
        for line in out.splitlines():
            c = ws.cell(ws.max_row + 1, 1, value=line)
            c.font = mono
            c.data_type = "s"
    for col, w in (("A", 50), ("B", 50), ("C", 14)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=False)
            if not c.font or c.font.name != "Consolas":
                c.font = mono
    wb.save(fp)
    return fp


def main() -> int:
    ap = argparse.ArgumentParser(description="Kör cluster-maj-diagnos-sviten (sond 1-4).")
    ap.add_argument("--skip-vm", action="store_true", help="Hoppa VM-sonder (1,2,4); kör bara lokal sond 3.")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    print("=" * 72)
    print("CLUSTER-MAJ DIAGNOS -- kör sond 1-4 i ordning")
    print("Reder ut: KeyError 'No_of_Sites' i feature_selection (block 2026-06-23)")
    print("=" * 72)

    results = []
    for num, fname, label, needs_vm in PROBES:
        if needs_vm and args.skip_vm:
            print(f"\n[HOPPAR] Sond {num} ({label}) -- kräver VM, --skip-vm satt.")
            results.append((num, fname, label, -1, "[HOPPAD: --skip-vm]"))
            continue
        print(f"\n{'#'*72}\n# SOND {num}: {label}\n{'#'*72}")
        rc, out = run_probe(fname)
        print(out)
        results.append((num, fname, label, rc, out))

    # Samlad dom
    print("\n" + "=" * 72)
    print("SAMLAD DOM:")
    for num, _, label, rc, _ in results:
        v = {0: "PASS", 2: "REVIEW", 127: "SAKNAS", -1: "HOPPAD"}.get(rc, f"rc={rc}")
        print(f"  Sond {num}: {v:<8} {label}")
    print("=" * 72)

    out_dir = Path(args.out) if args.out else (
        Path.cwd() / "verify_tool" / "receipts" / datetime.date.today().isoformat() / "cluster_maj_diagnosis")
    fp = write_receipt(results, out_dir)
    if fp:
        print(f"\n[Kvitto] {fp}")
    print("[KLART] Läs domarna ovan. Fixa FÖRST när alla fyra är förstådda.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
