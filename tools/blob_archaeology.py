#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
blob_archaeology.py -- FD.33 steg 1: fullstandig inventering + klassificering av Blob
======================================================================================
LAS-ONLY. Ror ingenting. Producerar den karta migreringen (blob_migrate_fd33.py)
exekverar mot: inventory.json + Excel-kvitto + konsoltabell.

VARFOR (FD.33 / BLOB_MALSTRUKTUR.md):
    Blob-output blandar scheman (datum-mappar vs fonster), kvitton finns bara lokalt,
    och natten 2026-07-03 bevisade att ogiltiga generationer (stub 5889 B, stale
    Model_Feed) ligger kvar bredvid giltiga. Innan NAGOT flyttas maste ALLT listas
    och domas. Detta ar arkeologin NEXT_SESSION kraver ("Inkl. arkeologi +
    stub/stale-generationsstad").

LARDOM INBYGGD (dagens, 2026-07-03): az-CLI-listning trunkeras vid ~5000 blobbar
    ("WARNING: Next Marker") -- darfor SDK har (list_blobs paginerar sjalv).

KLASSER (domslut per blob):
    VALID-CORE        exakt namn+storlek ur allowlist (dagens matta sanningar)
    KNOWN-INVALID     exakt namn ur denylist (NEXT_SESSION:s ogiltiga generationer)
    STUB              storlek under golv for sin filklass (LB.86: 5889 B-klassen)
    REGENERABLE-BULK  automl/model_objects/details/results-trad (per-KEY mellandata)
    FROZEN-FACIT      pipeline/00_frozen_facit/** (ankaret -- flyttas ALDRIG)
    STATUS            runstatus/*.json (fonster-baserat redan -- flyttas ej)
    UNKNOWN           allt annat -> manskliga ogon fore migrering

LOKALA KVITTON: verify_tool/receipts/<datum>/ + workspace/validation_receipts/
    mappas till fonster via DATE_TO_WINDOW-tabellen (Jens ager sanningen -- redigera
    dar om fel) + billig populations-sniff (6624=april-site, 6604=maj-site, 3791=
    maj-cluster, 4180=april-cluster, 125=bundle) som VERIFIERAR/flaggar. Omappbart
    hamnar i _unmapped (arligt, aldrig gissat).

KORNING (PowerShell, C:\\Projekt\\BCG -- token FARSK forst, LB.88!):
    az login --scope https://management.core.windows.net//.default
    $env:PRICINGMODEL_AUTH = "key"
    py -3.11 tools\\blob_archaeology.py
    py -3.11 tools\\blob_archaeology.py --no-local     # hoppa lokala kvitton
    py -3.11 tools\\blob_archaeology.py --also-prod    # sok aven prod-kontot (FD.35-koll)

OUTPUT:
    workspace\\fd33\\inventory.json            (migreringsplanens rainput)
    workspace\\validation_receipts\\blob_archaeology_<ts>.xlsx

Beroenden: blob.py (env-pekad, key-lage), azure-storage-blob, openpyxl.
Beroende av: blob_migrate_fd33.py (laser inventory.json).
Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia). Author: Claude advisor, FD.33-passet 2026-07-03.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

# -- Konsol immun mot CP1252 (encoding-klassen, LB.90-punkten) -----------------
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

REPO = Path(os.environ.get("BCG_REPO", r"C:\Projekt\BCG"))
if not REPO.exists():
    REPO = Path.cwd()

# -- Peka blob.py mot TEST-kontot i key-lage (samma monster som upload_frozen_facit)
os.environ.setdefault("PRICINGMODEL_STORAGE", "evbcgpricinginput")
os.environ.setdefault("PRICINGMODEL_RG", "ev-openai-swce-rg-test")
os.environ.setdefault("PRICINGMODEL_AUTH", "key")

sys.path.insert(0, str(REPO / "orchestration" / "infrastructure"))
sys.path.insert(0, str(REPO / "orchestration" / "shared"))

OUT_DIR = REPO / "workspace" / "fd33"
RECEIPT_DIR = REPO / "workspace" / "validation_receipts"

# =============================================================================
# SANNINGSTABELLER -- Jens ager dessa. Redigera HAR, aldrig i logiken.
# =============================================================================

# Fonster-ID (run_id = datafonster, run_status-kontraktet)
W_MAJ = "2022-07-01_2026-05-31"
W_APR = "2022-07-01_2026-04-30"

# Datum-mapp i Blob-output -> vilket FONSTER innehallet tillhor.
# (Matt 2026-07-03: 2026-06-17-prefixet BAR maj-artefakter -- site reparerades
#  in dit i natt. Datum-mapp != innehallsfonster ar exakt det FD.33 dodar.)
DATE_TO_WINDOW = {
    "2026-06-17": W_MAJ,
    "2026-06-20": W_MAJ,
    "2026-06-22": W_MAJ,
    "2026-06-24": W_MAJ,
    "2026-07-02": W_MAJ,
    "2026-07-03": W_MAJ,
}

# Lokala kvitto-datummappar -> fonster (redigera vid behov; omappat -> _unmapped)
RECEIPT_DATE_TO_WINDOW = {
    "2026-06-17": W_APR, "2026-06-18": W_APR,
    "2026-06-22": W_MAJ, "2026-06-23": W_MAJ, "2026-06-24": W_MAJ,
    "2026-06-25": W_MAJ, "2026-06-26": W_MAJ,
    "2026-07-02": W_MAJ, "2026-07-03": W_MAJ,
}

# Populations-sniff: tal som entydigt pekar ut fonster/familj (verifierare, ej domare)
POPULATION_HINTS = {
    "6604": ("site", W_MAJ), "6624": ("site", W_APR),
    "3791": ("cluster", W_MAJ), "4180": ("cluster", W_APR),
    "22477": ("r12", W_MAJ), "22,477": ("r12", W_MAJ),
}

# VALID-CORE: exakta (namn-suffix, byte) matta 2026-07-03 -- dagens sanningar.
ALLOWLIST_EXACT = [
    ("Final_Fallback_Data_20260703_125923.xlsx", 7_840_703),
    ("Model_Feed_2026-07-03.xlsx", 969_355),
    ("output_summary.xlsx", 561_503),   # site maj
    ("output_summary.xlsx", 325_044),   # cluster maj
]

# KNOWN-INVALID: NEXT_SESSION:s ogiltiga generationer + stub-kanalen
DENYLIST_SUBSTR = [
    "Model_Feed_2026-06-22", "Model_Feed_2026-07-02",
    "Final_Fallback_Data_20260622",        # juni-eran, fore cluster-arlighet (arkeologi-fynd 2026-07-03)
    "Final_Fallback_Data_20260702",        # 6.09 MB = stub-kontaminerade erans vav (arkeologi-fynd)
    "Final_Fallback_Data_20260703_0010",   # nattens site-stale vav
    "Final_Fallback_Data_20260703_0821",   # mellangeneration fore token-fix
    "Final_Fallback_Data_20260703_1126",   # Sigma-kraschens generation (data ok men ersatt)
    "Final_Fallback_Data_20260703_1225",   # token-dodens generation (ersatt av 125923)
]

# Storleksgolv per filklass (LB.86-stubbdetektering; 0 = ingen dom pa storlek)
SIZE_FLOOR = [
    (re.compile(r"output_summary(_ready)?\.xlsx$", re.I), 100_000),
    (re.compile(r"model_summary\.xlsx$", re.I), 1_000_000),
    (re.compile(r"Final_Fallback_Data_.*\.xlsx$", re.I), 1_000_000),
    (re.compile(r"Model_Feed_.*\.xlsx$", re.I), 100_000),
]

RE_BULK = re.compile(r"/(automl|model_objects|details|results)/", re.I)
RE_DATEDIR = re.compile(r"^(\d{4}-\d{2}-\d{2})/")

CONTAINERS = ["runstatus", "output", "input", "pipeline", "receipts", "quarantine"]

ROWS: list[dict] = []


def say(tag: str, msg: str) -> None:
    print(f"[{tag}] {msg}", flush=True)


def floor_for(name: str) -> int:
    for rx, fl in SIZE_FLOOR:
        if rx.search(name):
            return fl
    return 0


def classify_blob(container: str, name: str, size: int) -> tuple[str, str]:
    """-> (klass, not). Ordning: facit/status -> deny -> allow -> stub -> bulk -> unknown."""
    if container == "pipeline" and name.startswith("00_frozen_facit/"):
        return "FROZEN-FACIT", "ankare, flyttas aldrig"
    if container == "runstatus":
        return "STATUS", "fonster-baserad redan"
    if container in ("receipts", "quarantine"):
        return "ALREADY-NEW", "ligger redan i ny struktur"
    for bad in DENYLIST_SUBSTR:
        if bad in name:
            return "KNOWN-INVALID", f"denylist: {bad}"
    base = name.rsplit("/", 1)[-1]
    for allow_name, allow_size in ALLOWLIST_EXACT:
        if base == allow_name and size == allow_size:
            return "VALID-CORE", f"allowlist exakt ({size} B)"
    fl = floor_for(base)
    if fl and size < fl:
        return "STUB", f"{size} B < golv {fl} (LB.86)"
    if RE_BULK.search("/" + name):
        return "REGENERABLE-BULK", "per-KEY mellandata (automl-klassen)"
    return "UNKNOWN", ""


def window_for_blob(name: str) -> str:
    m = RE_DATEDIR.match(name)
    if m:
        return DATE_TO_WINDOW.get(m.group(1), "_unmapped")
    if name.startswith("00_frozen_facit/"):
        return "_frozen"
    return "_unmapped"


def inventory_account(label: str) -> None:
    import blob  # noqa: E402  (env redan satt)
    svc = blob._client()
    say("KONTO", f"{label}: {blob.STORAGE_ACCOUNT} (auth={blob._AUTH_MODE})")
    for cname in CONTAINERS:
        try:
            cc = svc.get_container_client(cname)
            if not cc.exists():
                say("skip", f"container '{cname}' finns ej")
                continue
        except Exception as e:
            say("skip", f"container '{cname}': {type(e).__name__}")
            continue
        n = 0
        for b in cc.list_blobs():  # SDK paginerar -- ingen Next Marker-falla
            klass, note = classify_blob(cname, b.name, b.size or 0)
            ROWS.append({
                "account": label, "container": cname, "blob": b.name,
                "bytes": int(b.size or 0),
                "modified": str(getattr(b, "last_modified", ""))[:19],
                "class": klass, "window": window_for_blob(b.name), "note": note,
            })
            n += 1
        say("LIST", f"{cname}: {n} blobbar")


def sniff_population(xlsx: Path) -> str:
    """Billig sniff: sok kanda populationstal i forsta bladets forsta ~80 rader."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active
        hits = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 80:
                break
            for v in row:
                s = str(v) if v is not None else ""
                for tok, (fam, win) in POPULATION_HINTS.items():
                    if tok in s:
                        hits.add(win)
        wb.close()
        if len(hits) == 1:
            return next(iter(hits))
        return ""
    except Exception:
        return ""


def inventory_local_receipts() -> None:
    roots = [
        ("verify_tool", REPO / "verify_tool" / "receipts"),
        ("workspace", REPO / "workspace" / "validation_receipts"),
    ]
    for root_label, root in roots:
        if not root.is_dir():
            say("skip", f"lokal kvittorot saknas: {root}")
            continue
        n = 0
        for fp in root.rglob("*"):
            if not fp.is_file():
                continue
            rel = fp.relative_to(root)
            parts = rel.parts
            date_dir = parts[0] if re.match(r"^\d{4}-\d{2}-\d{2}$", parts[0] or "") else ""
            win = RECEIPT_DATE_TO_WINDOW.get(date_dir, "")
            sniff = sniff_population(fp) if fp.suffix.lower() == ".xlsx" else ""
            if sniff and win and sniff != win:
                note = f"KONFLIKT: datumtabell={win} men population={sniff}"
                win = "_conflict"
            elif sniff and not win:
                win, note = sniff, "fonster via population-sniff"
            elif win:
                note = "fonster via datumtabell"
            else:
                win, note = "_unmapped", "ingen mappning -- manskligt oga"
            suite = parts[1] if len(parts) > 2 else ("extraction" if len(parts) == 2 else "misc")
            ROWS.append({
                "account": "LOCAL", "container": f"local:{root_label}",
                "blob": str(rel).replace(os.sep, "/"),
                "bytes": fp.stat().st_size,
                "modified": datetime.fromtimestamp(fp.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                "class": "LOCAL-RECEIPT", "window": win,
                "note": f"suite={suite}; {note}",
            })
            n += 1
        say("LOCAL", f"{root}: {n} kvittofiler")


def write_outputs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    inv = OUT_DIR / "inventory.json"
    inv.write_text(json.dumps(ROWS, ensure_ascii=False, indent=1), encoding="utf-8")
    say("SKRIVEN", str(inv))

    # Excel-kvitto (standard: datum/tid i namn + huvudrader)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        path = RECEIPT_DIR / f"blob_archaeology_{ts}.xlsx"
        wb = Workbook(); ws = wb.active; ws.title = "archaeology"
        summary: dict[str, int] = {}
        for r in ROWS:
            summary[r["class"]] = summary.get(r["class"], 0) + 1
        head = [("Receipt", "FD.33 blob archaeology (read-only inventory)"),
                ("Generated (UTC)", ts),
                ("Account (primary)", os.environ.get("PRICINGMODEL_STORAGE", "")),
                ("Rows", len(ROWS)),
                ("Classes", "; ".join(f"{k}={v}" for k, v in sorted(summary.items()))),
                ("Developer", "Jens Palmo (Senior Business Analyst, Evidensia)")]
        for k, v in head:
            ws.append([k, v]); ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.append([])
        cols = ["account", "container", "blob", "bytes", "modified", "class", "window", "note"]
        ws.append([c.upper() for c in cols])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        for r in ROWS:
            ws.append([r[c] for c in cols])
        widths = (10, 22, 70, 12, 18, 16, 26, 44)
        for col, w in zip("ABCDEFGH", widths):
            ws.column_dimensions[col].width = w
        wb.save(path)
        say("KVITTO", str(path))
    except Exception as e:
        say("WARN", f"Excel-kvitto ej skrivet ({type(e).__name__}: {e})")


def main() -> int:
    ap = argparse.ArgumentParser(description="FD.33 arkeologi: inventera + klassificera Blob (read-only).")
    ap.add_argument("--no-local", action="store_true", help="hoppa lokala kvitton")
    ap.add_argument("--also-prod", action="store_true",
                    help="inventera aven prod-kontot evipricingmodelstprod (FD.35-koll; kraver lasratt)")
    args = ap.parse_args()

    print("=" * 76)
    print("FD.33 ARKEOLOGI -- inventering + klassificering (ROR INGENTING)")
    print("=" * 76)

    inventory_account("TEST")

    if args.also_prod:
        # FD.35-sweep: finns kvarlamnad status/output pa prod? Mjuk degradering
        # om PIM-scope inte nar prod (observation loss != failure, AZ.7).
        global ROWS
        import importlib
        os.environ["PRICINGMODEL_STORAGE"] = "evipricingmodelstprod"
        os.environ.pop("PRICINGMODEL_RG", None)  # annan RG -- lat az sla upp brett
        try:
            import blob as _b
            importlib.reload(_b)
            _b._service_client = None
            inventory_account("PROD")
        except Exception as e:
            say("FD35", f"prod-kontot onaabart ({type(e).__name__}: {str(e)[:120]}) "
                        "-- FD.35-domen far vanta pa atkomst. Test-kontot ar oberort.")
        finally:
            os.environ["PRICINGMODEL_STORAGE"] = "evbcgpricinginput"
            os.environ["PRICINGMODEL_RG"] = "ev-openai-swce-rg-test"

    if not args.no_local:
        inventory_local_receipts()

    write_outputs()

    # Konsolsummering: det manskliga ogat far domsluten direkt
    print("-" * 76)
    by = {}
    for r in ROWS:
        by.setdefault(r["class"], []).append(r)
    for klass in ["VALID-CORE", "KNOWN-INVALID", "STUB", "REGENERABLE-BULK",
                  "FROZEN-FACIT", "STATUS", "LOCAL-RECEIPT", "ALREADY-NEW", "UNKNOWN"]:
        rows = by.get(klass, [])
        if not rows:
            continue
        mb = sum(r["bytes"] for r in rows) / 1e6
        print(f"  {klass:<17} {len(rows):>6} st  {mb:>10.1f} MB")
        if klass in ("KNOWN-INVALID", "STUB", "UNKNOWN"):
            for r in rows[:12]:
                print(f"      - {r['container']}/{r['blob']}  ({r['bytes']} B)")
            if len(rows) > 12:
                print(f"      ... +{len(rows)-12} till (se kvittot)")
    unmapped = [r for r in ROWS if r["window"] in ("_unmapped", "_conflict")]
    print("-" * 76)
    print(f"  Omappade/konflikt-fonster: {len(unmapped)} "
          f"(redigera DATE_TO_WINDOW-tabellerna i skripthuvudet vid behov)")
    print("=" * 76)
    print("NASTA: granska kvittot -> py -3.11 tools\\blob_migrate_fd33.py  (dry-run default)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
