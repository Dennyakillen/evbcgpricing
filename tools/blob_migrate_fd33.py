#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
blob_migrate_fd33.py -- FD.33 steg 2: exekvera BLOB_MALSTRUKTUR (familj-yttre, fonster-innerst)
================================================================================================
Konsumerar arkeologins inventory.json och EXEKVERAR EN PLAN, inte en improvisation
(BLOB_MALSTRUKTUR.md: "omlaggningen blir EXEKVERING AV EN PLAN").

ICKE-DESTRUKTIV BY DESIGN: allt ar server-side-KOPIA till nya prefix -- gamla
sokvagar rors inte. Pipelinen fortsatter fungera pa gamla vagar tills cutover-
committen (Etapp B: runners + app + blob._AFTER_INPUTS flippar TILLSAMMANS,
dry_run gron = bevis). Detta respekterar fyra-kartor-varningen: data forst
(reversibelt), kod-flip sen (atomart). Karantan/purge ar separata, explicita steg.

MALSTRUKTUR (BLOB_MALSTRUKTUR.md, kanonisk):
    output/    cluster|site|bundle/<window>/...   final/<window>/...
    receipts/  <suite>/<window>/...               (NY container; suite = extraction/
               rationality/provenance/probes/misc -- matchar appens PHASE_RECEIPT,
               medveten precisering av designens familj-slot)
    input/     parquet/transaction_data.parquet   data_prep/<window>/...
    runstatus/ + pipeline/00_frozen_facit/        OFORANDRADE

LAGEN (i ordning, alla utom --plan kraver farsk token -- LB.88: az login OMEDELBART fore):
    --plan        (DEFAULT) bygg + skriv plan, ror ingenting
    --commit      exekvera kopior/uppladdningar + MANIFEST.json (BB.11) + verifiera storlek
    --quarantine  kopiera KNOWN-INVALID/STUB -> container 'quarantine' (originalen KVAR)
    --purge       RADERA original som har verifierad karantan-kopia (enda destruktiva steget)
    --include-bulk    ta med REGENERABLE-BULK-trad (automl m.m.; default: hoppa + notera)
    --with-dataprep   ladda upp lokala dataprep-CSV:er till input/data_prep/<W_MAJ>/

KORNING (PowerShell, C:\\Projekt\\BCG):
    az login --scope https://management.core.windows.net//.default
    $env:PRICINGMODEL_AUTH = "key"
    py -3.11 tools\\blob_archaeology.py                       # 1. inventera
    py -3.11 tools\\blob_migrate_fd33.py                      # 2. granska planen
    py -3.11 tools\\blob_migrate_fd33.py --commit             # 3. utfor (kopior + manifest)
    py -3.11 tools\\blob_migrate_fd33.py --quarantine         # 4. stall undan ogiltigt
    #  --purge forst nar Etapp B (cutover) ar gron och karantanen granskad.

Beroenden: workspace/fd33/inventory.json (fran blob_archaeology.py), blob.py, azure-storage-blob, openpyxl.
Beroende av: Etapp B-cutover (runners/app pekar pa strukturen detta skript fyller).
Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia). Author: Claude advisor, FD.33-passet 2026-07-03.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import quote

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

REPO = Path(os.environ.get("BCG_REPO", r"C:\Projekt\BCG"))
if not REPO.exists():
    REPO = Path.cwd()

os.environ.setdefault("PRICINGMODEL_STORAGE", "evbcgpricinginput")
os.environ.setdefault("PRICINGMODEL_RG", "ev-openai-swce-rg-test")
os.environ.setdefault("PRICINGMODEL_AUTH", "key")

sys.path.insert(0, str(REPO / "orchestration" / "infrastructure"))
sys.path.insert(0, str(REPO / "orchestration" / "shared"))

INV_PATH = REPO / "workspace" / "fd33" / "inventory.json"
PLAN_PATH = REPO / "workspace" / "fd33" / "migration_plan.json"
RECEIPT_DIR = REPO / "workspace" / "validation_receipts"

W_MAJ = "2022-07-01_2026-05-31"

RE_DATE = r"\d{4}-\d{2}-\d{2}"
# (regex, dst_container, dst_mall) -- {window}/{tail}/{name} fylls ur matchen
RULES = [
    (re.compile(rf"^{RE_DATE}/cluster/(?P<tail>.+)$"), "output", "cluster/{window}/{tail}"),
    (re.compile(rf"^{RE_DATE}/site/(?P<tail>.+)$"),    "output", "site/{window}/{tail}"),
    (re.compile(rf"^{RE_DATE}/bundle/(?P<tail>.+)$"),  "output", "bundle/{window}/{tail}"),
    (re.compile(rf"^{RE_DATE}/(?P<name>output_summary\.xlsx)$"), "output", "site/{window}/{name}"),
    (re.compile(rf"^{RE_DATE}/(?P<name>Final_Fallback_Data_.+\.xlsx)$"), "output", "final/{window}/{name}"),
    (re.compile(rf"^{RE_DATE}/(?P<name>Model_Feed_.+\.xlsx)$"),          "output", "final/{window}/{name}"),
]
RULE_INPUT_PARQUET = (re.compile(r"^(?P<name>transaction_data\.parquet)$"),
                      "input", "parquet/{name}")
RULE_TX_MISPLACED = (re.compile(r"^00_frozen_facit/tx/(?P<name>.+\.csv)$"),
                     "input", "data_prep/" + W_MAJ + "/{name}")

# Lokala dataprep-CSV:er (opt-in, --with-dataprep). Masterdata EXKLUDERAD (7+ GB,
# regenererbar; parquet ar branslet). Bara det som faktiskt finns laddas upp.
LOCAL_DATAPREP = [
    REPO / "Pipeline" / "02. Elasticity" / "Sweden_Elasticity_Data_Prep_SQL" / "output" / "Sweden_weekly_model_data_P_C.csv",
    REPO / "Pipeline" / "02. Elasticity" / "Sweden_Elasticity_Data_Prep_SQL" / "output" / "Sweden_weekly_model_data_P_CH.csv",
    REPO / "Pipeline" / "02. Elasticity" / "Sweden_Elasticity_Data_Prep_SQL" / "output" / "Sweden_weekly_model_data_site_level.csv",
]

LOCAL_ROOTS = {
    "local:verify_tool": REPO / "verify_tool" / "receipts",
    "local:workspace":   REPO / "workspace" / "validation_receipts",
}

SUITE_RE = re.compile(r"suite=([A-Za-z_0-9]+)")


def say(tag: str, msg: str) -> None:
    print(f"[{tag}] {msg}", flush=True)


# ----------------------------------------------------------------------------- plan
def build_plan(inv: list[dict], include_bulk: bool, with_dataprep: bool) -> list[dict]:
    plan: list[dict] = []

    def add(action, src_c, src, dst_c, dst, size, klass, window, note=""):
        plan.append({"action": action, "src_container": src_c, "src": src,
                     "dst_container": dst_c, "dst": dst, "bytes": size,
                     "class": klass, "window": window, "note": note})

    for r in inv:
        c, name, size = r["container"], r["blob"], r["bytes"]
        klass, window = r["class"], r["window"]

        if r["account"] == "LOCAL":
            if klass != "LOCAL-RECEIPT":
                continue
            m = SUITE_RE.search(r.get("note", ""))
            suite = (m.group(1) if m else "misc").lower()
            win = window if window not in ("_unmapped", "_conflict") else f"_unmapped"
            date_dir = name.split("/")[0]
            fname = name.rsplit("/", 1)[-1]
            dst = (f"{suite}/{win}/{fname}" if win != "_unmapped"
                   else f"_unmapped/{date_dir}/{fname}")
            add("upload-local", c, name, "receipts", dst, size, klass, win)
            continue

        if klass in ("FROZEN-FACIT", "STATUS", "ALREADY-NEW"):
            # tx-CSV ligger fel-hyllad under frozen -- den ENDA frozen-flytten (kopia)
            m = RULE_TX_MISPLACED[0].match(name)
            if c == "pipeline" and m:
                add("copy", c, name, RULE_TX_MISPLACED[1],
                    RULE_TX_MISPLACED[2].format(**m.groupdict()), size, klass, W_MAJ,
                    "vaxande tx fel-hyllad under 00_frozen_facit -- kopieras till ratt hem")
            continue

        if klass in ("KNOWN-INVALID", "STUB"):
            add("quarantine", c, name, "quarantine", f"{c}/{name}", size, klass, window,
                r.get("note", ""))
            continue

        if klass == "REGENERABLE-BULK" and not include_bulk:
            add("skip-bulk", c, name, "", "", size, klass, window,
                "regenererbar mellandata -- hoppas (kor --include-bulk for att ta med)")
            continue

        if c == "input":
            m = RULE_INPUT_PARQUET[0].match(name)
            if m:
                add("copy", c, name, RULE_INPUT_PARQUET[1],
                    RULE_INPUT_PARQUET[2].format(**m.groupdict()), size, klass, "-")
                continue

        if c == "output":
            matched = False
            for rx, dst_c, tmpl in RULES:
                m = rx.match(name)
                if m:
                    if window in ("_unmapped", "_conflict"):
                        add("hold-unmapped", c, name, "", "", size, klass, window,
                            "datum-mapp saknar fonster i DATE_TO_WINDOW -- komplettera tabellen")
                    else:
                        add("copy", c, name, dst_c,
                            tmpl.format(window=window, **m.groupdict()), size, klass, window)
                    matched = True
                    break
            if matched:
                continue

        add("hold-unknown", c, name, "", "", size, klass, window,
            "ingen regel traffade -- manskligt oga (utoka RULES vid behov)")

    if with_dataprep:
        for fp in LOCAL_DATAPREP:
            if fp.exists():
                add("upload-local", "local:dataprep", str(fp), "input",
                    f"data_prep/{W_MAJ}/{fp.name}", fp.stat().st_size,
                    "LOCAL-DATAPREP", W_MAJ)
            else:
                say("skip", f"dataprep saknas lokalt: {fp.name}")
    return plan


# ----------------------------------------------------------------------------- exec
def _sas_and_svc():
    import blob  # noqa: E402
    from azure.storage.blob import (generate_account_sas, ResourceTypes,
                                    AccountSasPermissions)
    key = blob._read_account_key()
    sas = generate_account_sas(
        account_name=blob.STORAGE_ACCOUNT, account_key=key,
        resource_types=ResourceTypes(object=True, container=True),
        permission=AccountSasPermissions(read=True, list=True),
        expiry=datetime.now(timezone.utc) + timedelta(hours=2),
    )
    return blob, blob._client(), sas


def _ensure_container(svc, name: str) -> None:
    try:
        svc.create_container(name)
        say("SKAPAD", f"container '{name}'")
    except Exception as e:
        if "exists" not in str(e).lower():
            raise


def _copy_verified(blobmod, svc, sas, src_c, src, dst_c, dst, size) -> str:
    dst_bc = svc.get_container_client(dst_c).get_blob_client(dst)
    try:
        props = dst_bc.get_blob_properties()
        if props.size == size:
            return "SKIP-EXISTS"
    except Exception:
        pass
    src_url = f"{blobmod._ACCOUNT_URL}/{src_c}/{quote(src)}?{sas}"
    dst_bc.start_copy_from_url(src_url)
    for _ in range(240):  # samma konto -> oftast momentant; tak ~4 min
        p = dst_bc.get_blob_properties()
        st = (p.copy.status or "").lower()
        if st == "success":
            if p.size != size:
                raise RuntimeError(f"STORLEK SKILJER efter kopia: {dst} {p.size} != {size}")
            return "COPIED"
        if st in ("failed", "aborted"):
            raise RuntimeError(f"copy {st}: {dst}")
        time.sleep(1)
    raise RuntimeError(f"copy timeout: {dst}")


def _upload_verified(svc, dst_c, dst, local: Path) -> str:
    bc = svc.get_container_client(dst_c).get_blob_client(dst)
    size = local.stat().st_size
    try:
        if bc.get_blob_properties().size == size:
            return "SKIP-EXISTS"
    except Exception:
        pass
    with open(local, "rb") as fh:
        bc.upload_blob(fh, overwrite=True, max_concurrency=4)
    up = bc.get_blob_properties().size
    if up != size:
        raise RuntimeError(f"STORLEK SKILJER efter upload: {dst} {up} != {size}")
    return "UPLOADED"


def _write_manifests(svc, done: list[dict]) -> int:
    """BB.11: MANIFEST.json per (container, prefix tva niva-segment)."""
    groups: dict[tuple, list[dict]] = {}
    for d in done:
        segs = d["dst"].split("/")
        prefix = "/".join(segs[:2]) if len(segs) > 2 else (segs[0] if len(segs) > 1 else "")
        groups.setdefault((d["dst_container"], prefix), []).append(d)
    n = 0
    for (cont, prefix), items in groups.items():
        if not prefix:
            continue
        manifest = {
            "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "developer": "Jens Palmo (Senior Business Analyst, Evidensia)",
            "purpose": "BB.11 self-documenting output -- vad som ar vart att lasa, utan gissning",
            "prefix": f"{cont}/{prefix}",
            "files": [{"name": d["dst"].split("/")[-1], "blob": d["dst"],
                       "bytes": d["bytes"], "migrated_from": f"{d['src_container']}/{d['src']}",
                       "class": d["class"]} for d in sorted(items, key=lambda x: x["dst"])],
        }
        bc = svc.get_container_client(cont).get_blob_client(f"{prefix}/MANIFEST.json")
        bc.upload_blob(json.dumps(manifest, ensure_ascii=False, indent=1).encode("utf-8"),
                       overwrite=True)
        n += 1
    return n


def _local_path(row: dict) -> Path:
    if row["src_container"] == "local:dataprep":
        return Path(row["src"])
    root = LOCAL_ROOTS.get(row["src_container"])
    return (root / row["src"].replace("/", os.sep)) if root else Path(row["src"])


def write_receipt(plan: list[dict], results: dict[int, str], mode: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        path = RECEIPT_DIR / f"blob_migration_{mode}_{ts}.xlsx"
        wb = Workbook(); ws = wb.active; ws.title = "migration"
        agg: dict[str, int] = {}
        for i, p in enumerate(plan):
            agg[results.get(i, p["action"])] = agg.get(results.get(i, p["action"]), 0) + 1
        for k, v in [("Receipt", f"FD.33 blob migration ({mode})"),
                     ("Generated (UTC)", ts), ("Rows", len(plan)),
                     ("Outcome", "; ".join(f"{k}={v}" for k, v in sorted(agg.items()))),
                     ("Developer", "Jens Palmo (Senior Business Analyst, Evidensia)")]:
            ws.append([k, v]); ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.append([])
        ws.append(["RESULT", "ACTION", "SRC", "DST", "BYTES", "CLASS", "WINDOW", "NOTE"])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        for i, p in enumerate(plan):
            ws.append([results.get(i, "-"), p["action"],
                       f"{p['src_container']}/{p['src']}",
                       f"{p['dst_container']}/{p['dst']}" if p["dst"] else "",
                       p["bytes"], p["class"], p["window"], p["note"]])
        for col, w in zip("ABCDEFGH", (14, 14, 66, 60, 12, 16, 26, 44)):
            ws.column_dimensions[col].width = w
        wb.save(path)
        say("KVITTO", str(path))
    except Exception as e:
        say("WARN", f"kvitto ej skrivet ({type(e).__name__}: {e})")


def main() -> int:
    ap = argparse.ArgumentParser(description="FD.33 migrering: exekvera BLOB_MALSTRUKTUR (icke-destruktivt).")
    ap.add_argument("--commit", action="store_true")
    ap.add_argument("--quarantine", action="store_true")
    ap.add_argument("--purge", action="store_true")
    ap.add_argument("--include-bulk", action="store_true")
    ap.add_argument("--with-dataprep", action="store_true")
    args = ap.parse_args()

    if not INV_PATH.exists():
        say("ERROR", f"saknar {INV_PATH} -- kor tools\\blob_archaeology.py forst.")
        return 2
    inv = json.loads(INV_PATH.read_text(encoding="utf-8"))
    plan = build_plan(inv, args.include_bulk, args.with_dataprep)
    PLAN_PATH.write_text(json.dumps(plan, ensure_ascii=False, indent=1), encoding="utf-8")

    counts: dict[str, int] = {}
    vol: dict[str, float] = {}
    for p in plan:
        counts[p["action"]] = counts.get(p["action"], 0) + 1
        vol[p["action"]] = vol.get(p["action"], 0.0) + p["bytes"] / 1e6
    print("=" * 76)
    print(f"FD.33 MIGRERINGSPLAN  ({len(plan)} rader)  -> {PLAN_PATH}")
    for a in sorted(counts):
        print(f"  {a:<16} {counts[a]:>6} st  {vol[a]:>10.1f} MB")
    holds = [p for p in plan if p["action"].startswith("hold-")]
    for h in holds[:15]:
        print(f"    HOLD: {h['src_container']}/{h['src']}  ({h['note']})")
    if len(holds) > 15:
        print(f"    ... +{len(holds)-15} HOLD till (se plan/kvitto)")
    print("=" * 76)

    results: dict[int, str] = {}
    mode = "plan"

    if args.commit:
        mode = "commit"
        blobmod, svc, sas = _sas_and_svc()
        _ensure_container(svc, "receipts")
        done_copies: list[dict] = []
        fails = 0
        for i, p in enumerate(plan):
            try:
                if p["action"] == "copy":
                    results[i] = _copy_verified(blobmod, svc, sas,
                                                p["src_container"], p["src"],
                                                p["dst_container"], p["dst"], p["bytes"])
                    if results[i] != "SKIP-EXISTS":
                        say("COPY", f"{p['src']} -> {p['dst_container']}/{p['dst']}")
                    done_copies.append(p)
                elif p["action"] == "upload-local":
                    lp = _local_path(p)
                    results[i] = _upload_verified(svc, p["dst_container"], p["dst"], lp)
                    if results[i] != "SKIP-EXISTS":
                        say("UP", f"{lp.name} -> {p['dst_container']}/{p['dst']}")
                    done_copies.append(p)
            except Exception as e:
                results[i] = f"FAIL: {type(e).__name__}"
                say("FAIL", f"{p['src']} -> {p['dst']}: {str(e)[:160]}")
                fails += 1
        nman = _write_manifests(svc, done_copies)
        say("MANIFEST", f"{nman} MANIFEST.json skrivna (BB.11)")
        if fails:
            say("SUMMA", f"{fails} FAIL -- atgarda fore Etapp B-cutover.")

    if args.quarantine:
        mode = "quarantine" if not args.commit else mode
        blobmod, svc, sas = _sas_and_svc()
        _ensure_container(svc, "quarantine")
        for i, p in enumerate(plan):
            if p["action"] != "quarantine":
                continue
            try:
                results[i] = _copy_verified(blobmod, svc, sas,
                                            p["src_container"], p["src"],
                                            "quarantine", p["dst"], p["bytes"])
                say("QUAR", f"{p['src_container']}/{p['src']} -> quarantine/{p['dst']}")
            except Exception as e:
                results[i] = f"FAIL: {type(e).__name__}"
                say("FAIL", f"quarantine {p['src']}: {str(e)[:160]}")

    if args.purge:
        blobmod, svc, _ = _sas_and_svc()
        purged = 0
        for i, p in enumerate(plan):
            if p["action"] != "quarantine":
                continue
            qc = svc.get_container_client("quarantine").get_blob_client(p["dst"])
            try:
                if qc.get_blob_properties().size == p["bytes"]:
                    svc.get_container_client(p["src_container"]).get_blob_client(p["src"]).delete_blob()
                    results[i] = "PURGED"
                    purged += 1
                else:
                    results[i] = "PURGE-BLOCKED (karantan-storlek skiljer)"
            except Exception as e:
                results[i] = f"PURGE-BLOCKED ({type(e).__name__})"
        say("PURGE", f"{purged} original raderade (verifierad karantan-kopia kravdes for var och en).")

    write_receipt(plan, results, mode)
    print("KLART. Gamla sokvagar ORORDA (utom ev. --purge). "
          "Nasta: Etapp B cutover-commit (runners + app + _AFTER_INPUTS + dry_run gron).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
