#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bundle_csv_to_xlsx.py  --  CSV->xlsx-brygga for bundle model-data (permanent steg)
=================================================================================
Utvecklare: Jens Palmo (Senior Business Analyst, Evidensia). Forfattare: Claude.

VARFOR: model-data-creation skriver Bundle_Clinic_Data.csv (to_csv), men bundle-
modellen (regular_price.py) laser bundle_weekly_model_data_clinic_hospital.xlsx.
Inget BCG-skript konverterar CSV->xlsx (bevisat: bundle_chain_validator H2/H3).
Hittills gjordes det manuellt i Excel. Detta skript gor det programmatiskt sa
kedjan blir reproducerbar -- framtida fonster slipper manuellt steg.

FAITHFUL: ingen transformering av data. Laser CSV, skriver IDENTISKA kolumner till
EN sheet ('Sheet1') -- exakt strukturen modellen och BCG-originalet har:
  Clusters, week_starting_monday, Bundle_description, Bundle_code, Bundle_visits,
  basket_price, basket_revenue, bundle_visits_per_site, num_of_sites, FTE_Interpolated

RAM-latt: openpyxl write_only mode, rad-for-rad. Ingen Ray, ingen pandas-helladdning.

KOR PA VM (dar CSV:n bor, dar modellen laser):
    ~/bcg/cluster/.venv/bin/python ~/bundle_csv_to_xlsx.py \
        --csv  ~/bcg/bundle_dataprep/data/Bundle_Clinic_Data.csv \
        --xlsx ~/bcg/bundle/data/bundle_weekly_model_data_clinic_hospital.xlsx

ELLER lokalt (efter hamtning) -- samma argument, Windows-sokvagar.

Verifierar efter skrivning: radantal xlsx == radantal CSV (R7).
"""
from __future__ import annotations
import argparse
import csv
import sys
from pathlib import Path

EXPECTED_HEADER = [
    "Clusters", "week_starting_monday", "Bundle_description", "Bundle_code",
    "Bundle_visits", "basket_price", "basket_revenue",
    "bundle_visits_per_site", "num_of_sites", "FTE_Interpolated",
]

# Numeriska kolumner -> skriv som tal (modellen forvantar numeriskt, ej text).
NUMERIC_COLS = {
    "Bundle_visits", "basket_price", "basket_revenue",
    "bundle_visits_per_site", "num_of_sites", "FTE_Interpolated",
}


def log(tag, msg):
    print(f"[{tag}] {msg}", flush=True)


def to_num(val: str):
    """Konvertera till float/int dar mojligt; behall tomt som None, annars str."""
    if val is None or val == "":
        return None
    try:
        f = float(val)
        return int(f) if f.is_integer() else f
    except ValueError:
        return val  # lat icke-numeriskt passera oforandrat (faithful)


def main() -> int:
    ap = argparse.ArgumentParser(description="Konvertera bundle model-data CSV -> xlsx (faithful).")
    ap.add_argument("--csv", required=True, help="Bundle_Clinic_Data.csv (model-data-creation-output).")
    ap.add_argument("--xlsx", required=True, help="Mal: bundle_weekly_model_data_clinic_hospital.xlsx.")
    ap.add_argument("--sheet", default="Sheet1", help="Sheet-namn (default Sheet1, matchar BCG-original).")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        log("ERROR", "openpyxl saknas i denna miljo.")
        return 1

    csv_path = Path(args.csv)
    xlsx_path = Path(args.xlsx)
    if not csv_path.exists():
        log("ERROR", f"CSV saknas: {csv_path}")
        return 1

    log("READ", f"{csv_path}  ({csv_path.stat().st_size/1024/1024:.1f} MB)")

    # Backup av befintlig xlsx (LB.24 -- ror aldrig fryst utan backup)
    if xlsx_path.exists():
        import datetime
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = xlsx_path.with_suffix(f".pre_{stamp}.xlsx")
        xlsx_path.replace(bak)
        log("BACKUP", f"befintlig xlsx -> {bak.name}")

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title=args.sheet)

    n_rows = 0
    header_seen = None
    with open(csv_path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        for i, row in enumerate(reader):
            if i == 0:
                header_seen = row
                # verifiera header matchar forvantat (faithful-kontroll)
                if [h.strip() for h in row] != EXPECTED_HEADER:
                    log("WARN", "CSV-header avviker fran forvantat:")
                    log("WARN", f"  CSV:      {row}")
                    log("WARN", f"  Forvantat: {EXPECTED_HEADER}")
                    log("WARN", "Skriver anda (faithful), men kontrollera modellen.")
                ws.append(row)  # header som text
                continue
            # datarader: numeriska kolumner -> tal
            out = []
            for col_name, val in zip(header_seen, row):
                if col_name.strip() in NUMERIC_COLS:
                    out.append(to_num(val))
                else:
                    out.append(val if val != "" else None)
            ws.append(out)
            n_rows += 1

    log("WRITE", f"sparar {xlsx_path.name} ...")
    wb.save(xlsx_path)
    log("DONE", f"{xlsx_path}  ({xlsx_path.stat().st_size/1024/1024:.1f} MB)")

    # R7: verifiera radantal (xlsx datarader == CSV datarader)
    try:
        wb2 = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws2 = wb2.active
        xlsx_total = sum(1 for _ in ws2.iter_rows())  # inkl header
        xlsx_data = xlsx_total - 1
        log("R7", f"xlsx datarader = {xlsx_data:,}  (CSV datarader = {n_rows:,})")
        if xlsx_data == n_rows:
            log("VERDICT", "RATT -- radantal matchar, konvertering trogen.")
        else:
            log("VERDICT", f"GRANSKA -- {xlsx_data} != {n_rows}.")
    except Exception as e:
        log("R7", f"kunde ej verifiera ({e}) -- filen ar dock skriven.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
