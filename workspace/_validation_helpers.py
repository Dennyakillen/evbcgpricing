"""
_validation_helpers.py
=======================
Shared helpers for the validation suite: Excel receipt writing, formatting,
hashing, and path resolution.

Developer: Jens Palmö (Senior Business Analyst, Evidensia Djursjukvård AB)
Created:   2026-06-07

Used by all validate_*.py scripts in C:\Projekt\BCG\workspace\.
"""
import hashlib
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================ STANDARD PATHS ============================
BCG_ROOT = Path(r"C:\Projekt\BCG")
BUSINESS_ROOT = Path(r"C:\Projekt\Business_Analytics")

# Cluster model paths
CLUSTER_MODEL_ROOT = BCG_ROOT / "Pipeline" / "02. Elasticity" / "2. Product Cluster Level Models"
OUR_CSV = CLUSTER_MODEL_ROOT / "data" / "0828_Sweden_weekly_model_data_P_C.csv"

# BCG facit paths
BCG_FACIT_CSV = BUSINESS_ROOT / "bcg_inputs" / "0828_Sweden_weekly_model_data_P_C.csv"
BCG_CLUSTER_SEED_XLSX = BUSINESS_ROOT / "bcg_inputs" / "0808_Sweden_Clinic_Cluster_Mapping.xlsx"
BCG_FTE_XLSX = BUSINESS_ROOT / "bcg_inputs" / "Sweden__Interpolated_Productivity_time_date_june25.xlsx"

# Receipt directory - dated subfolder under archives (tracked in git per LF.3)
RECEIPT_ROOT = BCG_ROOT / "archives" / "validation_receipts"

# Standard windows
BCG_START = "2022-07-01"
BCG_END = "2025-06-28"


# ============================ FORMATTING ============================
def fmt_msek(amount):
    """Format amount as million SEK with 1 decimal."""
    return f"{amount/1e6:>10.1f} MSEK"


def fmt_pct(pct, decimals=2):
    """Format percentage with sign and decimals."""
    return f"{pct:>+7.{decimals}f}%"


def fmt_int(n):
    """Format integer with thousand separators."""
    return f"{n:>10,}"


def file_hash_short(path):
    """Compute short MD5 hash of a file for audit trail."""
    try:
        h = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()[:12]
    except Exception:
        return "n/a"


def now_iso():
    """Current timestamp in ISO format."""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def now_file_stamp():
    """Current timestamp suitable for file names."""
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")


def get_receipt_dir():
    """Get the dated receipt subfolder, creating it if needed."""
    today = datetime.now().strftime("%Y-%m-%d")
    dir_path = RECEIPT_ROOT / today
    dir_path.mkdir(parents=True, exist_ok=True)
    return dir_path


# ============================ EXCEL STYLING ============================
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(italic=True, color="666666")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN_BORDER = Border(
    left=Side(border_style="thin", color="CCCCCC"),
    right=Side(border_style="thin", color="CCCCCC"),
    top=Side(border_style="thin", color="CCCCCC"),
    bottom=Side(border_style="thin", color="CCCCCC"),
)


def style_header_row(ws, header_row=1):
    """Apply header styling to a row."""
    for cell in ws[header_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def autosize_columns(ws, max_width=60):
    """Autosize columns to fit content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def add_title(ws, title, subtitle=None):
    """Add a title and optional subtitle to a worksheet."""
    ws.append([title])
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws.append([subtitle])
        ws["A2"].font = SUBTITLE_FONT
    ws.append([])  # empty row


def add_status_cell(ws, cell_ref, value, pass_threshold=None, warn_threshold=None):
    """Add a colored status cell based on thresholds."""
    cell = ws[cell_ref]
    cell.value = value
    if pass_threshold is not None and value == pass_threshold:
        cell.fill = PASS_FILL
    elif warn_threshold is not None and value == warn_threshold:
        cell.fill = WARN_FILL
    else:
        cell.fill = FAIL_FILL


# ============================ RECEIPT WRITER ============================
def write_receipt(receipt_path, title, sheets):
    """
    Write a multi-sheet Excel receipt.

    Args:
        receipt_path: Path where Excel file is saved.
        title: Top-level title shown on each sheet.
        sheets: List of dicts with keys:
            - name: Sheet name (max 31 chars)
            - subtitle: Optional subtitle
            - headers: List of column headers
            - rows: List of row data (list of lists)
            - notes: Optional list of trailing notes
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_def in sheets:
        ws = wb.create_sheet(sheet_def["name"][:31])
        add_title(ws, title, sheet_def.get("subtitle"))

        # Headers
        ws.append(sheet_def["headers"])
        header_row = ws.max_row
        style_header_row(ws, header_row=header_row)

        # Data rows
        for row in sheet_def["rows"]:
            ws.append(row)

        # Notes
        notes = sheet_def.get("notes", [])
        if notes:
            ws.append([])
            for note in notes:
                ws.append([note])
                ws[ws.max_row][0].font = Font(italic=True, color="666666")

        autosize_columns(ws)

    wb.save(receipt_path)
    return receipt_path


# ============================ CONSOLE LOGGING ============================
def section(title, char="="):
    """Print a section header."""
    line = char * 78
    print(line)
    print(title)
    print(line)


def subsection(title):
    """Print a subsection header."""
    print()
    print("-" * 78)
    print(title)
    print("-" * 78)
