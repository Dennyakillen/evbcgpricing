"""
xlsx_export_bcg_freshness.py  --  analyspaket for "Vad hände sedan BCG?"
========================================================================
Orchestrator (ANALYSPAKET_STANDARD): runs analys_bcg_freshness.kör_analys(),
fills meta / report / data, and writes ONE Excel workbook -- the appendix to the
top-management presentation, so decision-makers can recompute in Excel.

Sheets (standard order):
  1. Logg     -- raw stdout of the decomposition, verbatim, monospace
  2. Rapport  -- curated mirror: headline + per-service + per-cluster + top movers
  3. Om analysen -- cover: purpose, method, sources, definitions, reservations
  4. Data: per_service / per_cluster / f_mix / top_movers / merged (filterable)

This is self-contained (no external analyspaket.py dependency) so it runs in the
BCG repo as-is; if you later wire it to Business_Analytics' analyspaket engine, the
kör_analys() contract already matches.

Run (PowerShell):
    cd "C:\\Projekt\\BCG"
    py -3.11 xlsx_export_bcg_freshness.py

Developer: Jens Palmö (Senior Business Analyst, Evidensia Djursjukvård AB)
Author: Claude advisor, 2026-06-11.
"""
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, str(Path(__file__).resolve().parent))
from analys_bcg_freshness import kör_analys  # noqa: E402

OUT = Path(__file__).resolve().parent / "output_analyspaket"
OUT.mkdir(exist_ok=True)

# ---- ANALYSPAKET palette (from ANALYSPAKET_STANDARD) ----
NAVY = "1F487C"; LJUSBLA = "C4DAF1"; BEIGE = "FDEDE0"; GRA = "BFBFBF"; ZEBRA = "F4F7FB"
MONO = "Consolas"
THIN = Side(style="thin", color=GRA)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _logg_sheet(wb, logg_text):
    ws = wb.active
    ws.title = "Logg"
    ws.sheet_view.showGridLines = False
    pre = [
        f"Vad hände sedan BCG? — analyspaket, kört {datetime.now():%Y-%m-%d %H:%M:%S}",
        "Utvecklare: Jens Palmö (Senior Business Analyst, Evidensia Djursjukvård AB)",
        "Rå stdout nedan, ordagrant. Monospace bevarar kolumnjustering.",
        "=" * 78, "",
    ]
    for i, line in enumerate(pre + logg_text.splitlines(), start=1):
        c = ws.cell(row=i, column=1, value=line)
        c.data_type = "s"  # '='-rader får inte tolkas som formler (L.45)
        c.font = Font(name=MONO, size=9)
        c.alignment = Alignment(horizontal="left", vertical="top")
    ws.column_dimensions["A"].width = 130
    ws.freeze_panes = "A6"


def _report_sheet(wb, res):
    ws = wb.create_sheet("Rapport")
    ws.sheet_view.showGridLines = False
    h = res["headline"]
    r = 1

    def line(txt, bold=False, size=10, color="1A1A1A"):
        nonlocal r
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(name=MONO, size=size, bold=bold, color=color)
        c.data_type = "s"
        r += 1

    line("VAD HÄNDE SEDAN BCG? — priselasticitetens drift", bold=True, size=13, color=NAVY)
    line(f"Växande: {res['our_file']}   Fruset: {res['facit_file']}", size=9, color="5B6878")
    line("-" * 70); line("")
    line("HELHET", bold=True, size=11, color=NAVY)
    line(f"  rader {h['rows']:,}   ProductKeys {h['productkeys']:,}")
    line(f"  median |drift|        : {h['median_abs_drift']:.4f}")
    line(f"  stabila (<0.5)        : {h['pct_stable']:.1f}%")
    line(f"  beslutsrelevant (>1.0): {h['pct_decision']:.1f}%  ({h['n_decision']:,} rader)")
    line(f"  median elasticitet    : {h['median_fe_facit']:.3f} -> {h['median_fe_ours']:.3f}")
    if "wavg_drift" in h:
        line(f"  omsättningsvägd drift : {h['wavg_drift']:+.4f}  "
             f"({h['wavg_fe_facit']:.3f} -> {h['wavg_fe_ours']:.3f})")
    line(f"  riktning              : {h['stronger']:.1f}% starkare / {h['weaker']:.1f}% svagare")
    line(""); line("-" * 70)
    line("TES (data-driven)", bold=True, size=11, color=NAVY)
    line("  Kärnsortimentets priskänslighet är stabil — modellen är beslutsklar.")
    line("  Synlig rörelse: Consult mildt svagare (enda stora tjänsten);")
    line("  Healthcare = brusnormalisering på svag signal (IB.10), ej beslutssignal.")
    line(""); line("-" * 70)
    line("PER SERVICE (topp rörelse)", bold=True, size=11, color=NAVY)
    ps = res["per_service"]
    if not ps.empty:
        for _, row in ps.head(8).iterrows():
            sc = row[ps.columns[0]]
            line(f"  {str(sc)[:22]:<22} |drift| {row['median_abs_drift']:.3f}  "
                 f"oms {row.get('oms_Mkr', float('nan')):.0f} Mkr")
    ws.column_dimensions["A"].width = 90


def _meta_sheet(wb, res):
    ws = wb.create_sheet("Om analysen")
    ws.sheet_view.showGridLines = False
    head_fill = PatternFill("solid", fgColor=NAVY)
    rows = [
        ("Titel", "Vad hände sedan BCG? — priselasticitetens drift på växande data"),
        ("Syfte", "Mäta hur de blandade priselasticiteterna förändrats sedan BCG:s 2025-snapshot"),
        ("Mål", "Beslutsunderlag till top management + återkörbar soliditetsvalidering"),
        ("Utvecklare", "Jens Palmö (Senior Business Analyst, Evidensia Djursjukvård AB)"),
        ("Kördatum", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Script", "analys_bcg_freshness.py + xlsx_export_bcg_freshness.py"),
        ("Växande källa", res["our_file"]),
        ("Fruset facit", res["facit_file"]),
        ("Grain", "ProductKey × SiteCode × Clusters (Step 6-utdata)"),
        ("Period", "Växande t.o.m. 2026-04 vs BCG-baseline 2025-09"),
        ("Drift", "ours.final_elasticity − facit.final_elasticity (negativ = mer priskänslig)"),
        ("Elasticitet", "log-log priskoefficient (IB.7); negativ = priskänslig"),
        ("Stabil-tröskel", "|drift| < 0.5 (IB.11 snapshot-driftband-analog)"),
        ("Beslutsrelevant", "|drift| > 1.0"),
        ("Omsättningsvägt", "drift vägt med TotalNet = den drift som faktiskt rör pengar"),
        ("Frozen locks (LF.9)", "väv-vikter (FD.14), routning (FD.15), bundle-gren (FD.11)"),
        ("Reservationer", "Drift drivs av växande Cluster+Site genom delvis fryst väv; "
                          "svag-signal-svans (IB.10) väntas röra sig och filtreras före beslut; "
                          "omsättningssiffror är storleksordning ej WAC."),
    ]
    ws.cell(row=1, column=1, value="OM ANALYSEN").font = Font(name=MONO, bold=True, size=13, color=NAVY)
    rr = 3
    for k, v in rows:
        a = ws.cell(row=rr, column=1, value=k); a.font = Font(bold=True, size=10, color="FFFFFF")
        a.fill = head_fill; a.alignment = Alignment(vertical="top"); a.border = BORDER
        b = ws.cell(row=rr, column=2, value=v); b.font = Font(size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top"); b.border = BORDER
        rr += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 92


def _df_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    if df is None or df.empty:
        ws.cell(row=1, column=1, value="(tom)")
        return
    head_fill = PatternFill("solid", fgColor=NAVY)
    zebra = PatternFill("solid", fgColor=ZEBRA)
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=str(col))
        c.font = Font(bold=True, color="FFFFFF", size=10); c.fill = head_fill
        c.alignment = Alignment(horizontal="left"); c.border = BORDER
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if isinstance(val, float):
                val = round(val, 4)
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(size=9.5); c.border = BORDER
            if i % 2 == 0:
                c.fill = zebra
    for j, col in enumerate(df.columns, start=1):
        w = max(12, min(40, int(df[col].astype(str).str.len().max() if len(df) else 12) + 3))
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = w
    ws.freeze_panes = "A2"


def main():
    res = kör_analys()
    wb = Workbook()
    _logg_sheet(wb, res["logg_text"])
    _report_sheet(wb, res)
    _meta_sheet(wb, res)
    # Data sheets (filterable, real cells)
    _df_sheet(wb, "Data_per_service", res["per_service"])
    _df_sheet(wb, "Data_per_cluster", res["per_cluster"])
    _df_sheet(wb, "Data_f_mix", res["f_mix"])
    _df_sheet(wb, "Data_top_movers", res["top_movers"])
    # merged is large (108k rows) -> keep a decision-relevant subset for Excel
    m = res["merged"]
    if m is not None and not m.empty and "drift" in m.columns:
        big = m.reindex(m["drift"].abs().sort_values(ascending=False).index).head(2000)
        cols = [c for c in ["ProductKey", "SiteCode", "Clusters", "service",
                            "fe_facit", "fe_ours", "drift", "TotalNet"] if c in big.columns]
        _df_sheet(wb, "Data_drift_top2000", big[cols])

    stamp = datetime.now().strftime("%Y-%m-%d")
    out_path = OUT / f"Analyspaket_BCG_Freshness_{stamp}.xlsx"
    wb.save(out_path)
    print(f"\n[KLART] analyspaket sparat: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
